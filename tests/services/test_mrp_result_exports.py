import base64
import datetime
import io

from openpyxl import load_workbook

from app.models import Item, ItemCategory, PlannedPurchase, PlannedRework, PlanningRun, Specification, Unit
from app.services.mrp_result_export import export_purchases_results_xlsx, export_rework_results_xlsx


def _mk_run(db) -> PlanningRun:
    run = PlanningRun(
        status="SUCCESS",
        started_by="test",
        horizon_days=10,
        pinned=False,
        config_version_id=None,
        config_snapshot={},
        warnings=[],
        kpi={},
        started_at=datetime.datetime.utcnow(),
        finished_at=datetime.datetime.utcnow(),
    )
    db.add(run)
    db.flush()
    return run


def _decode_workbook(data_base64: str):
    decoded = base64.b64decode(data_base64)
    return load_workbook(io.BytesIO(decoded))


def test_export_purchases_results_xlsx_groups_by_category(db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-exp-buy", unit_name="шт", short_name="шт", precision=0)
    category = ItemCategory(category_name="Группа закупки", category_ref1c="cat-buy")
    db.add_all([unit, category])
    db.flush()

    item_group = Item(
        item_code="BUY-EXP-1",
        item_name="Покупка с группой",
        item_article="BUY-ART-1",
        unit="u-exp-buy",
        category_id=category.category_id,
        status="active",
    )
    item_fallback = Item(
        item_code="BUY-EXP-2",
        item_name="Покупка без группы",
        item_article="BUY-ART-2",
        unit="u-exp-buy",
        status="active",
    )
    db.add_all([item_group, item_fallback])
    db.flush()

    run = _mk_run(db)
    db.add_all(
        [
            PlannedPurchase(
                run_id=run.run_id,
                item_id=item_group.item_id,
                requested_qty=5,
                planned_qty=5,
                qty=5,
                need_date=datetime.date(2025, 1, 10),
                order_date=datetime.date(2025, 1, 5),
                lead_time_days=5,
                bucket_date=datetime.date(2025, 1, 10),
                supplier_ref1c="supp-1",
            ),
            PlannedPurchase(
                run_id=run.run_id,
                item_id=item_fallback.item_id,
                requested_qty=2,
                planned_qty=2,
                qty=2,
                need_date=datetime.date(2025, 1, 11),
                order_date=datetime.date(2025, 1, 6),
                lead_time_days=5,
                bucket_date=datetime.date(2025, 1, 11),
                supplier_ref1c=None,
            ),
        ]
    )
    db.commit()

    result = export_purchases_results_xlsx(db=db, run_id=run.run_id)

    assert result["status"] == "ok"
    assert result["format"] == "xlsx"
    assert result["total_rows"] == 2
    assert result["total_groups"] == 2
    assert result["filename"].startswith(f"mrp_purchases_run_{run.run_id}_")

    wb = _decode_workbook(result["data_base64"])
    ws = wb.active

    assert ws.title == "Purchases"
    assert ws["A1"].value == "Товарная группа: Группа закупки"
    assert ws["A5"].value == "Товарная группа: Без товарной группы"

    headers = [ws.cell(row=2, column=idx).value for idx in range(1, 12)]
    assert headers == [
        "Наименование",
        "Артикул",
        "Поставщик",
        "Категория",
        "Количество",
        "ЕИ",
        "Дата потребности",
        "Дата заказа",
        "Срок пополнения, дн.",
        "Поставщик 1С",
        "Пометка",
    ]
    assert ws["A3"].value == "Покупка с группой"
    assert ws["C3"].value == "supp-1"
    assert ws["D3"].value == "Группа закупки"
    assert ws["E3"].value == 5
    assert ws["J3"].value == "supp-1"
    assert ws["A7"].value == "Покупка без группы"


def test_export_rework_results_xlsx_contains_component_status_and_spec(db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-exp-rw", unit_name="шт", short_name="шт", precision=0)
    category = ItemCategory(category_name="Группа переработки", category_ref1c="cat-rw")
    spec = Specification(spec_code="SPEC-EXP", spec_name="Спецификация экспорта", spec_ref1c="spec-exp")
    db.add_all([unit, category, spec])
    db.flush()

    item_partial = Item(
        item_code="RW-EXP-1",
        item_name="Переработка partial",
        item_article="RW-ART-1",
        unit="u-exp-rw",
        category_id=category.category_id,
        status="active",
    )
    item_blocked = Item(
        item_code="RW-EXP-2",
        item_name="Переработка blocked",
        item_article="RW-ART-2",
        unit="u-exp-rw",
        category_id=category.category_id,
        status="active",
    )
    db.add_all([item_partial, item_blocked])
    db.flush()

    run = _mk_run(db)
    db.add_all(
        [
            PlannedRework(
                run_id=run.run_id,
                item_id=item_partial.item_id,
                spec_id=spec.spec_id,
                requested_qty=7,
                planned_qty=4,
                qty=4,
                need_date=datetime.date(2025, 1, 10),
                order_date=datetime.date(2025, 1, 9),
                lead_time_days=1,
                bucket_date=datetime.date(2025, 1, 10),
                component_limit=4,
                component_blocked=False,
                component_partial=True,
                shortage={"planned_qty": 4},
            ),
            PlannedRework(
                run_id=run.run_id,
                item_id=item_blocked.item_id,
                spec_id=None,
                requested_qty=6,
                planned_qty=0,
                qty=0,
                need_date=datetime.date(2025, 1, 11),
                order_date=datetime.date(2025, 1, 10),
                lead_time_days=1,
                bucket_date=datetime.date(2025, 1, 11),
                component_limit=0,
                component_blocked=True,
                component_partial=False,
                shortage={"planned_qty": 0},
            ),
        ]
    )
    db.commit()

    result = export_rework_results_xlsx(db=db, run_id=run.run_id)

    assert result["status"] == "ok"
    assert result["format"] == "xlsx"
    assert result["total_rows"] == 2
    assert result["total_groups"] == 1
    assert result["filename"].startswith(f"mrp_rework_run_{run.run_id}_")

    wb = _decode_workbook(result["data_base64"])
    ws = wb.active

    assert ws.title == "Rework"
    assert ws["A1"].value == "Товарная группа: Группа переработки"

    headers = [ws.cell(row=2, column=idx).value for idx in range(1, 13)]
    assert headers == [
        "Наименование",
        "Артикул",
        "Количество",
        "Запрошено",
        "К плану",
        "ЕИ",
        "Дата потребности",
        "Дата запуска",
        "Срок пополнения, дн.",
        "Спецификация",
        "Лимит по комплектующим",
        "Статус комплектующих",
    ]
    assert ws["A3"].value == "Переработка partial"
    assert ws["J3"].value == "Спецификация экспорта"
    assert ws["L3"].value == "Частично ограничен"
    assert ws["A4"].value == "Переработка blocked"
    assert ws["L4"].value == "Заблокирован"
