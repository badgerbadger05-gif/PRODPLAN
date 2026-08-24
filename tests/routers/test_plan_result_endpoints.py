import base64
import datetime
import io
from datetime import timezone

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.database import get_db
from app.models import (
    CapacityLoad,
    Item,
    ItemCategory,
    LedgerGeneration,
    PhysicalImportBatch,
    PlannedOrder,
    PlannedOrderStage,
    PlannedPurchase,
    PlannedRework,
    PlanningRun,
    PlanningReadRow,
    PlanningReadSnapshot,
    PlanningTruthState,
    ProductionResource,
    ProductionStage,
    Specification,
    Unit,
)
from app.routers.plan import router as plan_router
from app.services.mrp_result_snapshot import build_mrp_result_snapshot


def _mk_run(db) -> PlanningRun:
    cutoff = datetime.datetime(2025, 1, 1, tzinfo=timezone.utc)
    physical = PhysicalImportBatch(
        batch_key="plan-result-endpoints-physical",
        status="completed",
        cutoff=cutoff,
        source_watermarks={},
        completed_at=cutoff,
    )
    db.add(physical)
    db.flush()
    generation = LedgerGeneration(
        generation_key="plan-result-endpoints-accepted",
        status="accepted",
        cutoff=cutoff,
        accepted_at=cutoff,
        capabilities={
            "execution_allocations": True,
            "planning_snapshots": True,
        },
        source_watermarks={},
        physical_import_batch_id=physical.id,
        algorithm_version="tests/1",
    )
    db.add(generation)
    db.flush()
    # SQLite normalizes timezone-aware DateTime values on round-trip; bind the
    # fixed run to that exact persisted cutoff rather than the Python input.
    db.expire(generation, ["cutoff"])
    persisted_cutoff = generation.cutoff
    db.add(PlanningTruthState(id=1, current_generation_id=generation.id))

    run = PlanningRun(
        status="FIXED_SNAPSHOT",
        started_by="test",
        horizon_days=10,
        pinned=False,
        config_version_id=None,
        config_snapshot={},
        warnings=[],
        kpi={},
        started_at=datetime.datetime.now(timezone.utc),
        finished_at=datetime.datetime.now(timezone.utc),
        ledger_generation_id=generation.id,
        ledger_cutoff=persisted_cutoff,
        active_freeze_version=1,
    )
    db.add(run)
    db.flush()
    return run


def _publish_result_snapshot(db, run: PlanningRun) -> PlanningReadSnapshot:
    """Build the immutable payload explicitly; HTTP GETs only read it."""
    db.flush()
    return build_mrp_result_snapshot(db, run.run_id)


def _publish_manual_purchase_snapshot(db, run: PlanningRun, rows: list[dict]) -> PlanningReadSnapshot:
    db.flush()
    snapshot = PlanningReadSnapshot(
        consumer="mrp_result",
        snapshot_key=f"run:{run.run_id}",
        ledger_generation_id=run.ledger_generation_id,
        cutoff=run.ledger_cutoff,
        truth_status="accepted",
        payload={
            "run_id": run.run_id,
            "row_counts": {"purchase": len(rows)},
            "total_qty": {
                "purchase": float(sum(float(row.get("qty") or 0) for row in rows)),
            },
        },
        published_at=datetime.datetime.now(timezone.utc),
    )
    db.add(snapshot)
    db.flush()
    for index, source_row in enumerate(rows):
        db.add(
            PlanningReadRow(
                snapshot_id=snapshot.id,
                row_key=f"purchase-{index}",
                row_kind="purchase",
                sort_key=f"2025-01-1{index}|00000000000{index}|000000000000",
                payload=source_row,
            )
        )
    db.flush()
    return snapshot


@pytest.fixture()
def client(db_session):
    app = FastAPI()
    app.include_router(plan_router, prefix="/api")

    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _decode_workbook(data_base64: str):
    decoded = base64.b64decode(data_base64)
    return load_workbook(io.BytesIO(decoded))


def test_rework_list_endpoint_returns_rows_with_shortage_fields(client, db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-api-rw", unit_name="шт", short_name="шт", precision=0)
    item = Item(
        item_code="RW-API-1",
        item_name="Rework API 1",
        item_article="RW-API-1",
        replenishment_method="Переработка",
        unit="u-api-rw",
                status="active",
    )
    spec = Specification(spec_code="SPEC-API-RW", spec_name="Spec API RW", spec_ref1c="spec-api-rw")
    db.add_all([unit, item, spec])
    db.flush()

    run = _mk_run(db)
    db.add(
        PlannedRework(
            run_id=run.run_id,
            item_id=item.item_id,
            spec_id=spec.spec_id,
            requested_qty=8.0,
            planned_qty=5.0,
            qty=5.0,
            need_date=datetime.date(2025, 1, 10),
            order_date=datetime.date(2025, 1, 8),
            lead_time_days=2,
            priority_index=None,
            bucket_date=datetime.date(2025, 1, 10),
            component_limit=5.0,
            component_blocked=False,
            component_partial=True,
            shortage={"planned_qty": 5.0, "component_limit": 5.0},
            ledger_generation_id=run.ledger_generation_id,
        )
    )
    snapshot = _publish_result_snapshot(db, run)
    db.commit()

    response = client.get(f"/api/v1/plan/results/{run.run_id}/rework")
    assert response.status_code == 200

    payload = response.json()
    assert payload["total"] == 1
    row = payload["rows"][0]
    assert row["item_name"] == "Rework API 1"
    assert row["spec_code"] == "SPEC-API-RW"
    assert row["requested_qty"] == 8.0
    assert row["planned_qty"] == 5.0
    assert row["component_partial"] is True
    assert row["shortage"]["component_limit"] == 5.0


def test_grouped_by_category_endpoints_return_group_sums_and_flags(client, db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-api-cat", unit_name="шт", short_name="шт", precision=0)
    purchase_group = ItemCategory(category_name="Группа закупки API", category_ref1c="cat-buy-api")
    rework_group = ItemCategory(category_name="Группа переработки API", category_ref1c="cat-rw-api")
    db.add_all([unit, purchase_group, rework_group])
    db.flush()

    purchase_item = Item(
        item_code="BUY-API-1",
        item_name="Buy API 1",
        item_article="BUY-API-1",
        replenishment_method="Покупка",
        unit="u-api-cat",
        category_id=purchase_group.category_id,
        # Поставщик строки закупки берётся здесь, из карточки номенклатуры.
        supplier_ref1c="supp-api-1",
        status="active",
    )
    purchase_item_fallback = Item(
        item_code="BUY-API-2",
        item_name="Buy API 2",
        item_article="BUY-API-2",
        replenishment_method="Покупка",
        unit="u-api-cat",
        status="active",
    )
    rework_item_full = Item(
        item_code="RW-API-G-1",
        item_name="RW API full",
        item_article="RW-API-G-1",
        replenishment_method="Переработка",
        unit="u-api-cat",
        category_id=rework_group.category_id,
        status="active",
    )
    rework_item_partial = Item(
        item_code="RW-API-G-2",
        item_name="RW API partial",
        item_article="RW-API-G-2",
        replenishment_method="Переработка",
        unit="u-api-cat",
        category_id=rework_group.category_id,
        status="active",
    )
    db.add_all([purchase_item, purchase_item_fallback, rework_item_full, rework_item_partial])
    db.flush()

    run = _mk_run(db)
    db.add_all(
        [
            PlannedPurchase(
                run_id=run.run_id,
                item_id=purchase_item.item_id,
                requested_qty=5,
                planned_qty=5,
                qty=5,
                need_date=datetime.date(2025, 1, 10),
                order_date=datetime.date(2025, 1, 5),
                lead_time_days=5,
                bucket_date=datetime.date(2025, 1, 10),
                ledger_generation_id=run.ledger_generation_id,
            ),
            PlannedPurchase(
                run_id=run.run_id,
                item_id=purchase_item_fallback.item_id,
                requested_qty=2,
                planned_qty=2,
                qty=2,
                need_date=datetime.date(2025, 1, 11),
                order_date=datetime.date(2025, 1, 6),
                lead_time_days=5,
                bucket_date=datetime.date(2025, 1, 11),
                supplier_ref1c=None,
                ledger_generation_id=run.ledger_generation_id,
            ),
            PlannedRework(
                run_id=run.run_id,
                item_id=rework_item_full.item_id,
                spec_id=None,
                requested_qty=5,
                planned_qty=5,
                qty=5,
                need_date=datetime.date(2025, 1, 10),
                order_date=datetime.date(2025, 1, 9),
                lead_time_days=1,
                bucket_date=datetime.date(2025, 1, 10),
                component_limit=5,
                component_blocked=False,
                component_partial=False,
                shortage={"planned_qty": 5},
                ledger_generation_id=run.ledger_generation_id,
            ),
            PlannedRework(
                run_id=run.run_id,
                item_id=rework_item_partial.item_id,
                spec_id=None,
                requested_qty=7,
                planned_qty=4,
                qty=4,
                need_date=datetime.date(2025, 1, 11),
                order_date=datetime.date(2025, 1, 10),
                lead_time_days=1,
                bucket_date=datetime.date(2025, 1, 11),
                component_limit=4,
                component_blocked=False,
                component_partial=True,
                shortage={"planned_qty": 4},
                ledger_generation_id=run.ledger_generation_id,
            ),
        ]
    )
    snapshot = _publish_result_snapshot(db, run)
    db.commit()

    purchase_response = client.get(f"/api/v1/plan/results/{run.run_id}/purchases/grouped-by-category")
    assert purchase_response.status_code == 200
    purchase_payload = purchase_response.json()
    assert purchase_payload["snapshot_id"] == snapshot.id
    assert purchase_payload["ledger_generation"] == run.ledger_generation_id
    assert purchase_payload["truth_status"] == "accepted"
    assert purchase_payload["total_groups"] == 2
    assert purchase_payload["total_orders"] == 2
    purchase_groups = {group["group_name"]: group for group in purchase_payload["groups"]}
    assert purchase_groups["Группа закупки API"]["sum_qty"] == 5.0
    assert purchase_groups["Группа закупки API"]["orders"][0]["supplier_ref1c"] == "supp-api-1"
    assert purchase_groups["Группа закупки API"]["orders"][0]["source_purchase_ids"]
    assert purchase_groups["Без товарной группы"]["sum_qty"] == 2.0

    rework_response = client.get(f"/api/v1/plan/results/{run.run_id}/rework/grouped-by-category")
    assert rework_response.status_code == 200
    rework_payload = rework_response.json()
    assert rework_payload["snapshot_id"] == snapshot.id
    assert rework_payload["ledger_generation"] == run.ledger_generation_id
    assert rework_payload["truth_status"] == "accepted"
    assert rework_payload["total_groups"] == 1
    assert rework_payload["total_orders"] == 2
    group = rework_payload["groups"][0]
    assert group["group_name"] == "Группа переработки API"
    assert group["sum_qty"] == 9.0
    assert group["sum_requested_qty"] == 12.0
    assert group["sum_planned_qty"] == 9.0
    assert group["blocked_orders"] == 0
    assert group["partial_orders"] == 1


def test_purchases_endpoint_supports_supplier_and_category_filters(
    client,
    db_session,
):
    run = _mk_run(db_session)
    purchase_1 = Item(
        item_code="PUR-API-1",
        item_name="Purchase with supplier",
        item_article="PUR-API-1",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
        category_id=None,
    )
    purchase_2 = Item(
        item_code="PUR-API-2",
        item_name="Purchase with supplier B",
        item_article="PUR-API-2",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
    )
    purchase_3 = Item(
        item_code="PUR-API-3",
        item_name="Purchase no supplier",
        item_article="PUR-API-3",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
    )
    category_a = ItemCategory(
        category_name="Категория A API",
        category_ref1c="cat-a-api",
    )
    category_b = ItemCategory(
        category_name="Категория B API",
        category_ref1c="cat-b-api",
    )
    db_session.add_all([purchase_1, purchase_2, purchase_3, category_a, category_b])
    db_session.flush()

    purchase_4 = Item(
        item_code="PUR-API-4",
        item_name="Purchase category A",
        item_article="PUR-API-4",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
        category_id=category_a.category_id,
    )
    purchase_5 = Item(
        item_code="PUR-API-5",
        item_name="Purchase category B",
        item_article="PUR-API-5",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
        category_id=category_b.category_id,
    )
    db_session.add_all(
        [
            purchase_4,
            purchase_5,
        ]
    )
    db_session.flush()
    _publish_manual_purchase_snapshot(
        db_session,
        run,
        [
            {
                "item_id": purchase_1.item_id,
                "item_name": "Purchase with supplier",
                "supplier_ref1c": "supp-1",
                "supplier_name": "Поставщик",
                "qty": 1,
            },
            {
                "item_id": purchase_2.item_id,
                "item_name": "Purchase with supplier B",
                "supplier_ref1c": "supp-2",
                "supplier_name": "Поставщик 2",
                "qty": 2,
            },
            {
                "item_id": purchase_3.item_id,
                "item_name": "Purchase no supplier",
                "supplier_ref1c": None,
                "supplier_name": "",
                "qty": 3,
            },
            {
                "item_id": purchase_4.item_id,
                "item_name": "Purchase category A",
                "supplier_ref1c": "supp-a-category",
                "supplier_name": "Кат",
                "category_id": category_a.category_id,
                "category_ref1c": "cat-a-api",
                "qty": 4,
            },
            {
                "item_id": purchase_5.item_id,
                "item_name": "Purchase category B",
                "supplier_ref1c": "supp-b-category",
                "supplier_name": "Кат",
                "category_id": category_b.category_id,
                "category_ref1c": "cat-b-api",
                "qty": 5,
            },
        ],
    )
    db_session.commit()

    filtered = client.get(
        f"/api/v1/plan/results/{run.run_id}/purchases?supplier_ref1c=supp-2"
    )
    assert filtered.status_code == 200
    payload = filtered.json()
    assert payload["total"] == 1
    assert payload["rows"][0]["supplier_ref1c"] == "supp-2"
    assert payload["rows"][0]["item_name"] == "Purchase with supplier B"

    filtered_by_category = client.get(
        f"/api/v1/plan/results/{run.run_id}/purchases?category_id={category_a.category_id}"
    )
    assert filtered_by_category.status_code == 200
    payload = filtered_by_category.json()
    assert payload["total"] == 1
    assert payload["rows"][0]["item_name"] == "Purchase category A"

    filtered_by_ref = client.get(
        f"/api/v1/plan/results/{run.run_id}/purchases?category_ref1c=cat-b-api"
    )
    assert filtered_by_ref.status_code == 200
    payload = filtered_by_ref.json()
    assert payload["total"] == 1
    assert payload["rows"][0]["item_name"] == "Purchase category B"

    missing_supplier = client.get(
        f"/api/v1/plan/results/{run.run_id}/purchases?supplier_ref1c=__missing_supplier_name"
    )
    assert missing_supplier.status_code == 200
    payload = missing_supplier.json()
    assert payload["total"] == 1
    assert payload["rows"][0]["item_name"] == "Purchase no supplier"


def test_purchases_endpoint_supports_missing_category_filter(
    client,
    db_session,
):
    run = _mk_run(db_session)
    missing_item = Item(
        item_code="PUR-MISS-1",
        item_name="Purchase missing category",
        item_article="PUR-MISS-1",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
    )
    category_item = Item(
        item_code="PUR-MISS-2",
        item_name="Purchase with category",
        item_article="PUR-MISS-2",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
    )
    db_session.add_all([missing_item, category_item])
    category = ItemCategory(category_name="Категория для контроля", category_ref1c="cat-keep")
    db_session.add(category)
    db_session.flush()
    category_row = Item(
        item_code="PUR-MISS-3",
        item_name="Purchase category",
        item_article="PUR-MISS-3",
        replenishment_method="Покупка",
        unit="шт",
        status="active",
        category_id=category.category_id,
    )
    db_session.add(category_row)
    db_session.flush()

    _publish_manual_purchase_snapshot(
        db_session,
        run,
        [
            {
                "item_id": missing_item.item_id,
                "item_name": "Purchase missing category",
                "supplier_ref1c": None,
                "supplier_name": "",
                "qty": 1,
            },
            {
                "item_id": category_item.item_id,
                "item_name": "Purchase with category",
                "supplier_ref1c": "supp-x",
                "supplier_name": "X",
                "category_id": category.category_id,
                "category_ref1c": "cat-keep",
                "qty": 2,
            },
            {
                "item_id": category_row.item_id,
                "item_name": "Purchase category",
                "supplier_ref1c": "supp-y",
                "supplier_name": "Y",
                "category_id": category.category_id,
                "category_ref1c": "cat-keep",
                "qty": 3,
            },
        ],
    )
    db_session.flush()
    db_session.commit()

    missing_category_response = client.get(
        f"/api/v1/plan/results/{run.run_id}/purchases?category_ref1c=__missing_category"
    )
    assert missing_category_response.status_code == 200
    payload = missing_category_response.json()
    assert payload["total"] == 1
    assert payload["rows"][0]["item_name"] == "Purchase missing category"


def test_export_endpoints_return_xlsx_payloads(client, db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-api-exp", unit_name="шт", short_name="шт", precision=0)
    category = ItemCategory(category_name="Группа экспорта API", category_ref1c="cat-exp-api")
    spec = Specification(spec_code="SPEC-EXP-API", spec_name="Спецификация API", spec_ref1c="spec-exp-api")
    db.add_all([unit, category, spec])
    db.flush()

    purchase_item = Item(
        item_code="BUY-EXP-API",
        item_name="Покупка API",
        item_article="BUY-EXP-API",
        replenishment_method="Покупка",
        unit="u-api-exp",
        category_id=category.category_id,
        status="active",
    )
    rework_item = Item(
        item_code="RW-EXP-API",
        item_name="Переработка API",
        item_article="RW-EXP-API",
        replenishment_method="Переработка",
        unit="u-api-exp",
        category_id=category.category_id,
        status="active",
    )
    db.add_all([purchase_item, rework_item])
    db.flush()

    run = _mk_run(db)
    db.add(
        PlannedPurchase(
            run_id=run.run_id,
            item_id=purchase_item.item_id,
            requested_qty=3,
            planned_qty=3,
            qty=3,
            need_date=datetime.date(2025, 1, 10),
            order_date=datetime.date(2025, 1, 5),
            lead_time_days=5,
            bucket_date=datetime.date(2025, 1, 10),
            supplier_ref1c="supp-exp-api",
            ledger_generation_id=run.ledger_generation_id,
        )
    )
    db.add(
        PlannedRework(
            run_id=run.run_id,
            item_id=rework_item.item_id,
            spec_id=spec.spec_id,
            requested_qty=6,
            planned_qty=4,
            qty=4,
            need_date=datetime.date(2025, 1, 11),
            order_date=datetime.date(2025, 1, 10),
            lead_time_days=1,
            bucket_date=datetime.date(2025, 1, 11),
            component_limit=4,
            component_blocked=False,
            component_partial=True,
            shortage={"planned_qty": 4},
            ledger_generation_id=run.ledger_generation_id,
        )
    )
    snapshot = _publish_result_snapshot(db, run)
    db.commit()

    purchase_response = client.get(f"/api/v1/plan/results/{run.run_id}/purchases/export?format=xlsx")
    assert purchase_response.status_code == 200
    purchase_payload = purchase_response.json()
    assert purchase_payload["format"] == "xlsx"
    assert purchase_payload["filename"].startswith(f"mrp_purchases_run_{run.run_id}_")
    purchase_wb = _decode_workbook(purchase_payload["data_base64"])
    assert purchase_wb.active.title == "Purchases"
    assert purchase_wb.active["A1"].value == "Товарная группа: Группа экспорта API"
    assert purchase_wb.active["A3"].value == "Покупка API"

    rework_response = client.get(f"/api/v1/plan/results/{run.run_id}/rework/export?format=xlsx")
    assert rework_response.status_code == 200
    rework_payload = rework_response.json()
    assert rework_payload["format"] == "xlsx"
    assert rework_payload["filename"].startswith(f"mrp_rework_run_{run.run_id}_")
    rework_wb = _decode_workbook(rework_payload["data_base64"])
    assert rework_wb.active.title == "Rework"
    assert rework_wb.active["A1"].value == "Товарная группа: Группа экспорта API"
    assert rework_wb.active["A3"].value == "Переработка API"
    assert rework_wb.active["J3"].value == "Спецификация API"
    assert rework_wb.active["L3"].value == "Частично ограничен"


def test_production_result_endpoints_keep_grouping_flags_and_export_contract(client, db_session):
    db = db_session

    unit = Unit(unit_ref1c="u-api-prod", unit_name="шт", short_name="шт", precision=0)
    stage = ProductionStage(stage_name="Сборка", stage_order=1, stage_ref1c="stage-prod-api")
    area = ProductionResource(
        resource_name="Участок API",
        shift_offset=0,
        planning_range=30,
        capacity=8.0,
        work_schedule="5/2",
        daily_work_hours=8.0,
        buffer_days=0,
    )
    item = Item(
        item_code="PROD-API-1",
        item_name="Production API 1",
        item_article="PROD-API-1",
        replenishment_method="Производство",
        unit="u-api-prod",
                status="active",
    )
    db.add_all([unit, stage, area, item])
    db.flush()

    run = _mk_run(db)
    order = PlannedOrder(
        run_id=run.run_id,
        item_id=item.item_id,
        requested_qty=10,
        planned_qty=6,
        qty=6,
        need_date=datetime.date(2025, 1, 10),
        start_date=datetime.date(2025, 1, 8),
        finish_date=datetime.date(2025, 1, 12),
        route_ref="route-api-prod",
        priority_index=1.5,
        bucket_date=datetime.date(2025, 1, 8),
        demand_ref="demand-api-prod",
        demand_date=datetime.date(2025, 1, 10),
        ledger_generation_id=run.ledger_generation_id,
    )
    db.add(order)
    db.flush()

    db.add(
        PlannedOrderStage(
            run_id=run.run_id,
            order_id=order.order_id,
            stage_id=stage.stage_id,
            area_id=area.resource_id,
            bucket_date=datetime.date(2025, 1, 8),
            hours=12,
        )
    )
    db.add(
        CapacityLoad(
            run_id=run.run_id,
            area_id=area.resource_id,
            bucket_date=datetime.date(2025, 1, 8),
            hours_planned=12,
            hours_available=8,
            overload_hours=4,
        )
    )
    snapshot = _publish_result_snapshot(db, run)
    db.commit()

    production_response = client.get(f"/api/v1/plan/results/{run.run_id}/production")
    assert production_response.status_code == 200
    production_payload = production_response.json()
    assert production_payload["total"] == 1
    row = production_payload["rows"][0]
    assert row["item_name"] == "Production API 1"
    assert row["qty"] == 6.0
    assert row["norm_hours_total"] == 12.0
    assert row["norm_hours_per_unit"] == 2.0
    assert row["flags"]["componentPartial"] is True
    assert row["flags"]["capacityShiftDays"] == 2
    assert row["flags"]["missingArea"] is False
    assert row["flags"]["missingNorm"] is False
    assert row["stages"][0]["area_name"] == "Участок API"

    grouped_response = client.get(f"/api/v1/plan/results/{run.run_id}/production/grouped")
    assert grouped_response.status_code == 200
    grouped_payload = grouped_response.json()
    assert grouped_payload["snapshot_id"] == snapshot.id
    assert grouped_payload["ledger_generation"] == run.ledger_generation_id
    assert grouped_payload["truth_status"] == "accepted"
    assert grouped_payload["total_groups"] == 1
    assert grouped_payload["total_orders"] == 1
    group = grouped_payload["groups"][0]
    assert group["area_name"] == "Участок API"
    assert group["norm_sum_hours"] == 12.0
    assert group["cap_overload_hours"] == 4.0
    assert group["cap_overloaded_buckets"] == 1
    assert group["orders"][0]["item_name"] == "Production API 1"
    assert group["orders"][0]["norm_hours_total"] == 12.0

    export_response = client.get(f"/api/v1/plan/results/{run.run_id}/production/export?format=xlsx")
    assert export_response.status_code == 200
    export_payload = export_response.json()
    assert export_payload["format"] == "xlsx"
    assert export_payload["filename"] == f"mrp_production_run_{run.run_id}.xlsx"
    export_wb = _decode_workbook(export_payload["data_base64"])
    ws = export_wb.active
    assert ws.title == "Production"
    assert ws["A1"].value == "Участок: Участок API"
    assert ws["A3"].value == "Production API 1"
    assert ws["C3"].value == 6
    assert ws["D3"].value == 12


@pytest.mark.parametrize(
    "path",
    [
        "production/grouped",
        "purchases/grouped",
        "purchases/grouped-by-category",
        "rework/grouped",
        "rework/grouped-by-category",
        "production/export?format=xlsx",
        "purchases/export?format=xlsx",
        "rework/export?format=xlsx",
    ],
)
def test_grouped_and_export_endpoints_fail_closed_without_persisted_snapshot(
    path, client, db_session
):
    run = _mk_run(db_session)
    db_session.commit()

    response = client.get(f"/api/v1/plan/results/{run.run_id}/{path}")

    assert response.status_code == 503
    assert response.json()["detail"]["code"] == "mrp_result_snapshot_required"
