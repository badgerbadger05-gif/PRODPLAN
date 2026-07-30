import base64
import io
from datetime import datetime
from decimal import Decimal
from types import SimpleNamespace

from sqlalchemy import select

from app.models import Item, ProductionOrder, ProductionProduct
from app.services.item_ledger.production_output_cache import (
    accepted_product_output,
    accepted_product_remaining_expr,
    update_accepted_product_output_cache,
)
from app.services.production_order_export import export_production_orders_xlsx


def test_remaining_is_derived_and_never_read_from_legacy_column():
    product = SimpleNamespace(
        quantity=Decimal("10"),
        produced_qty=Decimal("4"),
        remaining_qty=Decimal("999"),
    )

    output = accepted_product_output(product)

    assert output.planned_qty == Decimal("10")
    assert output.produced_qty == Decimal("4")
    assert output.remaining_qty == Decimal("6")


def test_cache_refresh_repairs_remaining_even_when_produced_did_not_change():
    product = SimpleNamespace(
        quantity=Decimal("10"),
        produced_qty=Decimal("4"),
        remaining_qty=Decimal("0"),
    )

    changed = update_accepted_product_output_cache(
        product,
        produced_qty=Decimal("4"),
    )

    assert changed is True
    assert product.produced_qty == Decimal("4")
    assert product.remaining_qty == Decimal("6")


def test_sql_projection_ignores_corrupt_remaining_column(db_session):
    item = Item(item_code="OUT-SQL", item_name="Output SQL", status="active")
    order = ProductionOrder(
        order_number="OUT-SQL",
        order_date=datetime(2026, 7, 30),
        deletion_mark=False,
    )
    db_session.add_all([item, order])
    db_session.flush()
    product = ProductionProduct(
        order_id=order.order_id,
        item_id=item.item_id,
        quantity=Decimal("10"),
        produced_qty=Decimal("4"),
        remaining_qty=Decimal("999"),
    )
    db_session.add(product)
    db_session.flush()

    expression = accepted_product_remaining_expr(
        ProductionProduct.quantity,
        ProductionProduct.produced_qty,
    )
    value = db_session.execute(
        select(expression).where(ProductionProduct.product_id == product.product_id)
    ).scalar_one()

    assert Decimal(str(value)) == Decimal("6")


def test_xlsx_export_ignores_corrupt_remaining_column(db_session):
    from openpyxl import load_workbook

    item = Item(
        item_code="OUT-XLSX",
        item_name="Output XLSX",
        unit="шт",
        status="active",
    )
    order = ProductionOrder(
        order_number="OUT-XLSX",
        order_date=datetime(2026, 7, 30),
        deletion_mark=False,
    )
    db_session.add_all([item, order])
    db_session.flush()
    db_session.add(
        ProductionProduct(
            order_id=order.order_id,
            item_id=item.item_id,
            quantity=Decimal("10"),
            produced_qty=Decimal("4"),
            remaining_qty=Decimal("999"),
        )
    )
    db_session.commit()

    payload = export_production_orders_xlsx(db_session)
    workbook = load_workbook(
        io.BytesIO(base64.b64decode(payload["data_base64"])),
        data_only=True,
    )
    row = list(workbook.active.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    assert row[4:7] == (10.0, 4.0, 6.0)
