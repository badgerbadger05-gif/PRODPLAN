import base64
import datetime
import io

from openpyxl import load_workbook

from app.models import Item, Supplier, SupplierOrder, SupplierOrderItem
from app.services.supplier_order_export import export_supplier_orders_xlsx


def test_supplier_order_export_includes_only_orders_used_by_mrp(db_session):
    db = db_session

    item = Item(
        item_code="SUP-EXP",
        item_name="Supplier Export Item",
        item_article="SUP-EXP",
        replenishment_method="Покупка",
        stock_qty=0,
        status="active",
    )
    supplier = Supplier(supplier_ref1c="supplier-ref", supplier_name="Supplier")
    db.add_all([item, supplier])
    db.flush()

    def add_order(number, state_name, deletion_mark, remaining_qty, state_key="default"):
        order = SupplierOrder(
            order_number=number,
            order_date=datetime.datetime(2026, 5, 8),
            order_ref1c=f"{number}-ref",
            supplier_id=supplier.supplier_id,
            is_posted=True,
            order_state_key=f"{number}-state" if state_key == "default" else state_key,
            order_state_name=state_name,
            deletion_mark=deletion_mark,
        )
        db.add(order)
        db.flush()
        db.add(
            SupplierOrderItem(
                order_id=order.order_id,
                item_id_ref=item.item_id,
                line_number=1,
                quantity=10,
                received_qty=10 - remaining_qty,
                remaining_qty=remaining_qty,
                price=5,
                amount=50,
                delivery_date=datetime.datetime(2026, 5, 10),
            )
        )

    # Учитываются только фазы «в пути» / «на складе» (deny-by-default).
    add_order("ATRANSIT", "В пути", False, 4)
    add_order("BSTOCK", "Принят на склад", False, 3)
    # «В закупку» теперь фаза «Нет товара» → не экспортируется.
    add_order("PURCHASING", "В закупку", False, 5)
    # Незамапленное/пустое состояние → UNKNOWN → не экспортируется.
    add_order("UNKNOWN", None, False, 3, state_key="unknown-state")
    add_order("LEGACY", None, False, 5, state_key=None)
    add_order("NEW", "Новый заказ", False, 4)
    add_order("CANCEL", "Отменен", False, 4)
    add_order("DONE", "Завершен", False, 4)
    add_order("ACCOUNTING", "Бухгалтерия", False, 4)
    add_order("DELETED", "В пути", True, 4)
    add_order("ZERO", "В пути", False, 0)
    db.commit()

    result = export_supplier_orders_xlsx(db)

    assert result["orders_count"] == 2
    assert result["total_rows"] == 2

    wb = load_workbook(io.BytesIO(base64.b64decode(result["data_base64"])))
    ws = wb.active
    assert ws.title == "ЗаказыПоставщику"
    assert ws.auto_filter.ref
    assert ws.cell(row=2, column=1).value == "Заказ №ATRANSIT от 2026-05-08 • В пути • Supplier"
    assert ws.cell(row=3, column=1).value == "ATRANSIT"
    assert ws.cell(row=3, column=3).value == "В пути"
    assert ws.cell(row=3, column=11).value == 4
    assert ws.cell(row=5, column=1).value == "Заказ №BSTOCK от 2026-05-08 • Принят на склад • Supplier"
    assert ws.cell(row=6, column=1).value == "BSTOCK"
    assert ws.cell(row=6, column=3).value == "Принят на склад"
    assert ws.cell(row=6, column=11).value == 3
