from __future__ import annotations

from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional, Tuple, Set, DefaultDict

from sqlalchemy.orm import Session
from sqlalchemy import func, and_, asc, desc
from collections import defaultdict
import json
import re
import math
import logging
logger = logging.getLogger("prodplan.planning")

from ..models import (
    PlanningConfigVersion,
    PlanningRun,
    PlannedOrder,
    PlannedOrderStage,
    PlannedPurchase,
    CapacityLoad,
    PeggingLink,
    Item,
    Unit,
    DefaultSpecification,
    SpecComponent,
    ProductionPlanEntry,
    ProductionResource,
    ResourceStage,
    ProductionStage,
    SpecOperation,
    Operation,
    ProductionKind,
    ResourceProductionKind,
    Specification,
)
from ..models import RootProduct
from .stage_logic import determine_parent_stage_and_norm
from .order_quantity_calculator import OrderQuantityCalculator
from .priority_manager import PriorityManager
from .capacity_scheduler import CapacityScheduler
from .pegging_builder import PeggingBuilder
from .warnings import make_warning, log_warning

# Default planning config fallback (aligned with Alembic seed)
DEFAULT_PLANNING_CONFIG: Dict[str, Any] = {
    "planning_horizon_days": 90,
    "mps_daily_horizon_days": 90,
    "weekly": {"enabled": True, "anchor_day": "Monday", "need_date_day": "Friday"},
    "procurement": {
        "default_lead_time_days": 30,
        "lead_time_min_policy": "max(default_lead_time_days, lead_time_from_item)",
        "lot_sizing": {"moq_source": "item_card_or_1", "multiple": 1, "rounding": "ceil"},
        "order_date_rounding_policy": "previous_workday",
    },
    "production": {"lot_sizing": {"min_batch": 1, "multiple": 1, "rounding": "ceil"}},
    "safety_stock_percent": 1,
    "capacity": {"use_resource_calendars": True, "consider_power_coefficients": True},
    "prioritization": {
        "weight_criticality": 0.4,
        "weight_importance": 0.3,
        "weight_cycle_time": 0.3,
        "default_importance": 1,
    },
    "toggles": {"include_wip": False, "enable_weekly_route_detail": False},
}


def _ensure_dict(raw: Any) -> Dict[str, Any]:
    """
    Try to robustly convert various JSON/Mapping representations to a plain dict.
    Gracefully handles:
      - dict-like objects
      - JSON string
      - objects implementing .items()
      - list of (key, value) pairs
    Returns {} if conversion is not possible.
    """
    if isinstance(raw, dict):
        return dict(raw)
    if isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                return dict(parsed)
            else:
                return {}
        except Exception:
            return {}
    # Mapping protocol
    try:
        # This can still raise if raw is an iterable of single values (e.g. list of strings)
        return dict(raw or {})
    except Exception:
        pass
    # Try items()
    try:
        return dict(getattr(raw, "items")())
    except Exception:
        pass
    # Try sequence of pairs
    try:
        return {k: v for (k, v) in list(raw)}
    except Exception:
        pass
    return {}


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    """Recursive dict merge: override values take precedence."""
    if not isinstance(base, dict) or not isinstance(override, dict):
        return override if override is not None else base
    result = dict(base)
    for k, v in (override or {}).items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = _deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def get_active_planning_config(db: Session) -> Tuple[int, Dict[str, Any]]:
    """Fetch active planning configuration snapshot and its version id."""
    cfg: Optional[PlanningConfigVersion] = (
        db.query(PlanningConfigVersion)
        .filter(PlanningConfigVersion.is_active.is_(True))
        .order_by(PlanningConfigVersion.created_at.desc())
        .first()
    )
    if not cfg:
        raise RuntimeError("Active planning configuration is not found. Seed config first.")
    cfg_dict = _ensure_dict(getattr(cfg, "config", {}))
    return int(cfg.id), cfg_dict


def list_planning_configs(db: Session, limit: int = 50, offset: int = 0) -> Dict[str, Any]:
    """
    List planning configuration versions with pagination.
    Returns: {"rows":[{id,version,is_active,created_at,created_by,comment}], "total", "limit", "offset"}
    """
    q = (
        db.query(PlanningConfigVersion)
        .order_by(PlanningConfigVersion.created_at.desc())
        .offset(max(0, int(offset)))
        .limit(max(1, min(int(limit or 50), 200)))
    )
    rows: List[PlanningConfigVersion] = q.all()
    total = db.query(func.count(PlanningConfigVersion.id)).scalar() or 0

    result: List[Dict[str, Any]] = []
    for r in rows:
        result.append(
            {
                "id": int(r.id),
                "version": int(r.version),
                "is_active": bool(r.is_active),
                "created_at": r.created_at.isoformat() if getattr(r, "created_at", None) else None,
                "created_by": getattr(r, "created_by", None),
                "comment": getattr(r, "comment", None),
            }
        )
    return {"rows": result, "total": int(total), "limit": int(limit), "offset": int(offset)}


def create_planning_config_version(
    db: Session,
    config: Any,
    comment: Optional[str] = None,
    created_by: Optional[str] = None,
    activate: bool = False,
) -> Dict[str, Any]:
    """
    Create a new planning configuration version.
    - Computes next version = max(version)+1
    - Optionally activates the new version atomically (deactivates old)
    Returns: {"id", "version", "is_active"}
    """
    cfg_dict = _ensure_dict(config)
    if not cfg_dict:
        raise RuntimeError("Config must be a non-empty object")

    current_max_version = db.query(func.max(PlanningConfigVersion.version)).scalar() or 0
    new_version = int(current_max_version) + 1

    rec = PlanningConfigVersion(
        version=int(new_version),
        is_active=False,
        config=cfg_dict,
        comment=comment,
        created_by=created_by,
    )
    db.add(rec)
    db.flush()
    config_id = int(rec.id)

    if activate:
        # Deactivate others and activate new
        db.query(PlanningConfigVersion).filter(PlanningConfigVersion.is_active.is_(True)).update(
            {"is_active": False}
        )
        rec.is_active = True

    db.commit()
    return {"id": config_id, "version": int(rec.version), "is_active": bool(rec.is_active)}


def activate_planning_config_version(db: Session, config_id: int) -> Dict[str, Any]:
    """
    Activate specified config version (set it active, deactivate previous active).
    Returns: {"id", "version", "is_active": True}
    """
    target: Optional[PlanningConfigVersion] = (
        db.query(PlanningConfigVersion).filter(PlanningConfigVersion.id == int(config_id)).first()
    )
    if not target:
        raise RuntimeError(f"PlanningConfigVersion {config_id} not found")

    # Deactivate previous active
    db.query(PlanningConfigVersion).filter(PlanningConfigVersion.is_active.is_(True)).update(
        {"is_active": False}
    )
    target.is_active = True
    db.commit()

    return {"id": int(target.id), "version": int(target.version), "is_active": True}


def get_active_planning_config_full(db: Session) -> Dict[str, Any]:
    """
    Return full active planning configuration record including config JSON.
    Returns: {"id","version","is_active","config","comment","created_by","created_at"}
    """
    rec: Optional[PlanningConfigVersion] = (
        db.query(PlanningConfigVersion)
        .filter(PlanningConfigVersion.is_active.is_(True))
        .order_by(PlanningConfigVersion.created_at.desc())
        .first()
    )
    if not rec:
        raise RuntimeError("Active planning configuration is not found. Seed config first.")
    return {
        "id": int(rec.id),
        "version": int(rec.version),
        "is_active": True,
        "config": _ensure_dict(getattr(rec, "config", {})),
        "comment": getattr(rec, "comment", None),
        "created_by": getattr(rec, "created_by", None),
        "created_at": rec.created_at.isoformat() if getattr(rec, "created_at", None) else None,
    }


def create_planning_run(
    db: Session,
    horizon_days: Optional[int] = None,
    use_weekly: Optional[bool] = None,
    config_overrides: Optional[Dict[str, Any]] = None,
    started_by: Optional[str] = None,
) -> int:
    """
    Create planning_run row with merged config snapshot.
    Note: actual calculation is not implemented yet; this creates a run placeholder.
    """
    try:
        try:
            cfg_id, cfg = get_active_planning_config(db)
        except Exception:
            cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)
    except Exception:
        cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)

    # apply simple overrides into a snapshot (non-destructive deep merge)
    overrides: Dict[str, Any] = {}
    if horizon_days is not None:
        overrides["planning_horizon_days"] = int(horizon_days)
    if use_weekly is not None:
        overrides.setdefault("weekly", {})
        overrides["weekly"]["enabled"] = bool(use_weekly)
    if config_overrides:
        overrides = _deep_merge(overrides, config_overrides)

    try:
        snapshot = _deep_merge(cfg, overrides)
    except Exception:
        snapshot = dict(cfg) if isinstance(cfg, dict) else dict(DEFAULT_PLANNING_CONFIG)

    run = PlanningRun(
        status="SUCCESS",  # placeholder until actual calc implemented
        started_by=started_by or "api",
        horizon_days=int(snapshot.get("planning_horizon_days") or snapshot.get("mps_daily_horizon_days", 90)),
        use_weekly=bool(snapshot.get("weekly", {}).get("enabled", True)),
        config_version_id=cfg_id,
        config_snapshot=snapshot,
        warnings=[],
        kpi={},
        started_at=datetime.utcnow(),
        finished_at=datetime.utcnow(),
    )
    db.add(run)
    db.flush()
    run_id = int(run.run_id)
    db.commit()
    return run_id


def list_planning_runs(db: Session, limit: int = 50, offset: int = 0) -> Dict[str, Any]:
    q = (
        db.query(PlanningRun)
        .order_by(PlanningRun.started_at.desc())
        .offset(max(0, int(offset)))
        .limit(max(1, min(int(limit or 50), 200)))
    )
    rows: List[PlanningRun] = q.all()

    result: List[Dict[str, Any]] = []
    for r in rows:
        # lightweight metrics
        order_cnt = db.query(func.count(PlannedOrder.order_id)).filter(PlannedOrder.run_id == r.run_id).scalar() or 0
        purch_cnt = db.query(func.count(PlannedPurchase.purchase_id)).filter(PlannedPurchase.run_id == r.run_id).scalar() or 0
        overload_cnt = (
            db.query(func.count(CapacityLoad.id))
            .filter(and_(CapacityLoad.run_id == r.run_id, CapacityLoad.overload_hours > 0))
            .scalar()
            or 0
        )
        result.append(
            {
                "run_id": int(r.run_id),
                "status": r.status,
                "started_at": r.started_at.isoformat() if r.started_at else None,
                "finished_at": r.finished_at.isoformat() if r.finished_at else None,
                "horizon_days": r.horizon_days,
                "use_weekly": r.use_weekly,
                "pinned": bool(getattr(r, "pinned", False)),
                "order_count": int(order_cnt),
                "purchase_count": int(purch_cnt),
                "overload_buckets": int(overload_cnt),
            }
        )

    total = db.query(func.count(PlanningRun.run_id)).scalar() or 0
    return {"rows": result, "total": int(total), "limit": limit, "offset": offset}


def get_run_summary(db: Session, run_id: int) -> Dict[str, Any]:
    r: Optional[PlanningRun] = db.query(PlanningRun).filter(PlanningRun.run_id == run_id).first()
    if not r:
        raise RuntimeError(f"Run {run_id} not found")

    order_cnt = db.query(func.count(PlannedOrder.order_id)).filter(PlannedOrder.run_id == run_id).scalar() or 0
    purch_cnt = db.query(func.count(PlannedPurchase.purchase_id)).filter(PlannedPurchase.run_id == run_id).scalar() or 0

    cap_rows: List[CapacityLoad] = db.query(CapacityLoad).filter(CapacityLoad.run_id == run_id).all()
    overload_total = float(sum(float(x.overload_hours or 0.0) for x in cap_rows))
    overloaded_buckets = int(sum(1 for x in cap_rows if float(x.overload_hours or 0.0) > 0.0))

    return {
        "run": {
            "run_id": int(r.run_id),
            "status": r.status,
            "started_at": r.started_at.isoformat() if r.started_at else None,
            "finished_at": r.finished_at.isoformat() if r.finished_at else None,
            "horizon_days": r.horizon_days,
            "use_weekly": r.use_weekly,
            "pinned": bool(getattr(r, "pinned", False)),
        },
        "counts": {"production_orders": int(order_cnt), "purchase_requests": int(purch_cnt)},
        "capacity": {"overload_total": overload_total, "overloaded_buckets": overloaded_buckets},
        "kpi": r.kpi or {},
        "warnings": r.warnings or [],
    }


def get_run_production(
    db: Session,
    run_id: int,
    item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Возвращает производственные заказы по прогону с денормализованными полями номенклатуры,
    поддержкой сортировки и агрегатом total_qty. Также рассчитывает нормативы:
      - norm_hours_total: сумма часов по PlannedOrderStage для заказа
      - norm_hours_per_unit: norm_hours_total / qty (если qty > 0)
    Колонки для сортировки: item_name | item_article | qty | need_date | start_date | priority_index
    """
    # Запрос с join к Item для денормализации и сортировки
    q = (
        db.query(
            PlannedOrder,
            Item.item_name,
            Item.item_article,
            Item.unit,
            Unit.short_name,
            Unit.unit_name,
            Unit.unit_code,
        )
        .outerjoin(Item, PlannedOrder.item_id == Item.item_id)
        .outerjoin(Unit, Item.unit == Unit.unit_ref1c)
        .filter(PlannedOrder.run_id == run_id)
    )
    if item_id is not None:
        q = q.filter(PlannedOrder.item_id == int(item_id))
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(PlannedOrder.bucket_type == bucket_type)

    # Получаем все данные до фильтрации по датам для агрегации
    rows_joined = q.all()

    # Фильтрация по датам после получения данных для корректной агрегации
    # Важно: заказы должны включаться в диапазон, если они пересекают его (start_date <= date_to AND finish_date >= date_from)
    filtered_rows = []
    for row in rows_joined:
        po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code = row
        include_row = True
        if date_from:
            # Включаем заказ, если его дата окончания >= начала диапазона (т.е. пересекает диапазон)
            if po.finish_date is None or _to_date(date_from) > po.finish_date:
                include_row = False
        if date_to:
            # Включаем заказ, если его дата начала <= конца диапазона (т.е. пересекает диапазон)
            if po.start_date is None or po.start_date > _to_date(date_to):
                include_row = False
        if include_row:
            filtered_rows.append(row)

    # Агрегация данных по item_id, start_date и unit
    aggregated_data: Dict[Tuple[int, str, str], Dict[str, Any]] = {}
    order_ids = []
    
    for row in filtered_rows:
        po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code = row
        order_ids.append(int(po.order_id))
        
        # Ключ агрегации: item_id, start_date, unit
        unit_display = in_unit_short or in_unit_name or in_unit_code or in_unit_guid
        agg_key = (int(po.item_id), po.start_date.isoformat() if po.start_date else "", unit_display or "")
        
        if agg_key not in aggregated_data:
            aggregated_data[agg_key] = {
                "item_id": int(po.item_id),
                "item_name": in_name,
                "item_article": in_article,
                "unit": unit_display,
                "qty": 0.0,
                "need_date": po.need_date.isoformat() if po.need_date else None,
                "start_date": po.start_date.isoformat() if po.start_date else None,
                "finish_date": po.finish_date.isoformat() if po.finish_date else None,
                "route_ref": po.route_ref,
                "priority_index": float(po.priority_index or 0.0) if po.priority_index is not None else None,
                "bucket_type": po.bucket_type,
                "bucket_date": po.bucket_date.isoformat() if po.bucket_date else None,
                "demand_ref": po.demand_ref,
                "demand_date": po.demand_date.isoformat() if po.demand_date else None,
                "stages": [],
                "norm_hours_total": 0.0,
                "norm_hours_per_unit": None,
            }
        
        # Суммируем количественные значения
        aggregated_data[agg_key]["qty"] += float(po.qty or 0.0)

    # Получаем этапы по отфильтрованным заказам
    stages: List[PlannedOrderStage] = []
    if order_ids:
        stages = (
            db.query(PlannedOrderStage)
            .filter(
                PlannedOrderStage.run_id == run_id,
                PlannedOrderStage.order_id.in_(order_ids),
            )
            .all()
        )

    # Группируем этапы по заказам
    stage_by_order: Dict[int, List[Dict[str, Any]]] = {}
    for s in stages:
        stage_by_order.setdefault(int(s.order_id), []).append(
            {
                "stage_id": int(s.stage_id),
                "area_id": int(s.area_id) if s.area_id is not None else None,
                "bucket_type": s.bucket_type,
                "bucket_date": s.bucket_date.isoformat() if s.bucket_date else None,
                "hours": float(s.hours or 0.0),
            }
        )

    # Суммируем часы по агрегированным данным
    for row in filtered_rows:
        po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code = row
        
        unit_display = in_unit_short or in_unit_name or in_unit_code or in_unit_guid
        agg_key = (int(po.item_id), po.start_date.isoformat() if po.start_date else "", unit_display or "")
        
        # Добавляем этапы к агрегированной записи
        order_stages = stage_by_order.get(int(po.order_id), [])
        aggregated_data[agg_key]["stages"].extend(order_stages)
        
        # Суммируем нормо-часы
        norm_total = float(sum(float(x.get("hours") or 0.0) for x in order_stages))
        aggregated_data[agg_key]["norm_hours_total"] += norm_total

    # Fallback cache: norm-hours per unit from specification operations (used when qty==0)
    try:
        item_ids_page: List[int] = list({int(r[0].item_id) for r in filtered_rows})
    except Exception:
        item_ids_page = []
    fallback_npu: Dict[int, float] = {}
    if item_ids_page:
        try:
            # Map item -> default spec
            defs = (
                db.query(DefaultSpecification)
                .filter(DefaultSpecification.item_id.in_(item_ids_page))
                .all()
            )
            item_to_spec: Dict[int, int] = {}
            spec_ids_set: Set[int] = set()
            for d in defs:
                try:
                    iid = int(d.item_id)
                    sid = int(d.spec_id)
                    item_to_spec[iid] = sid
                    spec_ids_set.add(sid)
                except Exception:
                    continue
            if spec_ids_set:
                # Aggregate sum of COALESCE(SpecOperation.time_norm, Operation.time_norm) per spec_id
                rows = (
                    db.query(
                        SpecOperation.spec_id.label("spec_id"),
                        func.sum(func.coalesce(SpecOperation.time_norm, Operation.time_norm)).label("sum_norm"),
                    )
                    .join(Operation, SpecOperation.operation_id == Operation.operation_id)
                    .filter(SpecOperation.spec_id.in_(list(spec_ids_set)))
                    .group_by(SpecOperation.spec_id)
                    .all()
                )
                spec_norm_sum: Dict[int, float] = {int(r.spec_id): float(getattr(r, "sum_norm", 0.0) or 0.0) for r in rows}
                for iid, sid in item_to_spec.items():
                    try:
                        npu_val = float(spec_norm_sum.get(int(sid), 0.0) or 0.0)
                        if npu_val > 0.0:
                            fallback_npu[int(iid)] = npu_val
                    except Exception:
                        continue
        except Exception as ex:
            logger.exception("fallback_npu build failed: %s", ex)
            fallback_npu = {}

    # Вычисляем нормо-часы на единицу и добавляем финальные данные
    final_data: List[Dict[str, Any]] = []
    for key, data in aggregated_data.items():
        # Вычисляем нормо-часы на единицу
        if data["qty"] > 1e-12:
            data["norm_hours_per_unit"] = float(data["norm_hours_total"] / data["qty"])
        else:
            # Применяем fallback, если количество равно 0
            item_id = data["item_id"]
            npu_fb = fallback_npu.get(item_id, 0.0)
            if npu_fb > 0.0:
                data["norm_hours_per_unit"] = npu_fb

        # Генерируем order_id для агрегированной записи (составной ID для идентификации)
        data["order_id"] = hash(f"{data['item_id']}_{data['start_date']}_{data['unit']}") % (10**10)
        
        final_data.append(data)

    # Сортировка агрегированных данных
    sort_map = {
        "item_name": lambda x: x.get("item_name", ""),
        "item_article": lambda x: x.get("item_article", ""),
        "qty": lambda x: x.get("qty", 0.0),
        "need_date": lambda x: x.get("need_date", ""),
        "start_date": lambda x: x.get("start_date", ""),
        "priority_index": lambda x: x.get("priority_index", 0.0),
    }
    
    sb = (sort_by or "start_date").strip().lower()
    sd = (sort_dir or "asc").strip().lower()
    key_fn = sort_map.get(sb, lambda x: x.get("start_date", ""))
    
    # Сортировка с учетом направления
    final_data.sort(key=key_fn, reverse=(sd == "desc"))
    
    # Вычисление итогов
    total = len(final_data)
    total_qty_val = float(sum(item.get("qty", 0.0) for item in final_data))

    # Применение пагинации
    start_idx = max(0, int(offset))
    end_idx = start_idx + max(1, min(int(limit or 100), 100))
    paginated_data = final_data[start_idx:end_idx]

    return {
        "rows": paginated_data,
        "total": int(total),
        "total_qty": float(total_qty_val),
        "limit": int(limit),
        "offset": int(offset),
    }


def get_run_purchases(
    db: Session,
    run_id: int,
    item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Возвращает заявки на закупку по прогону с денормализованными полями номенклатуры,
    поддержкой сортировки и агрегатом total_qty.
    Колонки для сортировки: item_name | item_article | qty | need_date | order_date | bucket_date | priority_index
    """
    # Join к Item для денормализации и сортировки
    q = (
        db.query(
            PlannedPurchase,
            Item.item_name,
            Item.item_article,
            Item.unit,
            Unit.short_name,
            Unit.unit_name,
            Unit.unit_code,
        )
        .outerjoin(Item, PlannedPurchase.item_id == Item.item_id)
        .outerjoin(Unit, Item.unit == Unit.unit_ref1c)
        .filter(PlannedPurchase.run_id == run_id)
    )
    if item_id is not None:
        q = q.filter(PlannedPurchase.item_id == int(item_id))
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(PlannedPurchase.bucket_type == bucket_type)

    rows_joined = q.all()

    # Фильтрация по датам после получения данных для корректной агрегации
    # Важно: для закупок используем bucket_date (need_date) для определения пересечения с диапазоном
    filtered_rows = []
    for row in rows_joined:
        r, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code = row
        include_row = True
        if date_from:
            # Включаем закупку, если её дата >= начала диапазона
            if r.bucket_date is None or r.bucket_date < _to_date(date_from):
                include_row = False
        if date_to:
            # Включаем закупку, если её дата <= конца диапазона
            if r.bucket_date is None or r.bucket_date > _to_date(date_to):
                include_row = False
        if include_row:
            filtered_rows.append(row)

    # Агрегация данных по item_id и unit
    aggregated_data: Dict[Tuple[int, str], Dict[str, Any]] = {}
    
    for row in filtered_rows:
        r, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code = row
        
        unit_display = in_unit_short or in_unit_name or in_unit_code or in_unit_guid
        agg_key = (int(r.item_id), unit_display or "")
        
        if agg_key not in aggregated_data:
            aggregated_data[agg_key] = {
                "item_id": int(r.item_id),
                "item_name": in_name,
                "item_article": in_article,
                "unit": unit_display,
                "qty": 0.0,
                "need_date": r.need_date.isoformat() if r.need_date else None,
                "order_date": r.order_date.isoformat() if r.order_date else None,
                "lead_time_days": int(r.lead_time_days),
                "priority_index": float(r.priority_index or 0.0) if r.priority_index is not None else None,
                "bucket_type": r.bucket_type,
                "bucket_date": r.bucket_date.isoformat() if r.bucket_date else None,
                "supplier_ref1c": r.supplier_ref1c,
            }
        
        # Суммируем количественные значения
        aggregated_data[agg_key]["qty"] += float(r.qty or 0.0)

    # Преобразуем агрегированные данные в список
    data: List[Dict[str, Any]] = []
    for key, values in aggregated_data.items():
        # Генерируем purchase_id для агрегированной записи (составной ID для идентификации)
        values["purchase_id"] = hash(f"{values['item_id']}_{values['unit']}_{values['need_date'] or ''}") % (10**10)
        data.append(values)

    # Сортировка агрегированных данных
    sort_map = {
        "item_name": lambda x: x.get("item_name", ""),
        "item_article": lambda x: x.get("item_article", ""),
        "qty": lambda x: x.get("qty", 0.0),
        "need_date": lambda x: x.get("need_date", ""),
        "order_date": lambda x: x.get("order_date", ""),
        "bucket_date": lambda x: x.get("bucket_date", ""),
        "priority_index": lambda x: x.get("priority_index", 0.0),
    }
    
    sb = (sort_by or "bucket_date").strip().lower()
    sd = (sort_dir or "asc").strip().lower()
    key_fn = sort_map.get(sb, lambda x: x.get("bucket_date", ""))
    
    # Сортировка с учетом направления
    data.sort(key=key_fn, reverse=(sd == "desc"))

    # Вычисление итогов
    total = len(data)
    total_qty_val = float(sum(item.get("qty", 0.0) for item in data))

    # Применение пагинации
    start_idx = max(0, int(offset))
    end_idx = start_idx + max(1, min(int(limit or 10), 10))
    paginated_data = data[start_idx:end_idx]

    return {
        "rows": paginated_data,
        "total": int(total),
        "total_qty": float(total_qty_val),
        "limit": int(limit),
        "offset": int(offset),
    }


def get_run_capacity(
    db: Session,
    run_id: int,
    area_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
) -> Dict[str, Any]:
    q = db.query(CapacityLoad).filter(CapacityLoad.run_id == run_id)
    if area_id is not None:
        q = q.filter(CapacityLoad.area_id == int(area_id))
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(CapacityLoad.bucket_type == bucket_type)
    if date_from:
        q = q.filter(CapacityLoad.bucket_date >= _to_date(date_from))
    if date_to:
        q = q.filter(CapacityLoad.bucket_date <= _to_date(date_to))

    total = q.count()
    rows: List[CapacityLoad] = (
        q.order_by(CapacityLoad.bucket_date.asc(), CapacityLoad.area_id.asc())
        .offset(max(0, int(offset)))
        .limit(max(1, min(int(limit or 200), 5000)))
        .all()
    )
    data = [
        {
            "area_id": int(r.area_id),
            "bucket_type": r.bucket_type,
            "bucket_date": r.bucket_date.isoformat() if r.bucket_date else None,
            "hours_planned": float(r.hours_planned or 0.0),
            "hours_available": float(r.hours_available or 0.0),
            "overload_hours": float(r.overload_hours or 0.0),
        }
        for r in rows
    ]
    return {"rows": data, "total": int(total), "limit": int(limit), "offset": int(offset)}


def get_run_pegging(
    db: Session,
    run_id: int,
    child_item_id: Optional[int] = None,
    parent_item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
) -> Dict[str, Any]:
    q = db.query(PeggingLink).filter(PeggingLink.run_id == run_id)
    if child_item_id is not None:
        q = q.filter(PeggingLink.child_item_id == int(child_item_id))
    if parent_item_id is not None:
        q = q.filter(PeggingLink.parent_item_id == int(parent_item_id))
    if date_from:
        q = q.filter(PeggingLink.need_date >= _to_date(date_from))
    if date_to:
        q = q.filter(PeggingLink.need_date <= _to_date(date_to))

    total = q.count()
    rows: List[PeggingLink] = (
        q.order_by(PeggingLink.need_date.asc(), PeggingLink.id.asc())
        .offset(max(0, int(offset)))
        .limit(max(1, min(int(limit or 200), 5000)))
        .all()
    )
    data = [
        {
            "id": int(r.id),
            "child_item_id": int(r.child_item_id),
            "parent_item_id": int(r.parent_item_id) if r.parent_item_id is not None else None,
            "demand_ref": r.demand_ref,
            "qty_contribution": float(r.qty_contribution or 0.0),
            "need_date": r.need_date.isoformat() if r.need_date else None,
            "parent_need_date": r.parent_need_date.isoformat() if r.parent_need_date else None,
        }
        for r in rows
    ]
    return {"rows": data, "total": int(total), "limit": int(limit), "offset": int(offset)}


# --- Preview calculation (gross requirement + netting) ---

def _read_last_stock_sync_at() -> Optional[str]:
    from pathlib import Path
    p = Path("config") / "last_sync_time.json"
    if not p.exists():
        return None
    try:
        import json
        data = json.loads(p.read_text("utf-8") or "{}")
        val = str(data.get("last_sync") or "").strip()
        return val or None
    except Exception:
        return None


def _iso_week_friday(d: date) -> date:
    # ISO: Monday=1 .. Sunday=7
    wd = d.isoweekday()
    monday = d - timedelta(days=wd - 1)
    friday = monday + timedelta(days=4)
    return friday


def compute_gross_requirements(
    db: Session,
    horizon_days: Optional[int] = None,
    use_weekly: Optional[bool] = None,
    config_overrides: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Step 1: Compute gross requirements (BOM expansion) in time buckets without netting or DB writes.
    - Buckets: daily (exact dates) and weekly (Friday of ISO week) according to policy.
    - Safety stock percent is applied to gross buckets as per policy.
    Returns: { meta, config, gross, stats }
    """
    # 1) Load active config and apply overrides (mirrors create_planning_run/compute_planning_preview)
    try:
        try:
            cfg_id, cfg = get_active_planning_config(db)
        except Exception:
            cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)
    except Exception:
        cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)

    overrides: Dict[str, Any] = {}
    if horizon_days is not None:
        overrides["planning_horizon_days"] = int(horizon_days)
    if use_weekly is not None:
        overrides.setdefault("weekly", {})
        overrides["weekly"]["enabled"] = bool(use_weekly)
    if config_overrides:
        overrides = _deep_merge(overrides, config_overrides)

    try:
        snapshot = _deep_merge(cfg, overrides)
    except Exception:
        snapshot = dict(cfg) if isinstance(cfg, dict) else dict(DEFAULT_PLANNING_CONFIG)

    horizon = int(snapshot.get("planning_horizon_days") or snapshot.get("mps_daily_horizon_days", 90))
    weekly_enabled = bool(snapshot.get("weekly", {}).get("enabled", True))
    ss_percent = float(snapshot.get("safety_stock_percent", 1) or 0.0)

    d0: date = date.today()
    dmax: date = d0 + timedelta(days=max(1, horizon) - 1)

    # Limits: configurable max BOM depth
    limits_cfg = ((snapshot.get("planning") or {}).get("limits") or {})
    try:
        max_bom_depth = int(limits_cfg.get("max_bom_depth", 200) or 200)
    except Exception:
        max_bom_depth = 200

    # 2) Fetch MPS daily entries in horizon (source of gross top-level demand)
    mps_rows: List[ProductionPlanEntry] = (
        db.query(ProductionPlanEntry)
        .filter(ProductionPlanEntry.date >= d0, ProductionPlanEntry.date <= dmax)
        .all()
    )

    # Determine Dcontig = max continuous day from d0 present in MPS
    present_days: Set[date] = {
        r.date.date() for r in mps_rows if float(r.planned_qty or 0.0) > 0.0
    }
    dcontig: Optional[date] = None
    cur = d0
    while cur <= dmax and cur in present_days:
        dcontig = cur
        cur = cur + timedelta(days=1)

    # 3) Build DefaultSpecification map and lazy components cache
    defaults: List[DefaultSpecification] = db.query(DefaultSpecification).all()
    default_spec_map: Dict[int, int] = {}
    for rec in defaults:
        try:
            default_spec_map[int(rec.item_id)] = int(rec.spec_id)
        except Exception:
            continue

    components_cache: Dict[int, List[SpecComponent]] = {}

    def get_components_for_spec(spec_id: int) -> List[SpecComponent]:
        if spec_id in components_cache:
            return components_cache[spec_id]
        comps = db.query(SpecComponent).filter(SpecComponent.spec_id == spec_id).all()
        components_cache[spec_id] = comps
        return comps

    # 4) Accumulators for gross buckets
    gross_daily: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))
    gross_weekly: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))

    def add_to_bucket(item_id: int, dt: date, qty: float, is_weekly: bool) -> None:
        if qty == 0.0:
            return
        if is_weekly:
            gross_weekly[item_id][dt] += qty
        else:
            gross_daily[item_id][dt] += qty

    # Recursive expansion with cycle guard
    def expand_bom(item_id: int, qty: float, is_weekly: bool, bucket_date: date, path: Set[int], depth: int = 0) -> None:
        if qty <= 0.0:
            return
        if depth > max_bom_depth:
            return
        if item_id in path:
            return
        spec_id = default_spec_map.get(item_id)
        if not spec_id:
            return
        new_path = set(path)
        new_path.add(item_id)
        comps = get_components_for_spec(spec_id)
        if not comps:
            return
        for c in comps:
            try:
                child_id = int(c.item_id)
                comp_qty = float(c.quantity or 0.0)
            except Exception:
                continue
            child_qty = qty * comp_qty
            if child_qty <= 0.0:
                continue
            # Accumulate child demand in the same bucket
            add_to_bucket(child_id, bucket_date, child_qty, is_weekly)
            # Recurse
            expand_bom(child_id, child_qty, is_weekly, bucket_date, new_path, depth + 1)

    # 5) Fill gross buckets from MPS (top-level + BOM expansion)
    for r in mps_rows:
        root_qty = float(r.planned_qty or 0.0)
        if root_qty <= 0.0:
            continue
        day = r.date.date()
        is_week = False
        if weekly_enabled:
            # If Dcontig is None => no daily window; everything beyond d0 is weekly
            if dcontig is None:
                is_week = True
            else:
                is_week = day > dcontig
        # Resolve bucket date
        bucket_dt = _iso_week_friday(day) if is_week else day
        # Accumulate root (top-level) demand
        try:
            root_item_id = int(r.item_id)
        except Exception:
            continue
        add_to_bucket(root_item_id, bucket_dt, root_qty, is_week)
        # Expand into components within the same bucket
        expand_bom(root_item_id, root_qty, is_week, bucket_dt, path=set(), depth=0)

    # 6) Apply safety stock percentage to gross (per bucket)
    factor = 1.0 + (ss_percent / 100.0 if ss_percent else 0.0)
    if abs(factor - 1.0) > 1e-9:
        for iid, dmap in gross_daily.items():
            for dt in list(dmap.keys()):
                dmap[dt] = float(dmap[dt]) * factor
        for iid, wmap in gross_weekly.items():
            for dt in list(wmap.keys()):
                wmap[dt] = float(wmap[dt]) * factor

    # 7) Convert to serializable structures
    def serialize_bucket(bmap: DefaultDict[int, DefaultDict[date, float]]) -> Dict[str, Dict[str, float]]:
        out: Dict[str, Dict[str, float]] = {}
        for iid, dtmap in bmap.items():
            out[str(int(iid))] = {dt.isoformat(): float(q) for dt, q in sorted(dtmap.items(), key=lambda x: x[0])}
        return out

    gross_ser = {"daily": serialize_bucket(gross_daily), "weekly": serialize_bucket(gross_weekly)}
    daily_bucket_count = sum(len(v) for v in gross_ser["daily"].values())
    weekly_bucket_count = sum(len(v) for v in gross_ser["weekly"].values())

    return {
        "meta": {
            "asOf": _read_last_stock_sync_at(),
            "d0": d0.isoformat(),
            "dmax": dmax.isoformat(),
            "dcontig": dcontig.isoformat() if dcontig else None,
        },
        "config": {
            "horizon_days": horizon,
            "weekly_enabled": weekly_enabled,
            "safety_stock_percent": ss_percent,
            "config_version_id": int(cfg_id),
        },
        "gross": gross_ser,
        "stats": {
            "items": len(set(gross_daily.keys()) | set(gross_weekly.keys())),
            "daily_buckets": int(daily_bucket_count),
            "weekly_buckets": int(weekly_bucket_count),
        },
    }
def compute_planning_preview(
    db: Session,
    horizon_days: Optional[int] = None,
    use_weekly: Optional[bool] = None,
    config_overrides: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """
    Compute gross requirements (BOM expansion) and net requirements (stock netting + safety stock)
    without writing anything to DB. Output buckets: daily (exact dates) and weekly (Friday of ISO week).
    """
    # 1) Load active config and apply overrides (same logic as create_planning_run)
    try:
        try:
            cfg_id, cfg = get_active_planning_config(db)
        except Exception:
            cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)
    except Exception:
        cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)
    overrides: Dict[str, Any] = {}
    if horizon_days is not None:
        overrides["planning_horizon_days"] = int(horizon_days)
    if use_weekly is not None:
        overrides.setdefault("weekly", {})
        overrides["weekly"]["enabled"] = bool(use_weekly)
    if config_overrides:
        overrides = _deep_merge(overrides, config_overrides)
    try:
        snapshot = _deep_merge(cfg, overrides)
    except Exception:
        snapshot = dict(cfg) if isinstance(cfg, dict) else dict(DEFAULT_PLANNING_CONFIG)

    horizon = int(snapshot.get("planning_horizon_days") or snapshot.get("mps_daily_horizon_days", 90))
    weekly_enabled = bool(snapshot.get("weekly", {}).get("enabled", True))
    ss_percent = float(snapshot.get("safety_stock_percent", 1) or 0.0)

    d0: date = date.today()
    dmax: date = d0 + timedelta(days=max(1, horizon) - 1)

    # Limits: configurable max BOM depth
    limits_cfg = ((snapshot.get("planning") or {}).get("limits") or {})
    try:
        max_bom_depth = int(limits_cfg.get("max_bom_depth", 200) or 200)
    except Exception:
        max_bom_depth = 200

    # 2) Fetch MPS daily entries in horizon (source of gross top-level demand)
    mps_rows: List[ProductionPlanEntry] = (
        db.query(ProductionPlanEntry)
        .filter(ProductionPlanEntry.date >= d0, ProductionPlanEntry.date <= dmax)
        .all()
    )

    # Determine Dcontig = max continuous day from d0 present in MPS
    present_days: Set[date] = {
        r.date.date() for r in mps_rows if float(r.planned_qty or 0.0) > 0.0
    }
    dcontig: Optional[date] = None
    cur = d0
    while cur <= dmax and cur in present_days:
        dcontig = cur
        cur = cur + timedelta(days=1)

    # 3) Build DefaultSpecification map and lazy components cache
    defaults: List[DefaultSpecification] = db.query(DefaultSpecification).all()
    default_spec_map: Dict[int, int] = {}
    for rec in defaults:
        try:
            default_spec_map[int(rec.item_id)] = int(rec.spec_id)
        except Exception:
            continue

    components_cache: Dict[int, List[SpecComponent]] = {}

    def get_components_for_spec(spec_id: int) -> List[SpecComponent]:
        if spec_id in components_cache:
            return components_cache[spec_id]
        comps = db.query(SpecComponent).filter(SpecComponent.spec_id == spec_id).all()
        components_cache[spec_id] = comps
        return comps

    # 4) Accumulators for gross buckets
    gross_daily: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))
    gross_weekly: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))

    def add_to_bucket(item_id: int, dt: date, qty: float, is_weekly: bool) -> None:
        if qty == 0.0:
            return
        if is_weekly:
            gross_weekly[item_id][dt] += qty
        else:
            gross_daily[item_id][dt] += qty

    # Recursive expansion with cycle guard
    def expand_bom(item_id: int, qty: float, is_weekly: bool, bucket_date: date, path: Set[int], depth: int = 0) -> None:
        if qty <= 0.0:
            return
        if depth > max_bom_depth:
            return
        if item_id in path:
            return
        spec_id = default_spec_map.get(item_id)
        if not spec_id:
            return
        new_path = set(path)
        new_path.add(item_id)
        comps = get_components_for_spec(spec_id)
        if not comps:
            return
        for c in comps:
            try:
                child_id = int(c.item_id)
                comp_qty = float(c.quantity or 0.0)
            except Exception:
                continue
            child_qty = qty * comp_qty
            if child_qty <= 0.0:
                continue
            # Accumulate child demand in the same bucket
            add_to_bucket(child_id, bucket_date, child_qty, is_weekly)
            # Recurse
            expand_bom(child_id, child_qty, is_weekly, bucket_date, new_path, depth + 1)

    # 5) Fill gross buckets from MPS (top-level + BOM expansion)
    for r in mps_rows:
        root_qty = float(r.planned_qty or 0.0)
        if root_qty <= 0.0:
            continue
        day = r.date.date()
        is_week = False
        if weekly_enabled:
            # If Dcontig is None => no daily window; everything beyond d0 is weekly
            if dcontig is None:
                is_week = True
            else:
                is_week = day > dcontig
        # Resolve bucket date
        bucket_dt = _iso_week_friday(day) if is_week else day
        # Accumulate root (top-level) demand
        try:
            root_item_id = int(r.item_id)
        except Exception:
            continue
        add_to_bucket(root_item_id, bucket_dt, root_qty, is_week)
        # Expand into components within the same bucket
        expand_bom(root_item_id, root_qty, is_week, bucket_dt, path=set(), depth=0)

    # 6) Apply safety stock percentage to gross (per bucket)
    factor = 1.0 + (ss_percent / 100.0 if ss_percent else 0.0)
    if abs(factor - 1.0) > 1e-9:
        for iid, dmap in gross_daily.items():
            for dt in list(dmap.keys()):
                dmap[dt] = float(dmap[dt]) * factor
        for iid, wmap in gross_weekly.items():
            for dt in list(wmap.keys()):
                wmap[dt] = float(wmap[dt]) * factor

    # 6.5) Adjust child gross by parent coverage from stock/WIP (avoid phantom component demand)
    # Build coverage maps per item and bucket based on available stock + WIP using FIFO across time
    item_ids: Set[int] = set(gross_daily.keys()) | set(gross_weekly.keys())
    coverage_daily: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))
    coverage_weekly: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))
    if item_ids:
        items: List[Item] = db.query(Item).filter(Item.item_id.in_(item_ids)).all()
        stock_by_item_cov: Dict[int, float] = {int(x.item_id): float(x.stock_qty or 0.0) for x in items}
        # WIP by item
        wip_by_item_cov: Dict[int, float] = {}
        try:
            from ..models import ProductionProduct
            wip_rows_cov = db.query(ProductionProduct.item_id, func.sum(ProductionProduct.quantity)).group_by(ProductionProduct.item_id).all()
            for iid_w, qty_w in wip_rows_cov:
                try:
                    wip_by_item_cov[int(iid_w)] = float(qty_w or 0.0)
                except Exception:
                    continue
        except Exception:
            wip_by_item_cov = {}
        # Compute per-bucket coverage for each item using FIFO by date
        for iid in list(item_ids):
            try:
                remaining = float(stock_by_item_cov.get(int(iid), 0.0) or 0.0) + float(wip_by_item_cov.get(int(iid), 0.0) or 0.0)
            except Exception:
                remaining = 0.0
            buckets_i: List[Tuple[str, date, float]] = []
            for dt, q in (gross_daily.get(int(iid), {}) or {}).items():
                buckets_i.append(("daily", dt, float(q or 0.0)))
            for dt, q in (gross_weekly.get(int(iid), {}) or {}).items():
                buckets_i.append(("weekly", dt, float(q or 0.0)))
            buckets_i.sort(key=lambda x: x[1])
            for btype, dt, q in buckets_i:
                if q <= 0.0:
                    continue
                cover = min(remaining, q)
                if cover > 0.0:
                    if btype == "daily":
                        coverage_daily[int(iid)][dt] += float(cover)
                    else:
                        coverage_weekly[int(iid)][dt] += float(cover)
                    remaining -= float(cover)
                if remaining <= 1e-12:
                    remaining = 0.0
        # Reduce child gross by parent coverage proportionally to per-unit quantities
        for parent_iid, spec_id in list(default_spec_map.items()):
            try:
                pid = int(parent_iid)
            except Exception:
                continue
            comps_local = get_components_for_spec(int(spec_id)) or []
            if not comps_local:
                continue
            # Daily buckets
            for dt, cov_par in list((coverage_daily.get(pid, {}) or {}).items()):
                if cov_par <= 0.0:
                    continue
                for comp in comps_local:
                    try:
                        child_id = int(getattr(comp, "item_id"))
                        per_unit = float(getattr(comp, "quantity", 0.0) or 0.0)
                    except Exception:
                        continue
                    if per_unit <= 0.0:
                        continue
                    red = float(cov_par) * per_unit
                    cur = float(gross_daily.get(child_id, {}).get(dt, 0.0) or 0.0)
                    new_val = cur - red
                    if new_val <= 0.0:
                        if child_id in gross_daily and dt in gross_daily[child_id]:
                            try:
                                del gross_daily[child_id][dt]
                            except Exception:
                                gross_daily[child_id][dt] = 0.0
                    else:
                        gross_daily[child_id][dt] = float(new_val)
            # Weekly buckets
            for dt, cov_par in list((coverage_weekly.get(pid, {}) or {}).items()):
                if cov_par <= 0.0:
                    continue
                for comp in comps_local:
                    try:
                        child_id = int(getattr(comp, "item_id"))
                        per_unit = float(getattr(comp, "quantity", 0.0) or 0.0)
                    except Exception:
                        continue
                    if per_unit <= 0.0:
                        continue
                    red = float(cov_par) * per_unit
                    cur = float(gross_weekly.get(child_id, {}).get(dt, 0.0) or 0.0)
                    new_val = cur - red
                    if new_val <= 0.0:
                        if child_id in gross_weekly and dt in gross_weekly[child_id]:
                            try:
                                del gross_weekly[child_id][dt]
                            except Exception:
                                gross_weekly[child_id][dt] = 0.0
                    else:
                        gross_weekly[child_id][dt] = float(new_val)

    # 7) Netting with stocks (WIP ignored)
    # Collect items involved
    item_ids: Set[int] = set(gross_daily.keys()) | set(gross_weekly.keys())
    if not item_ids:
        preview = {
            "meta": {
                "asOf": _read_last_stock_sync_at(),
                "d0": d0.isoformat(),
                "dmax": dmax.isoformat(),
                "dcontig": dcontig.isoformat() if dcontig else None,
            },
            "config": {
                "horizon_days": horizon,
                "weekly_enabled": weekly_enabled,
                "safety_stock_percent": ss_percent,
                "config_version_id": int(cfg_id),
            },
            "gross": {"daily": {}, "weekly": {}},
            "net": {"daily": {}, "weekly": {}},
            "stats": {"items": 0, "daily_buckets": 0, "weekly_buckets": 0},
        }
        return preview

    items: List[Item] = db.query(Item).filter(Item.item_id.in_(item_ids)).all()
    stock_by_item: Dict[int, float] = {int(x.item_id): float(x.stock_qty or 0.0) for x in items}

    # Учитываем WIP (заказы в производстве) из production_orders
    wip_by_item: Dict[int, float] = {}
    try:
        from ..models import ProductionProduct
        wip_rows = db.query(ProductionProduct.item_id, func.sum(ProductionProduct.quantity)).group_by(ProductionProduct.item_id).all()
        for iid, qty in wip_rows:
            try:
                wip_by_item[int(iid)] = float(qty or 0.0)
            except Exception:
                continue
    except Exception:
        wip_by_item = {}

    net_daily: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))
    net_weekly: DefaultDict[int, DefaultDict[date, float]] = defaultdict(lambda: defaultdict(float))

    for iid in item_ids:
        # Чистая потребность = Валовая - Остатки - WIP
        remaining = float(stock_by_item.get(iid, 0.0) or 0.0) + float(wip_by_item.get(iid, 0.0) or 0.0)
        # Build time-ordered buckets
        buckets: List[Tuple[str, date, float]] = []
        for dt, q in gross_daily.get(iid, {}).items():
            buckets.append(("daily", dt, float(q or 0.0)))
        for dt, q in gross_weekly.get(iid, {}).items():
            buckets.append(("weekly", dt, float(q or 0.0)))
        buckets.sort(key=lambda x: x[1])  # sort by date ascending

        for btype, dt, q in buckets:
            if q <= 0.0:
                continue
            if remaining >= q:
                net_q = 0.0
                remaining -= q
            else:
                net_q = q - max(remaining, 0.0)
                remaining = 0.0
            if net_q > 0.0:
                if btype == "daily":
                    net_daily[iid][dt] += net_q
                else:
                    net_weekly[iid][dt] += net_q

    # 8) Convert to serializable structures
    def serialize_bucket(bmap: DefaultDict[int, DefaultDict[date, float]]) -> Dict[str, Dict[str, float]]:
        out: Dict[str, Dict[str, float]] = {}
        for iid, dtmap in bmap.items():
            out[str(int(iid))] = {dt.isoformat(): float(q) for dt, q in sorted(dtmap.items(), key=lambda x: x[0])}
        return out

    gross_ser = {"daily": serialize_bucket(gross_daily), "weekly": serialize_bucket(gross_weekly)}
    net_ser = {"daily": serialize_bucket(net_daily), "weekly": serialize_bucket(net_weekly)}

    daily_bucket_count = sum(len(v) for v in gross_ser["daily"].values())
    weekly_bucket_count = sum(len(v) for v in gross_ser["weekly"].values())

    return {
        "meta": {
            "asOf": _read_last_stock_sync_at(),
            "d0": d0.isoformat(),
            "dmax": dmax.isoformat(),
            "dcontig": dcontig.isoformat() if dcontig else None,
        },
        "config": {
            "horizon_days": horizon,
            "weekly_enabled": weekly_enabled,
            "safety_stock_percent": ss_percent,
            "config_version_id": int(cfg_id),
        },
        "gross": gross_ser,
        "net": net_ser,
        "stats": {
            "items": len(item_ids),
            "daily_buckets": int(daily_bucket_count),
            "weekly_buckets": int(weekly_bucket_count),
        },
    }
# === Planning run execution (saving results to DB) ===

def _to_date(s: str) -> date:
    try:
        return date.fromisoformat(str(s))
    except Exception:
        # Best effort: try first 10 chars
        return date.fromisoformat(str(s)[:10])


def _previous_workday(d: date) -> date:
    # Simplified 5/2 calendar: Sat/Sun roll back to previous Friday
    wd = d.weekday()  # Mon=0..Sun=6
    if wd == 5:  # Saturday
        return d - timedelta(days=1)
    if wd == 6:  # Sunday
        return d - timedelta(days=2)
    return d


def _classify_flow(method: Optional[str]) -> str:
    """
    Return 'purchase' or 'production' based on replenishment_method text.
    Heuristic: contains 'покуп' or 'purchase' -> purchase, else production.
    """
    m = (method or "").strip().lower()
    if "покуп" in m or "purchase" in m or "закуп" in m:
        return "purchase"
    return "production"


def _load_active_config_snapshot(
    db: Session,
    horizon_days: Optional[int],
    use_weekly: Optional[bool],
    config_overrides: Optional[Dict[str, Any]],
) -> Tuple[int, Dict[str, Any]]:
    try:
        try:
            cfg_id, cfg = get_active_planning_config(db)
        except Exception:
            cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)
    except Exception:
        cfg_id, cfg = 0, dict(DEFAULT_PLANNING_CONFIG)

    overrides: Dict[str, Any] = {}
    if horizon_days is not None:
        overrides["planning_horizon_days"] = int(horizon_days)
    if use_weekly is not None:
        overrides.setdefault("weekly", {})
        overrides["weekly"]["enabled"] = bool(use_weekly)
    if config_overrides:
        overrides = _deep_merge(overrides, config_overrides)
    try:
        snapshot = _deep_merge(cfg, overrides)
    except Exception:
        snapshot = dict(cfg) if isinstance(cfg, dict) else dict(DEFAULT_PLANNING_CONFIG)
    return int(cfg_id), snapshot


def _normalize_lot_qty(qty: float, min_qty: Optional[float], multiple: Optional[float], rounding: str) -> float:
    """
    Normalize quantity according to lot sizing rules:
      - enforce minimum quantity (min_qty/MOQ) if provided (>0)
      - enforce multiple (batch multiple) using rounding policy: ceil|floor|round
    Returns non-negative float. If result becomes 0 or negative, returns 0.0
    """
    try:
        q = float(qty or 0.0)
    except Exception:
        q = 0.0
    if q <= 0.0:
        return 0.0

    # Ensure minimum
    try:
        mn = float(min_qty) if (min_qty is not None) else None
    except Exception:
        mn = None
    if mn is not None and mn > 0.0 and q < mn:
        q = mn

    # Apply multiple with rounding policy
    try:
        m = float(multiple) if (multiple is not None) else None
    except Exception:
        m = None
    mode = (rounding or "ceil").strip().lower()
    if m is not None and m > 0.0:
        ratio = q / m
        if mode == "floor":
            q = math.floor(ratio) * m
        elif mode == "round":
            q = round(ratio) * m
        else:
            # default: ceil
            q = math.ceil(ratio) * m

        # After rounding to multiple, still ensure minimum if defined
        if mn is not None and mn > 0.0 and q < mn:
            q = math.ceil(mn / m) * m

    if not (q > 0.0):
        return 0.0
    return float(q)


def _normalize_qty_for_production(snapshot: Dict[str, Any], qty: float, item: Optional[Item] = None, buffer_qty: Optional[float] = None) -> float:
    """
    Production lot sizing with buffer days and optimal batch support.
    Priority order:
    1. optimal_batch from item (if present and > buffer_qty)
    2. buffer_qty (if provided and > min_batch)
    3. Standard lot sizing (min_batch, multiple, rounding)
    
    Supported keys in snapshot['production']['lot_sizing']:
    - min_batch, multiple, rounding (ceil|floor|round)
    
    Defaults: min_batch=1, multiple=1, rounding='ceil'
    """
    lot_cfg: Dict[str, Any] = {}
    if isinstance(snapshot, dict):
        lot_cfg = (snapshot.get("production") or {}).get("lot_sizing") or {}
    
    try:
        min_batch = float(lot_cfg.get("min_batch", 1) or 1)
    except Exception:
        min_batch = 1.0
    try:
        multiple = float(lot_cfg.get("multiple", 1) or 1)
    except Exception:
        multiple = 1.0
    rounding = str(lot_cfg.get("rounding", "ceil") or "ceil").strip().lower()
    
    # Приоритет 1: оптимальная партия из номенклатуры
    optimal_batch = None
    if item is not None:
        try:
            optimal_batch = float(getattr(item, "optimal_batch", None) or 0.0)
            if optimal_batch <= 0.0:
                optimal_batch = None
        except Exception:
            optimal_batch = None
    
    # Приоритет 2: буферное количество
    base_qty = qty
    if buffer_qty is not None and buffer_qty > 0.0:
        base_qty = max(qty, float(buffer_qty))
    
    # Применяем оптимальную партию
    if optimal_batch is not None and optimal_batch > 0.0:
        # Если оптимальная партия больше базового количества, используем её
        if optimal_batch >= base_qty:
            base_qty = optimal_batch
        else:
            # Иначе используем кратное оптимальной партии для покрытия base_qty
            ratio = base_qty / optimal_batch
            if rounding == "floor":
                base_qty = math.floor(ratio) * optimal_batch
            elif rounding == "round":
                base_qty = round(ratio) * optimal_batch
            else:  # ceil
                base_qty = math.ceil(ratio) * optimal_batch
    
    # Стандартная нормализация с min_batch и multiple
    return _normalize_lot_qty(base_qty, min_batch, multiple, rounding)


def _normalize_qty_for_procurement(snapshot: Dict[str, Any], item: Optional[Item], qty: float) -> float:
    """
    Procurement lot sizing based on snapshot['procurement']['lot_sizing'].
    Supported keys: moq, moq_source, multiple, rounding (ceil|floor|round).
    - moq_source=item_card_or_1 tries item.moq, then falls back to 1 if not present.
    - If 'moq' key is provided explicitly, it takes precedence.
    Defaults: moq_source=item_card_or_1, multiple=1, rounding='ceil'
    """
    lot_cfg: Dict[str, Any] = {}
    if isinstance(snapshot, dict):
        lot_cfg = (snapshot.get("procurement") or {}).get("lot_sizing") or {}

    # Determine minimum quantity (MOQ)
    min_qty: Optional[float] = None
    if "moq" in lot_cfg:
        try:
            min_qty = float(lot_cfg.get("moq"))
        except Exception:
            min_qty = None

    moq_source = str(lot_cfg.get("moq_source", "item_card_or_1") or "item_card_or_1").strip().lower()
    if min_qty is None:
        if moq_source == "item_card_or_1":
            # Try to read from item card if present
            try:
                item_moq = getattr(item, "moq", None)
                if item_moq is not None:
                    min_qty = float(item_moq)
            except Exception:
                pass
            if min_qty is None:
                min_qty = 1.0
        else:
            # Unknown source -> fallback to 1
            min_qty = 1.0

    try:
        multiple = float(lot_cfg.get("multiple", 1) or 1)
    except Exception:
        multiple = 1.0
    rounding = str(lot_cfg.get("rounding", "ceil") or "ceil").strip().lower()
    return _normalize_lot_qty(qty, min_qty, multiple, rounding)
def run_planning_run(
    db: Session,
    horizon_days: Optional[int] = None,
    use_weekly: Optional[bool] = None,
    config_overrides: Optional[Dict[str, Any]] = None,
    started_by: Optional[str] = None,
) -> int:
    """
    Execute planning run:
      - Create PlanningRun with status RUNNING and config snapshot
      - Compute preview (gross + net)
      - Classify flows (production vs purchase)
      - Persist planned_order / planned_purchase
      - Build PlannedOrderStage (stage detection by children, norm_hours by spec ops, map to area)
      - Aggregate CapacityLoad (hours_planned/available/overload)
      - Compute priority_index for orders and purchases
      - Record PeggingLink (one-level child→parent by bucket)
      - Finalize run with SUCCESS and KPI/warnings
    """
    # 1) Snapshot config and create RUNNING run
    cfg_id, snapshot = _load_active_config_snapshot(db, horizon_days, use_weekly, config_overrides)
    run = PlanningRun(
        status="RUNNING",
        started_by=started_by or "api",
        horizon_days=int(snapshot.get("planning_horizon_days") or snapshot.get("mps_daily_horizon_days", 90)),
        use_weekly=bool(snapshot.get("weekly", {}).get("enabled", True)),
        config_version_id=cfg_id,
        config_snapshot=snapshot,
        warnings=[],
        kpi={},
        started_at=datetime.utcnow(),
        finished_at=None,
    )
    db.add(run)
    db.flush()
    run_id = int(run.run_id)

    warnings: List[Dict[str, Any]] = []

    try:
        # 2) Compute preview to obtain gross + net by buckets
        preview = compute_planning_preview(
            db=db,
            horizon_days=horizon_days,
            use_weekly=use_weekly,
            config_overrides=config_overrides or {},
        )
        net = (preview or {}).get("net") or {}
        net_daily: Dict[str, Dict[str, float]] = net.get("daily") or {}
        net_weekly: Dict[str, Dict[str, float]] = net.get("weekly") or {}

        if not net_daily and not net_weekly:
            # Fallback: if net not produced, use gross
            gross = (preview or {}).get("gross") or {}
            net_daily = gross.get("daily") or {}
            net_weekly = gross.get("weekly") or {}
            warnings.append(log_warning(logger, "PREVIEW_NO_NET", "Net requirements missing; used gross as fallback"))

        # 3) Load items dictionary for all involved item_ids
        all_item_ids: Set[int] = set()
        for iid in net_daily.keys():
            try:
                all_item_ids.add(int(iid))
            except Exception:
                continue
        for iid in net_weekly.keys():
            try:
                all_item_ids.add(int(iid))
            except Exception:
                continue

        items: List[Item] = []
        if all_item_ids:
            items = db.query(Item).filter(Item.item_id.in_(all_item_ids)).all()
        item_by_id: Dict[int, Item] = {int(x.item_id): x for x in items}

        # 4) Persist planned purchases and orders (collect created rows)
        created_orders: List[PlannedOrder] = []
        created_purchases: List[PlannedPurchase] = []

        default_lt = int(snapshot.get("procurement", {}).get("default_lead_time_days", 30) or 30)

        def _add_purchase(iid: int, need_dt: date, q: float, bucket_type: str):
            item = item_by_id.get(iid)
            lt = default_lt
            try:
                lt = max(default_lt, int(getattr(item, "replenishment_time", 0) or 0))
            except Exception:
                lt = default_lt

            # Lot sizing for procurement
            qn = _normalize_qty_for_procurement(snapshot, item, q)
            if qn <= 0.0:
                return

            order_dt = _previous_workday(need_dt - timedelta(days=lt))
            rec = PlannedPurchase(
                run_id=run_id,
                item_id=iid,
                qty=qn,
                need_date=need_dt,
                order_date=order_dt,
                lead_time_days=int(lt),
                priority_index=None,
                bucket_type=bucket_type,
                bucket_date=need_dt,
                supplier_ref1c=None,
            )
            db.add(rec)
            created_purchases.append(rec)

        def _add_order(iid: int, need_dt: date, q: float, bucket_type: str):
            """
            Создание производственного заказа с вычислением количества через OrderQuantityCalculator:
              - буферы участков по buffer_days через среднедневную потребность
              - приоритет оптимальной партии над буфером
              - ограничение по доступности комплектующих (остатки + WIP)
              - ограничение по горизонту планирования (сумма net-потребностей)
              - нормализация по min_batch/multiple/rounding
            Важно: локализованные кэши во избежание зависимости от порядка определения переменных ниже.
            """
            try:
                # 1) Default specs and specifications
                defaults_local: List[DefaultSpecification] = db.query(DefaultSpecification).all()
                default_spec_map_local: Dict[int, int] = {}
                for rec in defaults_local:
                    try:
                        default_spec_map_local[int(rec.item_id)] = int(rec.spec_id)
                    except Exception:
                        continue

                specifications_local: List[Specification] = db.query(Specification).all()
                spec_by_id_local: Dict[int, Specification] = {
                    int(spec.spec_id): spec for spec in specifications_local if spec.spec_id
                }

                # 2) Components loader (with small cache)
                _comp_cache_local: Dict[int, List[SpecComponent]] = {}

                def _get_components_local(spec_id: int) -> List[SpecComponent]:
                    if spec_id in _comp_cache_local:
                        return _comp_cache_local[spec_id]
                    comps = db.query(SpecComponent).filter(SpecComponent.spec_id == spec_id).all()
                    _comp_cache_local[spec_id] = comps
                    return comps

                # 3) Resources and production kind mapping
                resources_local: List[ProductionResource] = db.query(ProductionResource).all()
                res_by_id_local: Dict[int, ProductionResource] = {
                    int(r.resource_id): r for r in resources_local
                }
                rpk_rows = db.query(ResourceProductionKind).all()
                production_kinds_by_resource_local: Dict[int, Set[int]] = {}
                for rpk in rpk_rows:
                    production_kinds_by_resource_local.setdefault(
                        int(rpk.resource_id), set()
                    ).add(int(rpk.production_kind_id))

                # 4) Stocks and WIP
                stock_by_item_local: Dict[int, float] = {
                    int(x.item_id): float(getattr(x, "stock_qty", 0.0) or 0.0) for x in item_by_id.values()
                }
                wip_by_item_local: Dict[int, float] = {}
                try:
                    from ..models import ProductionProduct
                    wip_rows_local = (
                        db.query(ProductionProduct.item_id, func.sum(ProductionProduct.quantity))
                        .group_by(ProductionProduct.item_id)
                        .all()
                    )
                    for iid_w, qty_w in wip_rows_local:
                        try:
                            wip_by_item_local[int(iid_w)] = float(qty_w or 0.0)
                        except Exception:
                            continue
                except Exception:
                    wip_by_item_local = {}

                # 5) Total horizon demand by item (sum of net buckets)
                total_demand_by_item_local: Dict[int, float] = {}
                for iid_str2, buckets2 in (net_daily or {}).items():
                    try:
                        ik = int(iid_str2)
                    except Exception:
                        continue
                    total_demand_by_item_local[ik] = total_demand_by_item_local.get(ik, 0.0) + sum(
                        float(v or 0.0) for v in (buckets2 or {}).values()
                    )
                for iid_str2, buckets2 in (net_weekly or {}).items():
                    try:
                        ik = int(iid_str2)
                    except Exception:
                        continue
                    total_demand_by_item_local[ik] = total_demand_by_item_local.get(ik, 0.0) + sum(
                        float(v or 0.0) for v in (buckets2 or {}).values()
                    )

                horizon_days_local = int(
                    snapshot.get("planning_horizon_days") or snapshot.get("mps_daily_horizon_days", 90)
                )

                # 6) Compute quantities via calculator
                oqc = OrderQuantityCalculator(
                    snapshot=snapshot,
                    default_spec_map=default_spec_map_local,
                    spec_by_id=spec_by_id_local,
                    components_loader=_get_components_local,
                    item_by_id=item_by_id,
                    res_by_id=res_by_id_local,
                    production_kinds_by_resource=production_kinds_by_resource_local,
                    stock_by_item=stock_by_item_local,
                    wip_by_item=wip_by_item_local,
                    horizon_days=horizon_days_local,
                    total_demand_by_item=total_demand_by_item_local,
                )
                final_before, qn, comp_details, warn_list = oqc.compute(int(iid), float(q))
                if warn_list:
                    # Enrich component shortage warnings with parent item context and shortage amounts
                    enriched: List[Dict[str, Any]] = []
                    for w in (warn_list or []):
                        try:
                            if str(w.get("code") or "") == "COMPONENT_SHORTAGE":
                                ww = dict(w)
                                ww["item_id"] = int(iid)
                                itp = item_by_id.get(int(iid))
                                ww["item_name"] = getattr(itp, "item_name", None) if itp else None
                                ww["item_article"] = getattr(itp, "item_article", None) if itp else None
                                ww["need_date"] = order.need_date.isoformat() if order.need_date else None
                                # Parent-level shortage derived from requested vs component-limited possible
                                try:
                                    rq = float(ww.get("requested_qty", q))
                                except Exception:
                                    rq = float(q or 0.0)
                                try:
                                    mp = float(ww.get("max_producible_from_component", 0.0) or 0.0)
                                except Exception:
                                    mp = 0.0
                                ww["shortage_parent_qty"] = float(max(0.0, rq - mp))
                                enriched.append(ww)
                            else:
                                enriched.append(w)
                        except Exception:
                            enriched.append(w)
                    warnings.extend(enriched)
                if qn <= 0.0:
                    return
            except Exception:
                # Fail-safe fallback: place requested qty as-is if calculator failed
                qn = float(q or 0.0)
                if qn <= 0.0:
                    return

            # Enforce integer quantities for production orders (ceil)
            try:
                qn_int = math.ceil(float(qn))
            except Exception:
                qn_int = int(qn) if qn is not None else 0
            if qn_int <= 0:
                return

            rec = PlannedOrder(
                run_id=run_id,
                item_id=iid,
                qty=float(qn_int),
                need_date=need_dt,
                start_date=None,
                finish_date=None,
                route_ref=None,
                priority_index=None,
                bucket_type=bucket_type,
                bucket_date=need_dt,
                demand_ref=None,
                demand_date=need_dt,
            )
            db.add(rec)
            created_orders.append(rec)
            try:
                # keep original computation details and add rounded qty for diagnostics
                if isinstance(comp_details, dict):
                    comp_details = dict(comp_details)
                    comp_details["normalized_qty_rounded"] = float(qn_int)
                setattr(rec, "_comp_details", comp_details)
            except Exception:
                pass

        # Daily buckets
        for iid_str, buckets in net_daily.items():
            try:
                iid = int(iid_str)
            except Exception:
                continue
            item = item_by_id.get(iid)
            flow = _classify_flow(getattr(item, "replenishment_method", None)) if item else "production"

            for dt_str, qty in (buckets or {}).items():
                try:
                    q = float(qty or 0.0)
                except Exception:
                    q = 0.0
                if q <= 0.0:
                    continue
                need_dt = _to_date(dt_str)
                if flow == "purchase":
                    _add_purchase(iid, need_dt, q, "daily")
                else:
                    _add_order(iid, need_dt, q, "daily")

        # Weekly buckets
        for iid_str, buckets in net_weekly.items():
            try:
                iid = int(iid_str)
            except Exception:
                continue
            item = item_by_id.get(iid)
            flow = _classify_flow(getattr(item, "replenishment_method", None)) if item else "production"

            for dt_str, qty in (buckets or {}).items():
                try:
                    q = float(qty or 0.0)
                except Exception:
                    q = 0.0
                if q <= 0.0:
                    continue
                need_dt = _to_date(dt_str)
                if flow == "purchase":
                    _add_purchase(iid, need_dt, q, "weekly")
                else:
                    _add_order(iid, need_dt, q, "weekly")

        # Flush to get primary keys for orders/purchases
        db.flush()

        # 5) Build stage detection and norm caches based on default specs
        defaults: List[DefaultSpecification] = db.query(DefaultSpecification).all()
        default_spec_map: Dict[int, int] = {}
        for rec in defaults:
            try:
                default_spec_map[int(rec.item_id)] = int(rec.spec_id)
            except Exception:
                continue

        # Load specifications to get production kind info
        specifications: List[Specification] = db.query(Specification).all()
        spec_by_id: Dict[int, Specification] = {int(spec.spec_id): spec for spec in specifications if spec.spec_id}

        components_cache: Dict[int, List[SpecComponent]] = {}
        operations_cache: Dict[int, List[SpecOperation]] = {}

        def get_components_for_spec(spec_id: int) -> List[SpecComponent]:
            if spec_id in components_cache:
                return components_cache[spec_id]
            comps = db.query(SpecComponent).filter(SpecComponent.spec_id == spec_id).all()
            components_cache[spec_id] = comps
            return comps

        def get_operations_for_spec(spec_id: int) -> List[SpecOperation]:
            if spec_id in operations_cache:
                return operations_cache[spec_id]
            ops = db.query(SpecOperation).filter(SpecOperation.spec_id == spec_id).all()
            operations_cache[spec_id] = ops
            return ops

        # --- Roots and reachability helpers for diagnostics ---
        root_rows: List[RootProduct] = db.query(RootProduct).all()
        root_item_ids: List[int] = [
            int(r.item_id) for r in root_rows
            if getattr(r, "item_id", None) is not None
        ]

        # Cache (root_id, target_id) -> bool (reachability)
        _contains_cache: Dict[Tuple[int, int], bool] = {}

        def _child_items_of(item_id: int) -> List[int]:
            spec_id_local = default_spec_map.get(int(item_id))
            if not spec_id_local:
                return []
            comps_local = get_components_for_spec(int(spec_id_local)) or []
            out: List[int] = []
            for c in comps_local:
                try:
                    out.append(int(c.item_id))
                except Exception:
                    continue
            return out

        def _contains_target(root_id: int, target_id: int, depth: int = 0, seen: Optional[Set[int]] = None) -> bool:
            key = (int(root_id), int(target_id))
            if key in _contains_cache:
                return _contains_cache[key]
            if depth > 200:
                _contains_cache[key] = False
                return False
            if seen is None:
                seen = set()
            if int(root_id) in seen:
                _contains_cache[key] = False
                return False
            seen.add(int(root_id))
            for ch in _child_items_of(int(root_id)):
                if int(ch) == int(target_id):
                    _contains_cache[key] = True
                    return True
                if _contains_target(int(ch), int(target_id), depth + 1, seen):
                    _contains_cache[key] = True
                    return True
            _contains_cache[key] = False
            return False

        def _find_root_for_item(target_id: int) -> Optional[int]:
            for rid in root_item_ids:
                try:
                    if _contains_target(int(rid), int(target_id)):
                        return int(rid)
                except Exception:
                    continue
            return None

        def _get_item_safe(iid: int) -> Optional[Item]:
            it = item_by_id.get(int(iid))
            if it is not None:
                return it
            try:
                it = db.query(Item).filter(Item.item_id == int(iid)).first()
                if it is not None:
                    item_by_id[int(iid)] = it
                return it
            except Exception:
                return None

        # Map spec_id -> parent item_id using default_spec_map (default specs only)
        spec_to_item: Dict[int, int] = {}
        try:
            for iid, sid in (default_spec_map or {}).items():
                try:
                    spec_to_item[int(sid)] = int(iid)
                except Exception:
                    continue
        except Exception:
            spec_to_item = {}

        # Build reverse parent map: child_item_id -> set(parent_item_id)
        from collections import defaultdict as _dd2
        parent_map: DefaultDict[int, Set[int]] = _dd2(set)
        try:
            for sid, parent_iid in (spec_to_item or {}).items():
                try:
                    comps = get_components_for_spec(int(sid)) or []
                except Exception:
                    comps = []
                for c in comps:
                    try:
                        child_iid = int(c.item_id)
                        parent_map[child_iid].add(int(parent_iid))
                    except Exception:
                        continue
        except Exception:
            parent_map = _dd2(set)

        def _find_top_root_via_parents(target_id: int) -> Optional[int]:
            """
            Walk upward using parent_map to find the highest ancestor which is a RootProduct.
            Falls back to None if not found or on cycles.
            """
            try:
                visited: Set[int] = set()
                frontier: Set[int] = {int(target_id)}
                steps = 0
                while frontier and steps < _max_upwalk_steps:
                    next_frontier: Set[int] = set()
                    for cur in list(frontier):
                        cur = int(cur)
                        if cur in visited:
                            continue
                        visited.add(cur)
                        if cur in root_item_ids:
                            return cur
                        parents = parent_map.get(cur) or set()
                        for p in parents:
                            try:
                                ip = int(p)
                                if ip not in visited:
                                    next_frontier.add(ip)
                            except Exception:
                                continue
                    frontier = next_frontier
                    steps += 1
                return None
            except Exception:
                return None

        # Resources list
        resources: List[ProductionResource] = db.query(ProductionResource).all()

        # Resources and production kind mapping (new logic)
        resource_production_kinds: List[ResourceProductionKind] = db.query(ResourceProductionKind).all()
        production_kinds_by_resource: Dict[int, Set[int]] = {}
        for rpk in resource_production_kinds:
            production_kinds_by_resource.setdefault(int(rpk.resource_id), set()).add(int(rpk.production_kind_id))

        all_stages: List[ProductionStage] = db.query(ProductionStage).all()
        stage_name_map: Dict[int, str] = {int(s.stage_id): str(s.stage_name or "") for s in all_stages}

        # All production kinds for name lookup
        all_production_kinds: List[ProductionKind] = db.query(ProductionKind).all()
        production_kind_name_map: Dict[int, str] = {int(pk.id): str(pk.name or "") for pk in all_production_kinds}

        # Item-level cached analysis: stage_id, production_kind_id and norm_hours (per unit)
        item_stage_cache: Dict[int, Tuple[Optional[int], Optional[str]]] = {}
        item_norm_cache: Dict[int, float] = {}
        item_production_kind_cache: Dict[int, Optional[int]] = {}

        def analyze_parent_item(item_id: int) -> Tuple[Optional[int], Optional[str], float, Optional[int]]:
            """
            Delegates stage determination and norm-hours calculation to shared helper,
            aligned with 'Распределение этапов' page logic.
            Returns (stage_id or None, reason_if_ambiguous, norm_hours_single, production_kind_id or None).

            Domain overrides for painting:
            - If operations' majority stage corresponds to painting ('покраска'/'paint'), force painting.
            - If item looks like a painted part by name/article/code (e.g., contains color markers or '-SP' suffix),
              force painting.
            """
            # cache hit
            if item_id in item_stage_cache and item_id in item_norm_cache and item_id in item_production_kind_cache:
                st, reason = item_stage_cache[item_id]
                return st, reason, item_norm_cache[item_id], item_production_kind_cache[item_id]

            stg_id, rsn, nh = determine_parent_stage_and_norm(
                default_spec_map=default_spec_map,
                get_components_for_spec=get_components_for_spec,
                get_operations_for_spec=get_operations_for_spec,
                item_id=int(item_id),
            )

            # Get production kind from specification
            spec_id = default_spec_map.get(int(item_id))
            production_kind_id = None
            if spec_id:
                spec = spec_by_id.get(int(spec_id))
                if spec:
                    production_kind_id = spec.production_kind_id

            # Helpers
            def _painting_stage_id() -> Optional[int]:
                try:
                    for sid, nm in stage_name_map.items():
                        nml = (nm or "").strip().lower()
                        if "покраск" in nml or "paint" in nml:
                            return int(sid)
                except Exception:
                    return None
                return None

            def _is_painting_stage(sid: Optional[int]) -> bool:
                if sid is None:
                    return False
                name = (stage_name_map.get(int(sid)) or "").strip().lower()
                return ("покраск" in name) or ("paint" in name)

            def _looks_painted() -> bool:
                it = item_by_id.get(int(item_id))
                if not it:
                    return False
                txt = " ".join([
                    str(getattr(it, "item_name", "") or ""),
                    str(getattr(it, "item_article", "") or ""),
                    str(getattr(it, "item_code", "") or ""),
                ]).lower()
                # Heuristics: color adjectives and explicit paint keywords/suffixes
                color_keywords = [
                    "красн", "черн", "бел", "син", "голуб", "зел", "сер", "оранж",
                    "желт", "жёлт", "фиолет", "бордов", "слонов", "покраск", "покраш"
                ]
                if any(k in txt for k in color_keywords):
                    return True
                # Code suffixes like -SP, GSP
                code = str(getattr(it, "item_code", "") or "")
                if re.search(r"(-|_)?(g?sp)\b", code.lower()):
                    return True
                article = str(getattr(it, "item_article", "") or "")
                if re.search(r"(-|_)?(g?sp)\b", article.lower()):
                    return True
                return False

            # Painting override: from operations majority
            # Apply only if production_kind_id is not defined (to maintain priority of production kinds)
            if production_kind_id is None:
                try:
                    spec_id_local = default_spec_map.get(int(item_id))
                    op_major: Optional[int] = None
                    if spec_id_local:
                        ops_local = get_operations_for_spec(int(spec_id_local)) or []
                        from collections import defaultdict as _dd
                        op_counts: Dict[int, int] = _dd(int)
                        for op in ops_local:
                            sid = getattr(op, "stage_id", None)
                            if sid is None:
                                continue
                            try:
                                op_counts[int(sid)] += 1
                            except Exception:
                                continue
                        if op_counts:
                            max_cnt = max(op_counts.values())
                            top = [sid for sid, cnt in op_counts.items() if cnt == max_cnt]
                            if len(top) == 1:
                                op_major = int(top[0])

                    # Apply only when base stage is not determined yet
                    if stg_id is None and op_major is not None and _is_painting_stage(op_major):
                        stg_id = int(op_major)
                        rsn = "FORCE_PAINTING_FROM_OPERATIONS"
                except Exception:
                    # fail-safe: ignore override on any error
                    pass

                # Painting override: by item attributes (name/article/code) — behind config flag and only when stage is not defined
                try:
                    enable_by_name = bool(((snapshot.get("production") or {}).get("toggles") or {}).get("force_painting_by_name", False))
                    if enable_by_name and stg_id is None and _looks_painted():
                        paint_sid = _painting_stage_id()
                        if paint_sid is not None:
                            stg_id = int(paint_sid)
                            rsn = "FORCE_PAINTING_BY_NAME"
                except Exception:
                    pass

            item_stage_cache[item_id] = (stg_id, rsn)
            item_norm_cache[item_id] = float(nh or 0.0)
            item_production_kind_cache[item_id] = production_kind_id
            return stg_id, rsn, float(nh or 0.0), production_kind_id

        def pick_area_for_production_kind(production_kind_id: int) -> Optional[int]:
            """Pick area based on production kind mapping"""
            candidates = [rid for rid, pkset in production_kinds_by_resource.items() if production_kind_id in pkset]
            if not candidates:
                return None
            if len(candidates) == 1:
                return candidates[0]
            # Heuristic: return first candidate sorted by id
            return sorted(candidates)[0]

        # Stage-based area mapping removed (migration to production kinds)

        # 6) Backward scheduling of stages by days with resource capacity
        stages_created = 0

        # Resource dictionaries and helpers
        res_by_id: Dict[int, ProductionResource] = {int(r.resource_id): r for r in resources}

        def _is_workday(d: date) -> bool:
            # Mon..Fri are workdays
            return d.weekday() <= 4

        def _day_available_hours(res: Optional[ProductionResource], d: date) -> float:
            if not _is_workday(d) or res is None:
                return 0.0
            daily_hours = float(getattr(res, "daily_work_hours", 8.0) or 8.0)
            power_coeff = float(getattr(res, "capacity", 1.0) or 1.0)
            return daily_hours * power_coeff

        # Accumulator of planned hours per (area_id, day)
        capacity_usage_daily: DefaultDict[Tuple[int, date], float] = defaultdict(float)

        def _pick_area_for_day(stage_id_val: int, production_kind_id_val: Optional[int], d: date) -> Optional[int]:
            # Choose candidate resource with maximum free capacity on date d
            # Use ONLY production kind mapping (no stage fallback per migration plan)
            candidates = []
            if production_kind_id_val is not None:
                candidates = [rid for rid, pkset in production_kinds_by_resource.items() if production_kind_id_val in pkset]
            if not candidates:
                return None
            best_rid: Optional[int] = None
            best_free: float = -1.0
            for rid in candidates:
                res = res_by_id.get(int(rid))
                avail = _day_available_hours(res, d)
                used = float(capacity_usage_daily.get((int(rid), d), 0.0))
                free = max(0.0, avail - used)
                if free > best_free:
                    best_free = free
                    best_rid = int(rid)
            # Even if all candidates have zero free capacity, return a candidate to allow placement/overflow accounting
            return best_rid if best_rid is not None else int(sorted(candidates)[0])

        d0: date = date.today()
        # Capacity scheduler: used to cap quantities by available capacity in [d0..need_date]
        scheduler = CapacityScheduler(
            res_by_id=res_by_id,
            production_kinds_by_resource=production_kinds_by_resource,
            use_calendar_5_2=True,
        )

        # Limits: configurable max upward walk steps for finding top root via parents
        limits_cfg = ((snapshot.get("planning") or {}).get("limits") or {})
        try:
            _max_upwalk_steps = int(limits_cfg.get("max_upwalk_steps", 300) or 300)
        except Exception:
            _max_upwalk_steps = 300

        # Track missing mappings summary (production kind primary, stage kept for analytics/backward compatibility)
        missing_area_pk_counts: DefaultDict[int, int] = defaultdict(int)
        missing_area_stage_counts: DefaultDict[int, int] = defaultdict(int)

        # Pre-calculate priorities and sort orders before capacity scheduling
        # 1) Рассчитываем нормо-часы на единицу для номенклатур из созданных заказов (по операциям спецификации)
        item_norm_cache_pre: Dict[int, float] = {}
        seen_items: Set[int] = set(int(o.item_id) for o in created_orders)
        for iid_pre in seen_items:
            try:
                sid = default_spec_map.get(int(iid_pre))
                total_norm = 0.0
                if sid:
                    ops_local = get_operations_for_spec(int(sid)) or []
                    total_norm = float(sum(float(op.time_norm or 0.0) for op in ops_local))
                item_norm_cache_pre[int(iid_pre)] = float(total_norm)
            except Exception:
                item_norm_cache_pre[int(iid_pre)] = 0.0

        # 2) Считаем приоритеты единым менеджером
        pm = PriorityManager(snapshot)
        order_priorities: Dict[int, float] = pm.compute_order_priorities(
            db=db,
            created_orders=created_orders,
            item_norm_cache=item_norm_cache_pre,
            net_daily=net_daily,
            net_weekly=net_weekly,
            items=items,
        )

        # 3) Сортируем заказы по приоритету (от высшего к низшему) для разрешения конфликтов
        created_orders_sorted = sorted(
            created_orders,
            key=lambda x: order_priorities.get(int(x.order_id), 0.0),
            reverse=True,
        )

        for order in created_orders_sorted:
            item_id = int(order.item_id)
            qty = float(order.qty or 0.0)
            stage_id, reason, norm_single, production_kind_id = analyze_parent_item(item_id)
            if stage_id is None and production_kind_id is None:
                if reason in {"NO_CHILD_STAGE", "MIXED_CHILD_STAGES"}:
                    warnings.append(
                        log_warning(
                            logger,
                            str(reason),
                            f"Stage cannot be determined for item_id={item_id}",
                            item_id=int(item_id),
                        )
                    )
                # Skip if both stage and production kind are undefined
                continue

            # Pick candidate areas based on production kind first, then stage as fallback
            candidate_res_ids = []
            if production_kind_id is not None:
                candidate_res_ids = [rid for rid, pkset in production_kinds_by_resource.items() if production_kind_id in pkset]
            # No stage-based fallback per migration plan
            if not candidate_res_ids:
                if production_kind_id is not None:
                    # Enrich diagnostic: item + spec + production kind name
                    try:
                        item_rec = item_by_id.get(int(item_id))
                    except Exception:
                        item_rec = None
                    try:
                        spec_id_local = default_spec_map.get(int(item_id))
                    except Exception:
                        spec_id_local = None
                    spec_ref1c_local = None
                    spec_code_local = None
                    spec_name_local = None
                    if spec_id_local is not None:
                        try:
                            spec_local = spec_by_id.get(int(spec_id_local))
                            if spec_local is not None:
                                spec_ref1c_local = getattr(spec_local, "spec_ref1c", None)
                                spec_code_local = getattr(spec_local, "spec_code", None)
                                spec_name_local = getattr(spec_local, "spec_name", None)
                        except Exception:
                            spec_ref1c_local = None
                            spec_code_local = None
                            spec_name_local = None
                    pk_name = None
                    try:
                        pk_name = production_kind_name_map.get(int(production_kind_id))
                    except Exception:
                        pk_name = None

                    # Determine top-level root product for diagnostics (prefer upward walk via parents, fallback to downward reachability)
                    root_id_val = _find_top_root_via_parents(int(item_id)) or _find_root_for_item(int(item_id))
                    root_rec = _get_item_safe(int(root_id_val)) if root_id_val is not None else None

                    warnings.append(
                        log_warning(
                            logger,
                            "NO_AREA_FOR_PRODUCTION_KIND",
                            f"No resource area for production_kind_id={int(production_kind_id)}",
                            production_kind_id=int(production_kind_id),
                            production_kind_name=pk_name,
                            item_id=int(item_id),
                            item_code=getattr(item_rec, "item_code", None) if item_rec is not None else None,
                            item_name=getattr(item_rec, "item_name", None) if item_rec is not None else None,
                            item_article=getattr(item_rec, "item_article", None) if item_rec is not None else None,
                            root_item_id=int(root_id_val) if root_id_val is not None else int(item_id),
                            root_item_code=getattr(root_rec, "item_code", None) if root_rec is not None else (getattr(item_rec, "item_code", None) if item_rec is not None else None),
                            root_item_name=getattr(root_rec, "item_name", None) if root_rec is not None else (getattr(item_rec, "item_name", None) if item_rec is not None else None),
                            root_item_article=getattr(root_rec, "item_article", None) if root_rec is not None else (getattr(item_rec, "item_article", None) if item_rec is not None else None),
                            spec_code=spec_code_local,
                            spec_name=spec_name_local,
                            spec_id=int(spec_id_local) if spec_id_local is not None else None,
                            spec_ref1c=spec_ref1c_local,
                        )
                    )
                    missing_area_pk_counts[int(production_kind_id)] += 1
                else:
                    warnings.append(
                        log_warning(
                            logger,
                            "NO_PRODUCTION_KIND",
                            f"No production kind for item_id={int(item_id)}",
                            item_id=int(item_id),
                        )
                    )
                continue

            total_hours = float(norm_single or 0.0) * float(qty or 0.0)

            # Capacity limiting before scheduling: reduce qty if window capacity is insufficient
            if production_kind_id is not None and float(norm_single or 0.0) > 0.0 and float(qty or 0.0) > 0.0:
                try:
                    limited_qty, free_hours_window, workdays_window = scheduler.limit_qty_by_capacity(
                        qty=float(qty),
                        norm_hours_per_unit=float(norm_single),
                        production_kind_id=int(production_kind_id),
                        d0=d0,
                        need_date=order.need_date,
                        capacity_usage_daily=capacity_usage_daily,
                    )
                    # Округление количества после ограничения мощностью: только целые значения (ceil)
                    rounded_limited_qty = math.ceil(float(limited_qty))
                    if float(rounded_limited_qty) < float(qty) - 1e-9:
                        warnings.append(
                            log_warning(
                                logger,
                                "CAPACITY_LIMITED",
                                f"Order {int(order.order_id)} qty limited by capacity from {float(qty)} to {float(rounded_limited_qty)}",
                                order_id=int(order.order_id),
                                item_id=int(item_id),
                                from_qty=float(qty),
                                to_qty=float(rounded_limited_qty),
                                free_hours_window=float(free_hours_window),
                                workdays_window=int(workdays_window),
                                limited_qty_raw=float(limited_qty),
                            )
                        )
                        # Недопланированный объём = разница между нормализованным и ограниченным (с округлением вверх) количеством
                        try:
                            unmet_qty = float(qty) - float(rounded_limited_qty)
                        except Exception:
                            unmet_qty = 0.0
                        if unmet_qty > 1e-9:
                            # Попробуем определить предполагаемый ресурс (по виду производства)
                            try:
                                resource_hint = int(sorted(candidate_res_ids)[0]) if candidate_res_ids else None
                            except Exception:
                                resource_hint = None
                            # Детали расчёта количества (для корректного логирования дефицита мощности)
                            comp_details = getattr(order, "_comp_details", None)
                            req_qty_orig = float(comp_details.get("requested_qty", 0.0)) if isinstance(comp_details, dict) else None
                            final_before_cap = float(comp_details.get("final_qty_before_capacity", float(qty))) if isinstance(comp_details, dict) else float(qty)
                            normalized_qty_calc = float(comp_details.get("normalized_qty", float(qty))) if isinstance(comp_details, dict) else float(qty)
                            requested_before_capacity = float(normalized_qty_calc)
                            warnings.append(
                                log_warning(
                                    logger,
                                    "CAPACITY_SHORTAGE",
                                    f"Item {int(item_id)} unmet quantity due to capacity on need_date",
                                    order_id=int(order.order_id),
                                    item_id=int(item_id),
                                    production_kind_id=int(production_kind_id) if production_kind_id is not None else None,
                                    resource_id=resource_hint,
                                    need_date=order.need_date.isoformat() if order.need_date else None,
                                    requested_qty=float(req_qty_orig) if req_qty_orig is not None else None,
                                    final_qty_before_capacity=float(final_before_cap),
                                    normalized_qty=float(normalized_qty_calc),
                                    requested_qty_before_capacity=float(requested_before_capacity),
                                    limited_by_capacity=float(limited_qty),
                                    unmet_qty=float(unmet_qty),
                                    norm_hours_per_unit=float(norm_single or 0.0),
                                    unmet_norm_hours=float(unmet_qty) * float(norm_single or 0.0),
                                )
                            )
                        # Применяем ограничение с округлением вверх
                        order.qty = float(rounded_limited_qty)
                        qty = float(rounded_limited_qty)
                        total_hours = float(norm_single or 0.0) * float(qty or 0.0)
                except Exception:
                    # fail-safe: ignore capacity limiting if any error
                    pass

            # Zero-norm safeguard:
            # If total norm-hours is <= 0, we still create a stage slice with hours=0
            # so the order is grouped by stage/area in UI (instead of falling into '—').
            if total_hours <= 1e-9:
                sel_area_id = None
                if production_kind_id is not None:
                    sel_area_id = pick_area_for_production_kind(int(production_kind_id))
                if sel_area_id is None:
                    # Fallback: choose any mapped resource for stability (based on already computed candidates)
                    sel_area_id = int(sorted(candidate_res_ids)[0]) if candidate_res_ids else None
                if sel_area_id is not None:
                    pos = PlannedOrderStage(
                        run_id=run_id,
                        order_id=int(order.order_id),
                        stage_id=int(stage_id) if stage_id is not None else None,  # Stage still used for operations
                        area_id=int(sel_area_id),
                        bucket_type="daily",
                        bucket_date=order.need_date,
                        hours=0.0,
                    )
                    db.add(pos)
                    stages_created += 1
                    # Do not affect capacity usage; set order dates for consistency
                    order.start_date = order.need_date
                    order.finish_date = order.need_date
                else:
                    # Enrich zero-norm diagnostic with item/spec/pk info
                    try:
                        item_rec_zn = item_by_id.get(int(order.item_id))
                    except Exception:
                        item_rec_zn = None
                    try:
                        spec_id_zn = default_spec_map.get(int(order.item_id))
                    except Exception:
                        spec_id_zn = None
                    spec_ref1c_zn = None
                    spec_code_zn = None
                    spec_name_zn = None
                    if spec_id_zn is not None:
                        try:
                            spec_zn = spec_by_id.get(int(spec_id_zn))
                            if spec_zn is not None:
                                spec_ref1c_zn = getattr(spec_zn, "spec_ref1c", None)
                                spec_code_zn = getattr(spec_zn, "spec_code", None)
                                spec_name_zn = getattr(spec_zn, "spec_name", None)
                        except Exception:
                            spec_ref1c_zn = None
                            spec_code_zn = None
                            spec_name_zn = None
                    pk_name_zn = None
                    if production_kind_id is not None:
                        try:
                            pk_name_zn = production_kind_name_map.get(int(production_kind_id))
                        except Exception:
                            pk_name_zn = None

                    # Determine top-level root product for diagnostics (zero-norm case): prefer upward walk
                    root_id_zn = _find_top_root_via_parents(int(order.item_id)) or _find_root_for_item(int(order.item_id))
                    root_rec_zn = _get_item_safe(int(root_id_zn)) if root_id_zn is not None else None

                    warnings.append(
                        log_warning(
                            logger,
                            "NO_AREA_FOR_PRODUCTION_KIND_ZERO_NORM",
                            f"No resource area resolved for zero-norm production_kind_id={int(production_kind_id) if production_kind_id is not None else 'N/A'}",
                            production_kind_id=int(production_kind_id) if production_kind_id is not None else None,
                            production_kind_name=pk_name_zn,
                            order_id=int(order.order_id),
                            item_id=int(order.item_id),
                            item_code=getattr(item_rec_zn, "item_code", None) if item_rec_zn is not None else None,
                            item_name=getattr(item_rec_zn, "item_name", None) if item_rec_zn is not None else None,
                            item_article=getattr(item_rec_zn, "item_article", None) if item_rec_zn is not None else None,
                            root_item_id=int(root_id_zn) if root_id_zn is not None else int(order.item_id),
                            root_item_code=getattr(root_rec_zn, "item_code", None) if root_rec_zn is not None else (getattr(item_rec_zn, "item_code", None) if item_rec_zn is not None else None),
                            root_item_name=getattr(root_rec_zn, "item_name", None) if root_rec_zn is not None else (getattr(item_rec_zn, "item_name", None) if item_rec_zn is not None else None),
                            root_item_article=getattr(root_rec_zn, "item_article", None) if root_rec_zn is not None else (getattr(item_rec_zn, "item_article", None) if item_rec_zn is not None else None),
                            spec_code=spec_code_zn,
                            spec_name=spec_name_zn,
                            spec_id=int(spec_id_zn) if spec_id_zn is not None else None,
                            spec_ref1c=spec_ref1c_zn,
                        )
                    )
                # Skip regular scheduling in zero-norm case
                continue

            remaining = total_hours
            # Use unified capacity scheduler for backward allocation
            slices, residual = scheduler.schedule_backward(
                total_hours=float(total_hours),
                production_kind_id=production_kind_id,
                d0=d0,
                need_date=order.need_date,
                capacity_usage_daily=capacity_usage_daily,
            )
            used_dates: List[date] = [d for (_, d, _) in slices]

            for sel_area_id, day, hours in slices:
                pos = PlannedOrderStage(
                    run_id=run_id,
                    order_id=int(order.order_id),
                    stage_id=int(stage_id) if stage_id is not None else None,
                    area_id=int(sel_area_id),
                    bucket_type="daily",
                    bucket_date=day,
                    hours=float(hours),
                )
                db.add(pos)
                stages_created += 1

            # If some hours remain, place them to earliest workday at/ before d0 (allow overload) and warn
            fallback_used_date: Optional[date] = None
            if residual > 1e-6:
                fb = d0
                # find workday at/ before d0 (max 7-day lookback as safety)
                tries = 7
                while tries > 0 and not _is_workday(fb):
                    fb = fb - timedelta(days=1)
                    tries -= 1

                sel_area_id_fb = _pick_area_for_day(int(stage_id) if stage_id is not None else None, production_kind_id, fb) or int(sorted(candidate_res_ids)[0])

                pos = PlannedOrderStage(
                    run_id=run_id,
                    order_id=int(order.order_id),
                    stage_id=int(stage_id) if stage_id is not None else None,
                    area_id=int(sel_area_id_fb),
                    bucket_type="daily",
                    bucket_date=fb,
                    hours=float(residual),
                )
                db.add(pos)
                stages_created += 1
                capacity_usage_daily[(int(sel_area_id_fb), fb)] += float(residual)
                fallback_used_date = fb
                warnings.append(
                    log_warning(
                        logger,
                        "SCHED_OVERFLOW",
                        f"Order {int(order.order_id)} overflow: scheduled residual {float(residual):.3f}h on {fb.isoformat()}",
                        order_id=int(order.order_id),
                        residual_hours=float(residual),
                        date=fb.isoformat(),
                    )
                )
                residual = 0.0

            # Set start/finish dates on order
            if used_dates or fallback_used_date:
                all_dates = list(used_dates)
                if fallback_used_date:
                    all_dates.append(fallback_used_date)
                order.start_date = min(all_dates)
                order.finish_date = max(all_dates)

        db.flush()

        # 7) Aggregate CapacityLoad for daily and weekly buckets
        def _iso_friday(d: date) -> date:
            return _iso_week_friday(d)

        overload_total = 0.0
        overloaded_buckets = 0

        # Daily capacity load
        weekly_acc: DefaultDict[Tuple[int, date], float] = defaultdict(float)  # (area_id, friday) -> planned hours

        for (area_id, day), hours_planned in capacity_usage_daily.items():
            res = res_by_id.get(int(area_id))
            hours_available = _day_available_hours(res, day)
            overload = max(0.0, float(hours_planned or 0.0) - float(hours_available or 0.0))

            cap_d = CapacityLoad(
                run_id=run_id,
                area_id=int(area_id),
                bucket_type="daily",
                bucket_date=day,
                hours_planned=float(hours_planned or 0.0),
                hours_available=float(hours_available or 0.0),
                overload_hours=float(overload or 0.0),
            )
            db.add(cap_d)

            overload_total += float(overload or 0.0)
            if float(overload or 0.0) > 0.0:
                overloaded_buckets += 1

            # accumulate for weekly
            friday = _iso_friday(day)
            weekly_acc[(int(area_id), friday)] += float(hours_planned or 0.0)

        # Weekly capacity load (5/2 calendar → 5 workdays per week)
        for (area_id, friday), hours_planned in weekly_acc.items():
            res = res_by_id.get(int(area_id))
            daily_hours = float(getattr(res, "daily_work_hours", 8.0) or 8.0)
            power_coeff = float(getattr(res, "capacity", 1.0) or 1.0)
            hours_available = daily_hours * 5.0 * power_coeff
            overload = max(0.0, float(hours_planned or 0.0) - float(hours_available or 0.0))

            cap_w = CapacityLoad(
                run_id=run_id,
                area_id=int(area_id),
                bucket_type="weekly",
                bucket_date=friday,
                hours_planned=float(hours_planned or 0.0),
                hours_available=float(hours_available or 0.0),
                overload_hours=float(overload or 0.0),
            )
            db.add(cap_w)

            overload_total += float(overload or 0.0)
            if float(overload or 0.0) > 0.0:
                overloaded_buckets += 1

        db.flush()

        if overloaded_buckets > 0:
            warnings.append(
                log_warning(
                    logger,
                    "CAPACITY_OVERLOAD",
                    f"Detected {overloaded_buckets} overloaded buckets, total overload hours={overload_total:.3f}",
                    overloaded_buckets=int(overloaded_buckets),
                    overload_total=float(overload_total),
                )
            )

        # Summary of missing production_kind->area mappings
        if missing_area_pk_counts:
            warnings.append(
                log_warning(
                    logger,
                    "NO_AREA_FOR_PRODUCTION_KIND_SUMMARY",
                    "Missing resource area mapping for production kinds",
                    by_production_kind=[
                        {"production_kind_id": int(pkid), "production_kind_name": production_kind_name_map.get(int(pkid)), "count": int(cnt)}
                        for pkid, cnt in sorted(missing_area_pk_counts.items())
                    ],
                )
            )

        # 8) Pegging (one-level child→parent by bucket, per planned order)
        pb = PeggingBuilder()
        links = pb.build(
            run_id=int(run_id),
            orders=created_orders_sorted,
            default_spec_map=default_spec_map,
            get_components_for_spec=get_components_for_spec,
        )
        if links:
            db.add_all(links)

        db.flush()

        # 9) Присваиваем приоритеты: заказы — предрасчитанные, закупки — через PriorityManager
        for o in created_orders:
            o.priority_index = float(order_priorities.get(int(o.order_id), 0.0))

        pm2 = PriorityManager(snapshot)
        pm2.assign_purchase_priorities(created_purchases)

        db.flush()

        # 10) Finalize run with extended KPI
        kpi = {
            "orders_created": int(len(created_orders)),
            "purchases_created": int(len(created_purchases)),
            "items_involved": int(len(all_item_ids)),
            "stages_count": int(stages_created),
            "capacity_overload_total": float(overload_total),
            "overloaded_buckets": int(overloaded_buckets),
            "missing_area_stage_distinct": int(len(missing_area_stage_counts)),
            "missing_area_stage_total": int(sum(missing_area_stage_counts.values())),
        }

        run.status = "SUCCESS"
        run.kpi = kpi
        run.warnings = warnings
        run.finished_at = datetime.utcnow()

        db.commit()
        return run_id

    except Exception as e:
        try:
            run.status = "FAILED"
            run.finished_at = datetime.utcnow()
            w = warnings or []
            w.append({"code": "RUN_FAILED", "msg": str(e)})
            run.warnings = w
            db.commit()
        except Exception:
            db.rollback()
        raise

# --- Retention & pin management for planning runs ---

def cleanup_planning_runs(db: Session, older_than_days: int = 30, dry_run: bool = False) -> Dict[str, Any]:
    """
    Delete planning runs older than N days except those marked as pinned.
    Returns a report with affected run_ids. When dry_run=True, nothing is deleted.
    """
    try:
        days = max(1, int(older_than_days or 30))
    except Exception:
        days = 30
    cutoff_dt = datetime.utcnow() - timedelta(days=days)

    q = (
        db.query(PlanningRun)
        .filter(
            PlanningRun.started_at < cutoff_dt,
            PlanningRun.pinned.is_(False),
        )
    )
    runs: List[PlanningRun] = q.all()
    run_ids: List[int] = [int(r.run_id) for r in runs]

    report: Dict[str, Any] = {
        "dry_run": bool(dry_run),
        "older_than_days": int(days),
        "cutoff": cutoff_dt.isoformat(),
        "count": len(run_ids),
        "run_ids": run_ids,
    }

    if dry_run:
        return report

    if run_ids:
        # Cascade deletes will clear dependent tables (planned_order, planned_purchase, capacity_load, pegging_link, etc.)
        q.delete(synchronize_session=False)
        db.commit()

    return report


def set_run_pinned(db: Session, run_id: int, pinned: bool) -> Dict[str, Any]:
    """
    Set pinned flag for a planning run to protect from cleanup.
    """
    r: Optional[PlanningRun] = db.query(PlanningRun).filter(PlanningRun.run_id == int(run_id)).first()
    if not r:
        raise RuntimeError(f"Run {run_id} not found")
    r.pinned = bool(pinned)
    db.commit()
    return {"run_id": int(r.run_id), "pinned": bool(r.pinned)}
# ===== Backend-first aggregated/grouped helpers (compat with plan.py imports) =====

def get_capacity_summary(
    db: Session,
    run_id: int,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Сводка по мощности (часы/перегрузы) по участкам в заданном диапазоне.
    Возвращает: {"map": { area_id: { hours_planned, hours_available, overload_hours, overloaded_buckets } } }
    """
    q = db.query(CapacityLoad).filter(CapacityLoad.run_id == int(run_id))
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(CapacityLoad.bucket_type == bucket_type)
    if date_from:
        q = q.filter(CapacityLoad.bucket_date >= _to_date(date_from))
    if date_to:
        q = q.filter(CapacityLoad.bucket_date <= _to_date(date_to))

    rows: List[CapacityLoad] = q.all()
    result: Dict[int, Dict[str, Any]] = {}
    for r in rows:
        aid = int(r.area_id)
        rec = result.setdefault(aid, {
            "hours_planned": 0.0,
            "hours_available": 0.0,
            "overload_hours": 0.0,
            "overloaded_buckets": 0,
        })
        hp = float(r.hours_planned or 0.0)
        ha = float(r.hours_available or 0.0)
        ov = float(r.overload_hours or 0.0)
        rec["hours_planned"] += hp
        rec["hours_available"] += ha
        rec["overload_hours"] += ov
        if ov > 0.0:
            rec["overloaded_buckets"] += 1
    return {"map": result}


def get_run_purchases_grouped(
    db: Session,
    run_id: int,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000,
    offset: int = 0,
) -> Dict[str, Any]:
    """
    Сводная группировка закупок по (item_id, unit).
    Возвращает: {"rows":[{item_id,item_name,item_article,unit,qty,agg_key}], "total": N}
    """
    q = (
        db.query(
            PlannedPurchase,
            Item.item_name,
            Item.item_article,
            Item.unit,
            Unit.short_name,
            Unit.unit_name,
            Unit.unit_code,
        )
        .outerjoin(Item, PlannedPurchase.item_id == Item.item_id)
        .outerjoin(Unit, Item.unit == Unit.unit_ref1c)
        .filter(PlannedPurchase.run_id == int(run_id))
    )
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(PlannedPurchase.bucket_type == bucket_type)

    joined = q.all()
    
    # Фильтрация по датам после получения данных для корректной агрегации
    # Важно: для закупок используем bucket_date (need_date) для определения пересечения с диапазоном
    filtered_joined = []
    for r, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code in joined:
        include_row = True
        if date_from:
            # Включаем закупку, если её дата >= начала диапазона
            if r.bucket_date is None or r.bucket_date < _to_date(date_from):
                include_row = False
        if date_to:
            # Включаем закупку, если её дата <= конца диапазона
            if r.bucket_date is None or r.bucket_date > _to_date(date_to):
                include_row = False
        if include_row:
            filtered_joined.append((r, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code))

    agg: Dict[Tuple[int, str], Dict[str, Any]] = {}
    for r, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code in filtered_joined:
        unit_disp = in_unit_short or in_unit_name or in_unit_code or in_unit_guid or ""
        key = (int(r.item_id), str(unit_disp))
        rec = agg.setdefault(key, {
            "item_id": int(r.item_id),
            "item_name": in_name,
            "item_article": in_article,
            "unit": unit_disp,
            "qty": 0.0,
        })
        rec["qty"] += float(r.qty or 0.0)

    all_rows = []
    for (iid, unit_disp), rec in agg.items():
        rec["agg_key"] = f"{iid}|{unit_disp}"
        all_rows.append(rec)

    total = len(all_rows)
    # simple pagination on aggregated rows
    sliced = all_rows[int(offset): int(offset) + int(limit)]
    return {"rows": sliced, "total": int(total), "limit": int(limit), "offset": int(offset)}


def get_run_production_grouped(
    db: Session,
    run_id: int,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    area_id: Optional[int] = None,
    limit: int = 1000,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Группировка производственных заказов по «доминирующему участку» (dominant area, по сумме часов стадий).
    Возвращает:
      {
        "groups": [
          {
            "area_id": number,
            "area_name": string,
            "orders": [{ item_id, item_name, item_article, unit, qty, norm_hours_total, norm_hours_per_unit, agg_key }],
            "norm_sum_hours": number,
            "min_days_to_need": number|null,
            "cap_overload_hours": number,
            "cap_overloaded_buckets": number
          }
        ],
        "total": groups_count
      }
    """
    # 1) Базовые заказы с денормализацией номенклатуры
    q = (
        db.query(
            PlannedOrder,
            Item.item_name,
            Item.item_article,
            Item.unit,
            Unit.short_name,
            Unit.unit_name,
            Unit.unit_code,
        )
        .outerjoin(Item, PlannedOrder.item_id == Item.item_id)
        .outerjoin(Unit, Item.unit == Unit.unit_ref1c)
        .filter(PlannedOrder.run_id == int(run_id))
    )
    if bucket_type in {"daily", "weekly"}:
        q = q.filter(PlannedOrder.bucket_type == bucket_type)

    ord_rows = q.all()
    
    # Фильтрация по датам после получения данных для корректной агрегации
    # Важно: заказы должны включаться в диапазон, если они пересекают его (start_date <= date_to AND finish_date >= date_from)
    filtered_ord_rows = []
    for po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code in ord_rows:
        include_row = True
        if date_from:
            # Включаем заказ, если его дата окончания >= начала диапазона (т.е. пересекает диапазон)
            if po.finish_date is None or _to_date(date_from) > po.finish_date:
                include_row = False
        if date_to:
            # Включаем заказ, если его дата начала <= конца диапазона (т.е. пересекает диапазон)
            if po.start_date is None or po.start_date > _to_date(date_to):
                include_row = False
        if area_id is not None:
            # Для правильной фильтрации по area_id нужно сначала определить доминирующий участок
            # поэтому фильтрация будет происходить позже
            pass
        if include_row:
            filtered_ord_rows.append((po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code))

    # Fallback cache: per-item norm from specification operations for qty==0
    fallback_npu_grp: Dict[int, float] = {}
    try:
        item_ids_grp: List[int] = list({int(r[0].item_id) for r in filtered_ord_rows})
    except Exception:
        item_ids_grp = []
    if item_ids_grp:
        try:
            defs_g = (
                db.query(DefaultSpecification)
                .filter(DefaultSpecification.item_id.in_(item_ids_grp))
                .all()
            )
            item_to_spec_g: Dict[int, int] = {}
            spec_ids_g: Set[int] = set()
            for d in defs_g:
                try:
                    iid = int(d.item_id)
                    sid = int(d.spec_id)
                    item_to_spec_g[iid] = sid
                    spec_ids_g.add(sid)
                except Exception:
                    continue
            if spec_ids_g:
                rows_g = (
                    db.query(
                        SpecOperation.spec_id.label("spec_id"),
                        func.sum(func.coalesce(SpecOperation.time_norm, Operation.time_norm)).label("sum_norm"),
                    )
                    .join(Operation, SpecOperation.operation_id == Operation.operation_id)
                    .filter(SpecOperation.spec_id.in_(list(spec_ids_g)))
                    .group_by(SpecOperation.spec_id)
                    .all()
                )
                spec_norm_sum_g: Dict[int, float] = {int(r.spec_id): float(getattr(r, "sum_norm", 0.0) or 0.0) for r in rows_g}
                for iid, sid in item_to_spec_g.items():
                    try:
                        npu_val = float(spec_norm_sum_g.get(int(sid), 0.0) or 0.0)
                        if npu_val > 0.0:
                            fallback_npu_grp[int(iid)] = npu_val
                    except Exception:
                        continue
        except Exception as ex:
            logger.exception("fallback_npu_grp build failed: %s", ex)
            fallback_npu_grp = {}
    
    order_ids = [int(r[0].order_id) for r in filtered_ord_rows]
    
    # 2) Стадии только по выбранным заказам
    stages: List[PlannedOrderStage] = []
    if order_ids:
        stq = db.query(PlannedOrderStage).filter(
            PlannedOrderStage.run_id == int(run_id),
            PlannedOrderStage.order_id.in_(order_ids),
        )
        stages = stq.all()

    # map: order_id -> list[stage]
    st_by_order: Dict[int, List[PlannedOrderStage]] = {}
    for s in stages:
        st_by_order.setdefault(int(s.order_id), []).append(s)

    # Resource names
    res_rows: List[ProductionResource] = db.query(ProductionResource).all()
    area_name_map: Dict[int, str] = {int(r.resource_id): str(r.resource_name or "") for r in res_rows}

    # 3) Распределение заказов по dominat area и агрегирование
    from collections import defaultdict as _dd
    groups: Dict[int, Dict[str, Any]] = _dd(lambda: {
        "area_id": None,
        "area_name": "",
        "orders": [],
        "norm_sum_hours": 0.0,
        "min_days_to_need": None,
        "cap_overload_hours": 0.0,
        "cap_overloaded_buckets": 0,
    })

    today = date.today()
    
    # Агрегация заказов по item_id и unit до определения доминирующего участка
    item_aggregated: Dict[Tuple[int, str], Dict[str, Any]] = {}
    
    for po, in_name, in_article, in_unit_guid, in_unit_short, in_unit_name, in_unit_code in filtered_ord_rows:
        st_list = st_by_order.get(int(po.order_id), [])
        # суммируем часы по участкам, выбираем доминирующий
        hours_by_area: Dict[Optional[int], float] = _dd(float)
        total_hours = 0.0
        for s in st_list:
            aid = int(s.area_id) if s.area_id is not None else None
            h = float(getattr(s, "hours", 0.0) or 0.0)
            hours_by_area[aid] += h
            total_hours += h

        # fallback: если стадий нет — без участка (пропускаем из группировки)
        if not hours_by_area:
            continue

        dom_area_id: Optional[int] = None
        dom_hours = -1.0
        for aid, h in hours_by_area.items():
            if aid is not None and h > dom_hours:
                dom_area_id = int(aid)
                dom_hours = float(h)

        if dom_area_id is None:
            # если ни одной стадии с участком — пропускаем
            continue

        unit_disp = in_unit_short or in_unit_name or in_unit_code or in_unit_guid or ""
        # Ключ агрегации: item_id и unit
        agg_key = (int(po.item_id), unit_disp)
        
        if agg_key in item_aggregated:
            # Объединяем данные с существующей записью
            item_aggregated[agg_key]["qty"] += float(po.qty or 0.0)
            item_aggregated[agg_key]["norm_hours_total"] += float(total_hours)
            # Обновляем need_date, если текущая дата раньше
            current_need_date = item_aggregated[agg_key]["need_date"]
            if current_need_date is None or (po.need_date and po.need_date < _to_date(current_need_date)):
                item_aggregated[agg_key]["need_date"] = po.need_date.isoformat() if po.need_date else None
            # Также обновляем dom_area_id, если он отличается (хотя должен быть одинаковым для одного item_id)
            # В идеале, все заказы одного item_id должны иметь одинаковый доминирующий участок
        else:
            # Создаем новую запись
            qty_val = float(po.qty or 0.0)
            norm_per_unit = float(total_hours / qty_val) if qty_val > 1e-12 else None
            # Expanded fallback: also when norm_per_unit computed as ~0 (total_hours==0 with qty>0)
            if norm_per_unit is None or (norm_per_unit is not None and norm_per_unit <= 1e-12):
                try:
                    npu_fb = float(fallback_npu_grp.get(int(po.item_id), 0.0) or 0.0)
                    if npu_fb > 0.0:
                        norm_per_unit = npu_fb
                except Exception:
                    pass

            item_aggregated[agg_key] = {
                "agg_key": f"{int(po.item_id)}|{unit_disp}",
                "item_id": int(po.item_id),
                "item_name": in_name,
                "item_article": in_article,
                "unit": unit_disp,
                "qty": qty_val,
                "norm_hours_total": float(total_hours),
                "norm_hours_per_unit": float(norm_per_unit) if norm_per_unit is not None else None,
                "need_date": po.need_date.isoformat() if po.need_date else None,
                "dom_area_id": dom_area_id  # временно сохраняем для последующего распределения по группам
            }

    # Распределяем агрегированные заказы по группам (по доминирующему участку)
    for item_data in item_aggregated.values():
        dom_area_id = item_data["dom_area_id"]
        
        if area_id is not None and dom_area_id != int(area_id):
            continue
            
        g = groups[dom_area_id]
        g["area_id"] = dom_area_id
        g["area_name"] = area_name_map.get(dom_area_id, f"Участок #{dom_area_id}")
        g["orders"].append({
            "agg_key": item_data["agg_key"],
            "item_id": item_data["item_id"],
            "item_name": item_data["item_name"],
            "item_article": item_data["item_article"],
            "unit": item_data["unit"],
            "qty": item_data["qty"],
            "norm_hours_total": item_data["norm_hours_total"],
            "norm_hours_per_unit": item_data["norm_hours_per_unit"],
        })
        g["norm_sum_hours"] += item_data["norm_hours_total"]
        if item_data["need_date"]:
            try:
                need_date = _to_date(item_data["need_date"])
                delta = (need_date - today).days
                if g["min_days_to_need"] is None or delta < g["min_days_to_need"]:
                    g["min_days_to_need"] = int(delta)
            except Exception:
                pass

    # 4) Индикаторы мощности по группам
    if groups:
        aids = list(groups.keys())
        capq = db.query(CapacityLoad).filter(CapacityLoad.run_id == int(run_id), CapacityLoad.area_id.in_(aids))
        if bucket_type in {"daily", "weekly"}:
            capq = capq.filter(CapacityLoad.bucket_type == bucket_type)
        if date_from:
            # ORIGINAL: capq = capq.filter(CapacityLoad.bucket_date >= _to_date(date_from))
            # FIXED: используем правильное поле bucket_date для фильтрации
            capq = capq.filter(CapacityLoad.bucket_date >= _to_date(date_from))
        if date_to:
            # ORIGINAL: capq = capq.filter(CapacityLoad.bucket_date <= _to_date(date_to))
            # FIXED: используем правильное поле bucket_date для фильтрации
            capq = capq.filter(CapacityLoad.bucket_date <= _to_date(date_to))
        for r in capq.all():
            g = groups.get(int(r.area_id))
            if not g:
                continue
            ov = float(r.overload_hours or 0.0)
            g["cap_overload_hours"] += ov
            if ov > 0.0:
                g["cap_overloaded_buckets"] += 1

    # Применяем фильтрацию по area_id к результатам группировки
    if area_id is not None:
        filtered_groups = [g for g in groups.values() if g["area_id"] == int(area_id)]
    else:
        filtered_groups = list(groups.values())
    
    total = len(filtered_groups)
    # простой слайс (пагинация по группам)
    sliced = filtered_groups[int(offset): int(offset) + int(limit)]
    return {"groups": sliced, "total": int(total), "limit": int(limit), "offset": int(offset)}


# NOTE: Функция get_run_production_agenda_day удалена, так как фильтрация по одной дате больше не требуется.
# ORIGINAL FUNCTION:
# def get_run_production_agenda_day(
#     db: Session,
#     run_id: int,
#     day_date: str,
#     area_id: Optional[int] = None,
# ) -> Dict[str, Any]:
#     """
#     «Повестка дня» для конкретной даты: группировка по участкам.
#     Пересчёт количества из часов по норме на единицу (norm_hours_total / qty для заказа).
#     Возвращает: {"groups":[{area_id,area_name,orders,norm_sum_hours,sum_qty,cap_overload_hours}]}
#     """
#     day = _to_date(day_date)
#
#     # Страиваем карты: order_id -> total_hours, norm_per_unit
#     ord_q = db.query(PlannedOrder).filter(PlannedOrder.run_id == int(run_id))
#     orders = ord_q.all()
#     if not orders:
#         return {"groups": []}
#     ord_total_hours: Dict[int, float] = {}
#     for o in orders:
#         ord_total_hours[int(o.order_id)] = 0.0
#
#     stq = db.query(PlannedOrderStage).filter(
#         PlannedOrderStage.run_id == int(run_id),
#         PlannedOrderStage.bucket_type == "daily",
#         PlannedOrderStage.bucket_date == day,
#     )
#     if area_id is not None:
#         stq = stq.filter(PlannedOrderStage.area_id == int(area_id))
#     day_stages = stq.all()
#
#     # Для подсчёта norm_per_unit нужно знать все часы по заказу, не только за день
#     st_all_q = db.query(PlannedOrderStage).filter(PlannedOrderStage.run_id == int(run_id))
#     st_all = st_all_q.all()
#     for s in st_all:
#         oid = int(s.order_id)
#         ord_total_hours[oid] = ord_total_hours.get(oid, 0.0) + float(getattr(s, "hours", 0.0) or 0.0)
#
#     # Карта заказов и ЕИ
#     item_ids: Set[int] = set(int(o.item_id) for o in orders)
#     items: List[Item] = []
#     if item_ids:
#         items = db.query(Item).filter(Item.item_id.in_(item_ids)).all()
#     item_map: Dict[int, Item] = {int(x.item_id): x for x in items}
#
#     unit_guids: Set[str] = {getattr(i, "unit", None) for in items if getattr(i, "unit", None)}
#     units: List[Unit] = []
#     if unit_guids:
#         units = db.query(Unit).filter(Unit.unit_ref1c.in_(unit_guids)).all()
#     unit_map: Dict[str, Unit] = {str(u.unit_ref1c): u for u in units}
#
#     # Fallback cache for per-item norm from specification operations (for cases when qty=0)
#     fallback_npu_day: Dict[int, float] = {}
#     try:
#         item_ids_for_fallback: List[int] = list(item_ids)
#     except Exception:
#         item_ids_for_fallback = []
#     if item_ids_for_fallback:
#         try:
#             defs_d = (
#                 db.query(DefaultSpecification)
#                 .filter(DefaultSpecification.item_id.in_(item_ids_for_fallback))
#                 .all()
#             )
#             item_to_spec_d: Dict[int, int] = {}
#             spec_ids_d: Set[int] = set()
#             for d in defs_d:
#                 try:
#                     iid = int(d.item_id)
#                     sid = int(d.spec_id)
#                     item_to_spec_d[iid] = sid
#                     spec_ids_d.add(sid)
#                 except Exception:
#                     continue
#             if spec_ids_d:
#                 rows_d = (
#                     db.query(
#                         SpecOperation.spec_id.label("spec_id"),
#                         func.sum(func.coalesce(SpecOperation.time_norm, Operation.time_norm)).label("sum_norm"),
#                     )
#                     .join(Operation, SpecOperation.operation_id == Operation.operation_id)
#                     .filter(SpecOperation.spec_id.in_(list(spec_ids_d)))
#                     .group_by(SpecOperation.spec_id)
#                     .all()
#                 )
#                 spec_norm_sum_d: Dict[int, float] = {int(r.spec_id): float(getattr(r, "sum_norm", 0.0) or 0.0) for r in rows_d}
#                 for iid, sid in item_to_spec_d.items():
#                     try:
#                         npu_val = float(spec_norm_sum_d.get(int(sid), 0.0) or 0.0)
#                         if npu_val > 0.0:
#                             fallback_npu_day[int(iid)] = npu_val
#                     except Exception:
#                         continue
#         except Exception as ex:
#             logger.exception("fallback_npu_day build failed: %s", ex)
#             fallback_npu_day = {}
#     # Ресурсы
#     res_rows: List[ProductionResource] = db.query(ProductionResource).all()
#     area_name_map: Dict[int, str] = {int(r.resource_id): str(r.resource_name or "") for r in res_rows}
#
#     # --- Дополнительно: определения доминирующего участка по всем стадиям и список заказов на день ---
#     # Карта (order_id -> карта часов по участкам)
#     hours_by_area_by_order: Dict[int, Dict[Optional[int], float]] = {}
#     for s in st_all:
#         try:
#             oid = int(s.order_id)
#             aid = int(s.area_id) if s.area_id is not None else None
#             h = float(getattr(s, "hours", 0.0) or 0.0)
#             m = hours_by_area_by_order.setdefault(oid, {})
#             m[aid] = float(m.get(aid, 0.0) + h)
#         except Exception:
#             continue
#
#     # Доминирующий участок заказа (по сумме часов всех стадий)
#     dom_area_by_order: Dict[int, Optional[int]] = {}
#     for oid, amap in hours_by_area_by_order.items():
#         best_aid: Optional[int] = None
#         best_h = -1.0
#         for aid, hh in amap.items():
#             if aid is not None and float(hh) > best_h:
#                 best_h = float(hh)
#                 best_aid = int(aid)
#         dom_area_by_order[int(oid)] = best_aid
#
#     # Заказы, назначенные на этот день (по bucket_date), даже если на самом деле часов в этот день нет
#     ord_day_q = db.query(PlannedOrder).filter(
#         PlannedOrder.run_id == int(run_id),
#         PlannedOrder.bucket_type == "daily",
#         PlannedOrder.bucket_date == day,
#     )
#     orders_for_day: List[PlannedOrder] = ord_day_q.all()
#
#     from collections import defaultdict as _dd
#     groups: Dict[int, Dict[str, Any]] = _dd(lambda: {
#         "area_id": None,
#         "area_name": "",
#         "orders": [],
#         "norm_sum_hours": 0.0,
#         "sum_qty": 0.0,
#         "cap_overload_hours": 0.0,
#         "hours_available_day": 0.0,
#         "cap_overload_percent": None,
#     })
#
#     # Индексы по order_id
#     ord_by_id: Dict[int, PlannedOrder] = {int(o.order_id): o for o in orders}
#     processed_oids: Set[int] = set()
#
#     for s in day_stages:
#         if s.area_id is None:
#             continue
#         aid = int(s.area_id)
#         if area_id is not None and aid != int(area_id):
#             continue
#         oid = int(s.order_id)
#         o = ord_by_id.get(oid)
#         if not o:
#             continue
#         processed_oids.add(oid)
#
#         total_hours = float(ord_total_hours.get(oid, 0.0) or 0.0)
#         qty = float(getattr(o, "qty", 0.0) or 0.0)
#         npu = float(total_hours / qty) if qty > 1e-12 else None
#         # Expanded fallback: also when computed npu is ~0 (total_hours==0 with qty>0)
#         if npu is None or (npu is not None and npu <= 1e-12):
#             try:
#                 npu_fb = float(fallback_npu_day.get(int(o.item_id), 0.0) or 0.0)
#                 if npu_fb > 0.0:
#                     npu = npu_fb
#             except Exception:
#                 pass
#         hours_today = float(getattr(s, "hours", 0.0) or 0.0)
#         # Реальный выпуск в штуках по часам дня
#         qty_today_real = (hours_today / npu) if (npu and npu > 1e-12) else 0.0
#         # Требование: на день показывать только целые штуки, округляя вверх
#         qty_today_int = int(math.ceil(qty_today_real)) if (npu and npu > 1e-12) else 0
#         # Остаток нормочасов до полного изделия добавляем в общий норматив группы в шапке
#         remainder_hours = 0.0
#         if (npu and npu > 1e-12) and qty_today_int > 0:
#             try:
#                 remainder_hours = max(0.0, float(qty_today_int) - float(qty_today_real)) * float(npu)
#             except Exception:
#                 remainder_hours = 0.0
#
#         it = item_map.get(int(o.item_id))
#         unit_guid = getattr(it, "unit", None) if it else None
#         unit_rec = unit_map.get(str(unit_guid)) if unit_guid else None
#         unit_disp = unit_guid
#         if unit_rec:
#             unit_disp = (
#                 getattr(unit_rec, "short_name", None)
#                 or getattr(unit_rec, "unit_name", None)
#                 or getattr(unit_rec, "unit_code", None)
#                 or unit_guid
#             )
#
#         g = groups[aid]
#         g["area_id"] = aid
#         g["area_name"] = area_name_map.get(aid, f"Участок #{aid}")
#         # Если на выбранный день по заказу часов 0 — трактуем как «перегруз на день»:
#         # показываем полный объём заказа в display_*, ставим флаг overload, добавляем часы в cap_overload_hours
#         overloaded_today = float(hours_today) <= 1e-9
#         row = {
#             "order_id": int(o.order_id),
#             "agg_key": f"{int(o.item_id)}|{unit_disp or ''}",
#             "item_id": int(o.item_id),
#             "item_name": getattr(it, "item_name", None) if it else None,
#             "item_article": getattr(it, "item_article", None) if it else None,
#             "unit": unit_disp,
#             "qty": float(qty_today_int),
#             "norm_hours_total": float(hours_today),
#             "norm_hours_per_unit": float(npu) if (npu is not None) else None,
#         }
#         if overloaded_today:
#             # Показываем полный объём заказа и суммарные нормо‑часы:
#             # если по заказу нет часов вообще (total_hours≈0), используем fallback-норму на единицу.
#             try:
#                 nhpu = float(npu) if (npu is not None) else 0.0
#             except Exception:
#                 nhpu = 0.0
#             display_total = float(total_hours) if float(total_hours) > 1e-9 else float(qty) * float(nhpu)
#             row["display_qty"] = float(qty)
#             row["display_norm_hours_total"] = float(display_total)
#             row["overload"] = True
#             try:
#                 g["cap_overload_hours"] = float(g.get("cap_overload_hours", 0.0) or 0.0) + float(display_total or 0.0)
#             except Exception:
#                 pass
#         else:
#             row["overload"] = False
#         g["orders"].append(row)
#         # В общий норматив группы добавляем фактические часы дня + остаток до целого изделия
#         g["norm_sum_hours"] += float(hours_today) + float(remainder_hours)
#         # Суммарный выпуск за день — целыми штуками
#         g["sum_qty"] += float(qty_today_int)
#
#     # Добавляем заказы, назначенные на этот день, но не попавшие в day_stages (hours_today == 0)
#     for o in orders_for_day:
#         try:
#             oid = int(o.order_id)
#         except Exception:
#             continue
#         if oid in processed_oids:
#             continue
#
#         # Определяем доминирующий участок по всем стадиям заказа
#         dom_aid = dom_area_by_order.get(oid)
#         if dom_aid is None:
#             continue
#         if area_id is not None and int(dom_aid) != int(area_id):
#             continue
#
#         total_hours = float(ord_total_hours.get(oid, 0.0) or 0.0)
#         qty_full = float(getattr(o, "qty", 0.0) or 0.0)
#         # Норма на единицу (если нет — из фолбэка по сумме операций спецификации)
#         npu2 = float(total_hours / qty_full) if qty_full > 1e-12 else None
#         if npu2 is None or (npu2 is not None and npu2 <= 1e-12):
#             try:
#                 npu_fb2 = float(fallback_npu_day.get(int(o.item_id), 0.0) or 0.0)
#                 if npu_fb2 > 0.0:
#                     npu2 = npu_fb2
#             except Exception:
#                 pass
#
#         it2 = item_map.get(int(o.item_id))
#         unit_guid2 = getattr(it2, "unit", None) if it2 else None
#         unit_rec2 = unit_map.get(str(unit_guid2)) if unit_guid2 else None
#         unit_disp2 = unit_guid2
#         if unit_rec2:
#             unit_disp2 = (
#                 getattr(unit_rec2, "short_name", None)
#                 or getattr(unit_rec2, "unit_name", None)
#                 or getattr(unit_rec2, "unit_code", None)
#                 or unit_guid2
#             )
#
#         g2 = groups[int(dom_aid)]
#         g2["area_id"] = int(dom_aid)
#         g2["area_name"] = area_name_map.get(int(dom_aid), f"Участок #{int(dom_aid)}")
#         # Заказ на этот день, но часов в этот день нет — показываем «0 за день», а в display_* — полный объём
#         g2["orders"].append({
#             "order_id": int(o.order_id),
#             "agg_key": f"{int(o.item_id)}|{unit_disp2 or ''}",
#             "item_id": int(o.item_id),
#             "item_name": getattr(it2, "item_name", None) if it2 else None,
#             "item_article": getattr(it2, "item_article", None) if it2 else None,
#             "unit": unit_disp2,
#             "qty": 0.0,  # за день 0
#             "norm_hours_total": 0.0,  # за день 0
#             "norm_hours_per_unit": float(npu2) if (npu2 is not None) else None,
#             # Расширения для отображения «перегруза»
#             "display_qty": float(qty_full),
#             # Если по заказу нет часов (не был расписан), оцениваем суммарные нормо‑часы через норму на ед.
#             "display_norm_hours_total": float(total_hours) if float(total_hours) > 1e-9 else float(qty_full) * float(npu2 or 0.0),
#             "overload": True,
#         })
#         # Учитываем невыполненный объём как перегруз по участку: используем оценку display_norm_hours_total
#         try:
#             add_ov = float(total_hours) if float(total_hours) > 1e-9 else float(qty_full) * float(npu2 or 0.0)
#             g2["cap_overload_hours"] = float(g2.get("cap_overload_hours", 0.0) or 0.0) + float(add_ov or 0.0)
#         except Exception:
#             pass
#         # norm_sum_hours/sum_qty считаются по фактическим часам дня; здесь ничего не добавляем
#
#     # Индикаторы перегруза за день
#     if groups:
#         aids = list(groups.keys())
#         capq = db.query(CapacityLoad).filter(
#             CapacityLoad.run_id == int(run_id),
#             CapacityLoad.bucket_type == "daily",
#             CapacityLoad.bucket_date == day,
#             CapacityLoad.area_id.in_(aids),
#         )
#         for r in capq.all():
#             g = groups.get(int(r.area_id))
#             if g:
#                 g["cap_overload_hours"] += float(r.overload_hours or 0.0)
#
#     # Дополнительно: интегрируем недопланированные объёмы из предупреждений CAPACITY_SHORTAGE на выбранную дату
#     try:
#         run_rec = db.query(PlanningRun).filter(PlanningRun.run_id == int(run_id)).first()
#     except Exception:
#         run_rec = None
#     capacity_shortages: List[Dict[str, Any]] = []
#     if run_rec is not None:
#         try:
#             for w in (run_rec.warnings or []):
#                 try:
#                     if str(w.get("code") or "") != "CAPACITY_SHORTAGE":
#                         continue
#                     nd = (w.get("need_date") or "")[:10]
#                     if nd != day.isoformat():
#                         continue
#                     capacity_shortages.append(w)
#                 except Exception:
#                     continue
#         except Exception:
#             capacity_shortages = []
#
#     if capacity_shortages:
#         # Справочник соответствия вида производства -> ресурсы
#         rpk_rows = db.query(ResourceProductionKind).all()
#         pk_to_resources: Dict[int, List[int]] = {}
#         for rpk in rpk_rows:
#             try:
#                 pk_to_resources.setdefault(int(rpk.production_kind_id), []).append(int(rpk.resource_id))
#             except Exception:
#                 continue
#
#         for w in capacity_shortages:
#             try:
#                 iid = int(w.get("item_id"))
#             except Exception:
#                 iid = None
#             try:
#                 pkid = int(w.get("production_kind_id")) if w.get("production_kind_id") is not None else None
#             except Exception:
#                 pkid = None
#             try:
#                 aid_warn = int(w.get("resource_id")) if w.get("resource_id") is not None else None
#             except Exception:
#                 aid_warn = None
#
#             # Определяем участок
#             target_aid: Optional[int] = aid_warn
#             if target_aid is None and pkid is not None:
#                 cand = pk_to_resources.get(int(pkid)) or []
#                 if cand:
#                     target_aid = int(sorted(cand)[0])
#
#             if target_aid is None:
#                 # не удалось определить участок — пропускаем
#                 continue
#
#             # Обеспечиваем наличие группы
#             g = groups[int(target_aid)]
#             if g.get("area_id") is None:
#                 g["area_id"] = int(target_aid)
#                 g["area_name"] = area_name_map.get(int(target_aid), f"Участок #{int(target_aid)}")
#
#             # Данные строки
#             unmet_qty = float(w.get("unmet_qty", 0.0) or 0.0)
#             unmet_hours = float(w.get("unmet_norm_hours", 0.0) or 0.0)
#             npu_w = float(w.get("norm_hours_per_unit", 0.0) or 0.0)
#
#             # Достаём справочник номенклатуры/ЕИ
#             itw = item_map.get(int(iid)) if iid is not None else None
#             unit_guid_w = getattr(itw, "unit", None) if itw else None
#             unit_rec_w = unit_map.get(str(unit_guid_w)) if unit_guid_w else None
#             unit_disp_w = unit_guid_w
#             if unit_rec_w:
#                 unit_disp_w = (
#                     getattr(unit_rec_w, "short_name", None)
#                     or getattr(unit_rec_w, "unit_name", None)
#                     or getattr(unit_rec_w, "unit_code", None)
#                     or unit_guid_w
#                 )
#
#             g["orders"].append({
#                 "order_id": None,
#                 "agg_key": f"{int(iid) if iid is not None else 0}|{unit_disp_w or ''}",
#                 "item_id": int(iid) if iid is not None else 0,
#                 "item_name": getattr(itw, "item_name", None) if itw else None,
#                 "item_article": getattr(itw, "item_article", None) if itw else None,
#                 "unit": unit_disp_w,
#                 "qty": 0.0,
#                 "norm_hours_total": 0.0,
#                 "norm_hours_per_unit": float(npu_w) if npu_w > 0.0 else None,
#                 "display_qty": float(unmet_qty),
#                 "display_norm_hours_total": float(unmet_hours),
#                 "overload": True,
#             })
#             try:
#                 g["cap_overload_hours"] = float(g.get("cap_overload_hours", 0.0) or 0.0) + float(unmet_hours or 0.0)
#             except Exception:
#                 pass
#
#     # Post-process groups: compute available hours for the day, overload percent and display fields
#     def _is_workday(d: date) -> bool:
#         return d.weekday() <= 4
#
#     res_by_id_local: Dict[int, ProductionResource] = {int(r.resource_id): r for r in res_rows}
#
#     for aid, g in groups.items():
#         res = res_by_id_local.get(int(aid))
#         hours_available_day = 0.0
#         if res is not None and _is_workday(day):
#             daily_hours = float(getattr(res, "daily_work_hours", 8.0) or 8.0)
#             power_coeff = float(getattr(res, "capacity", 1.0) or 1.0)
#             hours_available_day = daily_hours * power_coeff
#         g["hours_available_day"] = float(hours_available_day)
#         if hours_available_day > 1e-9:
#             g["cap_overload_percent"] = float((g.get("cap_overload_hours") or 0.0) / hours_available_day * 100.0)
#         else:
#             g["cap_overload_percent"] = None
#
#         # Не форсируем перегруз для всех строк группы.
#         # Сохраняем ранее вычисленные флаги overload на уровне строк:
#         # - строки с 0 часами на день (назначены на день) уже имеют overload=True и display_*,
#         # - синтетические CAPACITY_SHORTAGE строки тоже overload=True,
#         # - обычные строки с часами на день — overload=False.
#         for row in g.get("orders", []):
#             if "overload" not in row:
#                 row["overload"] = False
#
#     return {"groups": list(groups.values())}
def generate_shortage_report(db: Session, run_id: int) -> Dict[str, Any]:
    """
    Generates an XLSX report for component shortages based on planning run warnings.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io
        import base64
    except ImportError:
        raise RuntimeError("openpyxl is required for XLSX export. Please install it.")

    run = db.query(PlanningRun).filter(PlanningRun.run_id == run_id).first()
    if not run:
        raise RuntimeError(f"Run {run_id} not found")

    warnings = run.warnings or []
    
    # New branch: detailed component-caused shortages per parent item
    comp_warnings = [w for w in (warnings or []) if str(w.get("code") or "") == "COMPONENT_SHORTAGE"]
    if comp_warnings:
        # Collect parent and component item ids
        parent_ids: List[int] = []
        component_ids: List[int] = []
        for w in comp_warnings:
            try:
                pid = int(w.get("item_id"))
                parent_ids.append(pid)
            except Exception:
                pass
            try:
                cid = int(w.get("component_id"))
                component_ids.append(cid)
            except Exception:
                pass

        # Preload dictionaries for names and codes
        all_ids = list({i for i in (parent_ids + component_ids) if i is not None})
        items = db.query(Item).filter(Item.item_id.in_(all_ids)).all() if all_ids else []
        imap: Dict[int, Item] = {int(it.item_id): it for it in items}
        # Preload units
        unit_guids = {getattr(it, "unit", None) for it in items if getattr(it, "unit", None)}
        units = db.query(Unit).filter(Unit.unit_ref1c.in_(list(unit_guids))).all() if unit_guids else []
        unit_map: Dict[str, Unit] = {str(u.unit_ref1c): u for u in units}

        # Build child->parents map from PeggingLink for this run (to recover parent when warning lacks item_id)
        links = db.query(PeggingLink).filter(PeggingLink.run_id == int(run_id)).all()
        child_to_parents: Dict[int, Set[int]] = {}
        for ln in links or []:
            try:
                ch = int(getattr(ln, "child_item_id"))
                pr = getattr(ln, "parent_item_id")
                if pr is not None:
                    child_to_parents.setdefault(ch, set()).add(int(pr))
            except Exception:
                continue

        # Build XLSX
        import io, base64
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except Exception as e:
            raise RuntimeError(f"openpyxl is required for XLSX export. Please install it. {e}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Дефицит по компонентам"
        headers = [
            "Код изделия",
            "Изделие",
            "Артикул изделия",
            "ЕИ",
            "Запрошено, шт",
            "Возможный выпуск, шт (по компоненту)",
            "Дефицит изделия, шт",
            "Компонент",
            "Код компонента",
            "Артикул компонента",
            "ЕИ компонента",
            "Лимитирующий",
        ]
        ws.append(headers)
        # Bold header
        for cell in ws[1]:
            cell.font = Font(bold=True)
        # Row fills
        from openpyxl.styles import PatternFill
        green_fill = PatternFill(fill_type="solid", start_color="FFC6EFCE")  # light green
        yellow_fill = PatternFill(fill_type="solid", start_color="FFFFF2CC") # light yellow

        # Collect raw rows for grouping and sorting
        # Tuple: (parent_id, parent_code, parent_name, parent_article, parent_unit, req, mp, shortage_parent, comp_name, comp_code, comp_article, comp_unit)
        raw_rows: List[Tuple[int, str, str, str, str, float, float, float, str, str, str, str]] = []
        for w in comp_warnings:
            # Component id is mandatory for this report; skip if missing
            try:
                cid = int(w.get("component_id"))
            except Exception:
                continue

            # Parents set: from warning.item_id if present; otherwise via PeggingLink mapping
            parents: List[Optional[int]] = []
            pid_val = w.get("item_id")
            if pid_val is not None:
                try:
                    parents = [int(pid_val)]
                except Exception:
                    parents = []
            if not parents:
                parents = sorted(list(child_to_parents.get(int(cid), set()))) or [None]

            c = imap.get(int(cid))
            comp_unit = unit_map.get(getattr(c, "unit", None))
            comp_unit_code = getattr(comp_unit, "short_name", None) or getattr(comp_unit, "unit_code", None) or getattr(comp_unit, "unit_name", None)

            # Quantities
            try:
                req = float(w.get("requested_qty", 0.0) or 0.0)
            except Exception:
                req = 0.0
            try:
                mp = float(w.get("max_producible_from_component", 0.0) or 0.0)
            except Exception:
                mp = 0.0
            try:
                shortage_parent = float(w.get("shortage_parent_qty", max(0.0, req - mp)))
            except Exception:
                shortage_parent = max(0.0, req - mp)

            for pid in parents:
                p = imap.get(int(pid)) if (pid is not None) else None
                unit = unit_map.get(getattr(p, "unit", None))
                unit_code = getattr(unit, "short_name", None) or getattr(unit, "unit_code", None) or getattr(unit, "unit_name", None)
                p_code = getattr(p, "item_code", "") if p else ""
                raw_rows.append((
                    int(pid) if (pid is not None) else 0,
                    str(p_code or ""),
                    getattr(p, "item_name", "") if p else "",
                    getattr(p, "item_article", "") if p else "",
                    str(unit_code or ""),
                    req,
                    mp,
                    shortage_parent,
                    getattr(c, "item_name", "") if c else "",
                    getattr(c, "item_code", "") if c else "",
                    getattr(c, "item_article", "") if c else "",
                    str(comp_unit_code or ""),
                ))

        # Group by parent item (pid, item_name, item_article, unit) with deduplication:
        # req_max = max(requested), mp_min = min(possible across components)
        # shortage = req_max - mp_min
        from collections import defaultdict
        grouped: Dict[Tuple[int, str, str, str], Dict[str, Any]] = {}
        for pid, p_code, p_name, p_article, p_unit, req, mp, shortage_parent, c_name, c_code, c_article, c_unit in raw_rows:
            # Skip unknown parents (avoid empty green rows)
            if (pid is None or int(pid) == 0) and not p_name:
                continue
            key = (int(pid), p_name, p_article, p_unit)
            if key not in grouped:
                grouped[key] = {
                    "req_max": 0.0,
                    "mp_min": None,
                    "comp_map": {}  # key -> {name, code, article, unit, req_max, mp_min}
                }
            g = grouped[key]
            # Parent metrics
            if float(req or 0.0) > float(g["req_max"] or 0.0):
                g["req_max"] = float(req or 0.0)
            if g["mp_min"] is None or float(mp or 0.0) < float(g["mp_min"] or 0.0):
                g["mp_min"] = float(mp or 0.0)
            # Component deduplication by code (fallback to name|unit)
            comp_key = str(c_code or f"{c_name}|{c_unit}")
            cm = g["comp_map"].get(comp_key)
            if cm is None:
                g["comp_map"][comp_key] = {
                    "name": c_name,
                    "code": c_code,
                    "article": c_article,
                    "unit": c_unit,
                    "req_max": float(req or 0.0),
                    "mp_min": float(mp or 0.0),
                }
            else:
                if float(req or 0.0) > float(cm["req_max"]):
                    cm["req_max"] = float(req or 0.0)
                if float(mp or 0.0) < float(cm["mp_min"]):
                    cm["mp_min"] = float(mp or 0.0)
        
        # Sort by computed shortage descending
        def _item_shortage(v: Dict[str, Any]) -> float:
            reqx = float(v.get("req_max", 0.0) or 0.0)
            mpx = float(v.get("mp_min", 0.0) or 0.0) if v.get("mp_min") is not None else 0.0
            return max(0.0, reqx - mpx)
        sorted_items = sorted(grouped.items(), key=lambda x: _item_shortage(x[1]), reverse=True)

        # Append grouped data with coloring: parent rows green, component rows yellow
        total_rows = 0
        grand_total_shortage = 0.0
        for (pid, p_name, p_article, p_unit), data in sorted_items:
            req_par = float(data.get("req_max", 0.0) or 0.0)
            mp_par = float(data.get("mp_min", 0.0) or 0.0) if data.get("mp_min") is not None else 0.0
            shortage_par = max(0.0, req_par - mp_par)
            # Skip parents without shortage to avoid clutter and empty greens
            if shortage_par <= 1e-9:
                continue
            # Parent item code from dictionary
            p_obj = imap.get(int(pid)) if pid else None
            p_code = getattr(p_obj, "item_code", "") if p_obj else ""
    
            # Parent row
            ws.append([
                p_code,
                p_name,
                p_article,
                p_unit,
                req_par,
                mp_par,
                shortage_par,
                "", "", "", "",
                "",  # Limiting flag (only for component rows)
            ])
            # Style parent row (green + bold)
            pr = ws.max_row
            from openpyxl.styles import Font as _Font
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=pr, column=col)
                cell.fill = green_fill
                cell.font = _Font(bold=True)
            total_rows += 1
            grand_total_shortage += shortage_par
    
            # Component rows (deduped)
            comp_list = list(data.get("comp_map", {}).values())
            # Determine limiting mp (minimum across components)
            min_mp = None
            for cm in comp_list:
                v = float(cm.get("mp_min", 0.0) or 0.0)
                if min_mp is None or v < min_mp:
                    min_mp = v
            for cm in comp_list:
                req = float(cm.get("req_max", 0.0) or 0.0)
                mp = float(cm.get("mp_min", 0.0) or 0.0)
                is_lim = (min_mp is not None) and (abs(mp - min_mp) <= 1e-9)
                ws.append([
                    p_code, p_name, p_article, p_unit,  # parent columns filled for context
                    req,
                    mp,
                    "",  # shortage at component row — пусто
                    cm.get("name", ""),
                    cm.get("code", ""),
                    cm.get("article", ""),
                    cm.get("unit", ""),
                    "Да" if is_lim else "",
                ])
                # Style component row (yellow)
                rr = ws.max_row
                for col in range(1, len(headers) + 1):
                    ws.cell(row=rr, column=col).fill = yellow_fill
                total_rows += 1
    
        # Summary row with overall shortage
        ws.append(["", "ИТОГО дефицит, шт", "", "", "", "", grand_total_shortage, "", "", "", "", ""])
        sr = ws.max_row
        for col in range(1, len(headers) + 1):
            ws.cell(row=sr, column=col).font = Font(bold=True)

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = "A2"

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        b64 = base64.b64encode(bio.read()).decode("utf-8")
        return {
            "status": "ok",
            "format": "xlsx",
            "data_base64": b64,
            "filename": f"mrp_shortage_report_run_{run_id}.xlsx",
            "total_rows": int(total_rows),
        }

    # Fallback to previous behavior (capacity/stock generic shortages)
    shortage_warnings = [
        w for w in warnings
        if w.get("code") in ("INSUFFICIENT_STOCK_FOR_PRODUCTION", "CAPACITY_SHORTAGE")
    ]

    # --- Logging for debug ---
    import logging
    logger = logging.getLogger(__name__)
    logger.warning(f"Found {len(shortage_warnings)} shortage warnings to process for run {run_id}.")
    for w in shortage_warnings:
        logger.warning(f"Processing warning: {w}")
    # --- End Logging ---

    if not shortage_warnings:
        return {"status": "ok", "message": "No shortages found.", "total_rows": 0}

    # --- Data Aggregation ---
    
    # item_id -> {shortage_qty, required_for: set(parent_item_name)}
    shortage_map: Dict[int, Dict[str, Any]] = defaultdict(lambda: {"shortage_qty": 0.0, "required_for": set()})
    
    for w in shortage_warnings:
        try:
            item_id = int(w.get("item_id"))
            
            # For INSUFFICIENT_STOCK_FOR_PRODUCTION
            if w.get("code") == "INSUFFICIENT_STOCK_FOR_PRODUCTION":
                shortage_qty = float(w.get("shortage_qty", 0.0))
                parent_item_name = w.get("parent_item_name")
                shortage_map[item_id]["shortage_qty"] += shortage_qty
                if parent_item_name:
                    shortage_map[item_id]["required_for"].add(str(parent_item_name))

            # For CAPACITY_SHORTAGE, the "unmet_qty" is for the parent item,
            # but it implies a shortage of its components. We need to find the components.
            elif w.get("code") == "CAPACITY_SHORTAGE":
                # This warning is for a parent item that couldn't be produced.
                # We need to find its direct children (components) and list them as potential shortages.
                # The actual quantity is harder to determine without a full BOM explosion for the unmet part.
                # For now, we will just list the components that are needed for the parent.
                parent_item_id = int(w.get("item_id"))
                parent_item_name = w.get("item_name")
                
                # Fetch possibly multiple default specifications for the parent and aggregate their components
                spec_rows = db.query(DefaultSpecification.spec_id).filter(DefaultSpecification.item_id == parent_item_id).all()
                spec_ids: List[int] = []
                for row in (spec_rows or []):
                    try:
                        sid = int(getattr(row, "spec_id", row[0]))
                        spec_ids.append(sid)
                    except Exception:
                        continue
                if spec_ids:
                    components = db.query(SpecComponent).filter(SpecComponent.spec_id.in_(spec_ids)).all()
                    for comp in components:
                        # Quantity for components is unknown from capacity shortage; mark as 0 but record dependency
                        shortage_map[int(comp.item_id)]["shortage_qty"] += 0
                        if parent_item_name:
                            shortage_map[int(comp.item_id)]["required_for"].add(str(parent_item_name))

        except (ValueError, TypeError):
            continue

    item_ids = list(shortage_map.keys())
    if not item_ids:
        return {"status": "ok", "message": "No valid shortage items found.", "total_rows": 0}

    items = db.query(Item).filter(Item.item_id.in_(item_ids)).all()
    item_map = {i.item_id: i for i in items}

    # --- XLSX Generation ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицит Комплектующих"

    headers = ["Наименование", "Артикул", "Дефицит, кол-во", "Требуется для"]
    ws.append(headers)
    
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    total_rows = 0
    for item_id, data in sorted(shortage_map.items(), key=lambda x: x[0]):
        item = item_map.get(item_id)
        if not item:
            continue
        
        required_for_str = ", ".join(sorted(list(data["required_for"])))
        
        row_data = [
            item.item_name,
            item.item_article,
            data["shortage_qty"] if data["shortage_qty"] > 0 else "н/д",
            required_for_str
        ]
        ws.append(row_data)
        total_rows += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    b64 = base64.b64encode(bio.read()).decode("utf-8")

    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": b64,
        "filename": f"mrp_shortage_report_run_{run_id}.xlsx",
        "total_rows": total_rows,
    }