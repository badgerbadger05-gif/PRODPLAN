from __future__ import annotations

from typing import Any, Dict, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
import logging
logger = logging.getLogger("specification")
import base64
import io

from ..database import get_db

from app.utils.numeric import to_float as _to_float
from app.services.specification_service import (
    # helpers still used for response shaping in the router (e.g. xlsx export)
    _round_qty,
    _round_time,
    _build_units_map,
    _item_payload,
    _build_full_tree,
    _walk_nodes,
    _normal_filter_value,
    _safe_filename_part,
    # endpoint-backing services
    build_tree_children,
    build_tree_root_node,
    build_tree_debug,
    build_debug_report,
    build_units_debug_report,
    build_full_debug,
    search_items,
    build_flattened,
    build_where_used,
    build_quality,
    get_item_by_code_or_id,
)

router = APIRouter(prefix="/v1/specification", tags=["specification"])


# ------- endpoint

@router.get("/tree")
def get_specification_tree(
    item_code: Optional[str] = Query(None, description="Код изделия (альтернатива item_id/item_ref1c)"),
    item_id: Optional[int] = Query(None, description="ID изделия (альтернатива item_code/item_ref1c)"),
    item_ref1c: Optional[str] = Query(None, description="GUID изделия (Ref_Key из 1С, альтернатива item_code/item_id)"),
    root_qty: float = Query(1.0, description="Количество корневого изделия для расчёта"),
    parent_id: Optional[str] = Query(None, description="Идентификатор узла (для ленивой подгрузки детей)"),
    depth: int = Query(0, ge=0, le=2, description="Глубина разворота (0 - только корень, 1 - корень + дети)"),
    debug: bool = Query(False, description="Возвращать диагностическую информацию в meta.debug"),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Возвращает узлы спецификации (дерево) для QTable в режиме tree.

    Режимы:
      - Корень (без parent_id): возвращает 1 узел типа 'item' по item_code|item_id.
      - Дети узла (с parent_id): возвращает список дочерних 'item' и 'operation'.

    Идентификаторы узлов:
      - item:{item_id}:{tree_qty}
      - op:{spec_operation_id}:{parent_item_id}:{parent_tree_qty}
    """
    try:
        logger.info(f"[spec.tree] request parent_id={parent_id} item_code={item_code} item_id={item_id} root_qty={root_qty} depth={depth}")
        units_map = _build_units_map(db)
        if parent_id:
            # Lazy-load children for given node
            nodes = build_tree_children(db, parent_id, units_map)
            if nodes is None:
                return {"nodes": [], "meta": {"parentId": parent_id, "mode": "children"}}
            return {
                "nodes": nodes,
                "meta": {
                    "parentId": parent_id,
                    "mode": "children",
                }
            }

        # Root node case
        if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
            logger.error("[spec.tree] missing all of item_code, item_id, item_ref1c")
            raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")

        item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
        if not item:
            logger.error(f"[spec.tree] item not found item_code={item_code} item_id={item_id}")
            raise HTTPException(status_code=404, detail="Item not found")

        r_qty = _to_float(root_qty, 1.0)
        node = build_tree_root_node(db, item, root_qty, depth, units_map)

        meta: Dict[str, Any] = {
            "rootId": node["id"],
            "requested": {
                "item_code": item_code,
                "item_id": int(item.item_id),
                "root_qty": _round_qty(r_qty, 3),
                "depth": int(depth or 0),
            }
        }

        if debug:
            meta["debug"] = build_tree_debug(db, item, node)

        resp = {
            "nodes": [node],
            "meta": meta
        }
        logger.info(f"[spec.tree] root response children={len(node.get('children', []))} hasChildren={node.get('hasChildren')}")
        return resp

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[spec.tree] error: {e}")
        raise HTTPException(status_code=500, detail=f"Specification tree error: {e}")
# --- DEBUG endpoint: quick diagnostics without reading server logs
@router.get("/debug")
def get_specification_debug(
    item_code: Optional[str] = Query(None, description="Код изделия (альтернатива item_id/item_ref1c)"),
    item_id: Optional[int] = Query(None, description="ID изделия (альтернатива item_code/item_ref1c)"),
    item_ref1c: Optional[str] = Query(None, description="GUID изделия (Ref_Key, альтернатива item_code/item_id)"),
    root_qty: float = Query(1.0, description="Количество корневого изделия для расчёта"),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Диагностика разрешения спецификации и построения первого уровня детей.
    Удобно открыть в браузере:
      http://localhost:8000/api/v1/specification/debug?item_code=КОД&root_qty=1

    Возвращает:
      {
        "item": { id, code, name },
        "default_spec_id": int|null,
        "resolved_spec_id": int|null,
        "used_fallback": bool,
        "components_count": int,
        "operations_count": int,
        "children_count": int,
        "children_sample": [ { id, type, name, operationName, stageName } ... up to 10 ]
      }
    """
    try:
        logger.info(f"[spec.debug] request item_code={item_code} item_id={item_id} root_qty={root_qty}")

        if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
            raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")

        item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")

        return build_debug_report(db, item, root_qty)

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[spec.debug] error: {e}")
        raise HTTPException(status_code=500, detail=f"Specification debug error: {e}")
@router.get("/units-debug")
def get_units_debug(
    item_code: Optional[str] = Query(None, description="Код изделия для проверки его ЕИ"),
    item_id: Optional[int] = Query(None, description="ID изделия для проверки его ЕИ"),
    item_ref1c: Optional[str] = Query(None, description="GUID изделия (Ref_Key) для проверки его ЕИ"),
    unit_guid: Optional[str] = Query(None, description="GUID ЕИ (Ref_Key) для прямой проверки"),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Диагностика: что выгружено в таблицу units и как резолвится единица измерения у изделия/в узле спецификации.

    Возвращает:
      {
        "units_total": int,
        "units_sample": [ { unit_ref1c, unit_code, unit_name, short_name, iso_code, base_unit_ref1c, ratio, precision } ... up to 10 ],
        "check": {
          "unit_guid": "...",
          "unit_row": {...} | null
        },
        "item": {
          "id": int,
          "code": str,
          "name": str,
          "unit_guid": str|null,
          "mapped_label": str|null,
          "unit_row": {...} | null
        } | null
      }
    """
    return build_units_debug_report(db, item_code, item_id, item_ref1c, unit_guid)
# ------- full tree (recursive) -------


@router.get("/full")
def get_specification_full(
    item_code: Optional[str] = Query(None, description="Код изделия (альтернатива item_id/item_ref1c)"),
    item_id: Optional[int] = Query(None, description="ID изделия (альтернатива item_code/item_ref1c)"),
    item_ref1c: Optional[str] = Query(None, description="GUID изделия (Ref_Key, альтернатива item_code/item_id)"),
    root_qty: float = Query(1.0, description="Количество корневого изделия для расчёта"),
    max_depth: int = Query(15, ge=1, le=50, description="Максимальная глубина разворота дерева"),
    debug: bool = Query(False, description="Возвращать диагностическую информацию в meta.debug"),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    """
    Полное дерево спецификации (с операциями) одним запросом.
    Формат узлов полностью совместим с /v1/specification/tree.

    Пример:
      GET /api/v1/specification/full?item_code=XXX&amp;root_qty=1&amp;max_depth=15
    """
    try:
        logger.info(f"[spec.full] request item_code={item_code} item_id={item_id} root_qty={root_qty} max_depth={max_depth}")

        if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
            raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")

        item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
        if not item:
            raise HTTPException(status_code=404, detail="Item not found")

        units_map = _build_units_map(db)
        root_node = _build_full_tree(
            db=db,
            root_item=item,
            root_qty=_to_float(root_qty, 1.0),
            units_map=units_map,
            max_depth=int(max_depth or 15),
        )

        meta: Dict[str, Any] = {
            "rootId": root_node.get("id"),
            "requested": {
                "item_code": item_code,
                "item_id": int(item.item_id),
                "root_qty": _round_qty(_to_float(root_qty, 1.0), 3),
                "max_depth": int(max_depth or 15),
            }
        }

        if debug:
            meta["debug"] = build_full_debug(db, item)

        logger.info(f"[spec.full] built tree root_id={meta['rootId']}")
        return {"nodes": [root_node], "meta": meta}

    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[spec.full] error: {e}")
        raise HTTPException(status_code=500, detail=f"Specification full error: {e}")


@router.get("/search")
def search_specification_items(
    q: str = Query("", description="Поиск по артикулу, коду, названию или GUID"),
    has_spec: Optional[bool] = Query(None, description="Фильтр по наличию разрешаемой спецификации"),
    quality: Optional[str] = Query(None, description="Фильтр качества: no_spec|multiple_defaults"),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    term = str(q or "").strip()
    rows = search_items(db, term, has_spec, quality, int(limit))
    return {"items": rows, "meta": {"q": term, "count": len(rows), "limit": int(limit)}}


@router.get("/flattened")
def get_specification_flattened(
    item_code: Optional[str] = Query(None),
    item_id: Optional[int] = Query(None),
    item_ref1c: Optional[str] = Query(None),
    root_qty: float = Query(1.0),
    max_depth: int = Query(15, ge=1, le=50),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
        raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")
    item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return build_flattened(db, item, root_qty, int(max_depth or 15))


@router.get("/where-used")
def get_specification_where_used(
    item_code: Optional[str] = Query(None),
    item_id: Optional[int] = Query(None),
    item_ref1c: Optional[str] = Query(None),
    max_depth: int = Query(8, ge=1, le=25),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
        raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")
    item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return build_where_used(db, item, int(max_depth))


@router.get("/quality")
def get_specification_quality(
    item_code: Optional[str] = Query(None),
    item_id: Optional[int] = Query(None),
    item_ref1c: Optional[str] = Query(None),
    max_depth: int = Query(15, ge=1, le=50),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
        raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")
    item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    return build_quality(db, item, int(max_depth or 15))


@router.get("/export")
def export_specification_xlsx(
    item_code: Optional[str] = Query(None),
    item_id: Optional[int] = Query(None),
    item_ref1c: Optional[str] = Query(None),
    root_qty: float = Query(1.0),
    max_depth: int = Query(20, ge=1, le=50),
    replenishment_method: Optional[str] = Query(None, description="Optional exact filter by replenishment method"),
    db: Session = Depends(get_db),
) -> Dict[str, Any]:
    if item_code is None and item_id is None and (item_ref1c is None or str(item_ref1c).strip() == ""):
        raise HTTPException(status_code=400, detail="Either item_code, item_id or item_ref1c is required")

    item = get_item_by_code_or_id(db, item_code=item_code, item_id=item_id, item_ref1c=item_ref1c)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}")

    units_map = _build_units_map(db)
    root_payload = _item_payload(db, item, units_map)
    root_node = _build_full_tree(
        db=db,
        root_item=item,
        root_qty=_to_float(root_qty, 1.0),
        units_map=units_map,
        max_depth=int(max_depth or 20),
    )
    flat_nodes = _walk_nodes([root_node])
    method_filter = _normal_filter_value(replenishment_method)
    if method_filter:
        tree_nodes = [
            node for node in flat_nodes
            if _normal_filter_value(node.get("replenishmentMethod")) == method_filter
        ]
    else:
        tree_nodes = flat_nodes

    flattened_payload = build_flattened(
        db,
        item,
        _to_float(root_qty, 1.0),
        int(max_depth or 20),
    )
    flattened_rows = list(flattened_payload.get("items") or [])
    if method_filter:
        flattened_rows = [
            row for row in flattened_rows
            if _normal_filter_value(row.get("replenishment_method")) == method_filter
        ]

    where_used_rows = build_where_used(
        db,
        item,
        10,
    ).get("items") or []
    quality_rows = build_quality(
        db,
        item,
        int(max_depth or 20),
    ).get("issues") or []

    wb = Workbook()
    header_fill = PatternFill(fill_type="solid", fgColor="FFD9EAF7")
    title_fill = PatternFill(fill_type="solid", fgColor="FF4F81BD")
    title_font = Font(bold=True, color="FFFFFFFF")

    def setup_sheet(ws, title: str, headers: List[str], rows: List[List[Any]]) -> None:
        ws.title = title
        root_title = f"{root_payload.get('item_article') or root_payload.get('item_code')} · {root_payload.get('item_name') or ''}"
        ws.append([root_title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(headers)))
        ws.cell(row=1, column=1).fill = title_fill
        ws.cell(row=1, column=1).font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="left")
        ws.append(headers)
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A3"
        last_col = get_column_letter(max(1, len(headers)))
        ws.auto_filter.ref = f"A2:{last_col}{max(2, ws.max_row)}"
        for col_idx in range(1, len(headers) + 1):
            values = [headers[col_idx - 1]]
            for row in rows:
                values.append(row[col_idx - 1] if col_idx - 1 < len(row) else "")
            width = min(max(max((len(str(v)) for v in values), default=10) + 2, 10), 55)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def item_path(node: Dict[str, Any]) -> str:
        path = node.get("path") or []
        return " / ".join(str(p.get("article") or p.get("name") or "") for p in path if p.get("article") or p.get("name"))

    tree_headers = ["Уровень", "Тип", "Номенклатура/операция", "Артикул", "Этап", "Метод", "Кол-во", "Ед./Норма", "Итого", "Проблемы", "Путь"]
    tree_rows: List[List[Any]] = []
    for node in tree_nodes:
        is_operation = node.get("type") == "operation"
        stage = node.get("stage") or {}
        computed = node.get("computed") or {}
        operation = node.get("operation") or {}
        tree_rows.append([
            int(node.get("level") or 0),
            "Операция" if is_operation else "Номенклатура",
            str(operation.get("name") or "Операция") if is_operation else str(node.get("name") or ""),
            str(node.get("article") or ""),
            str(stage.get("name") or "") if isinstance(stage, dict) else "",
            "" if is_operation else str(node.get("replenishmentMethod") or ""),
            "" if node.get("qtyPerParent") is None else _round_qty(_to_float(node.get("qtyPerParent")), 3),
            f"{_round_time(_to_float(node.get('timeNormNh')), 3)} н/ч" if is_operation else str(node.get("unit") or ""),
            "" if computed.get("treeQty") is None else _round_qty(_to_float(computed.get("treeQty")), 3),
            ", ".join(str(w) for w in node.get("warnings") or []),
            item_path(node),
        ])
    setup_sheet(wb.active, "Дерево", tree_headers, tree_rows)

    flat_headers = ["Компонент", "Код", "Артикул", "Метод", "Итого", "Ед.", "Вхождений", "Уровни", "Этапы", "Проблемы"]
    flat_rows = [[
        str(row.get("name") or ""),
        str(row.get("item_code") or ""),
        str(row.get("article") or ""),
        str(row.get("replenishment_method") or ""),
        _round_qty(_to_float(row.get("total_qty")), 3),
        str(row.get("unit") or ""),
        int(row.get("occurrences") or 0),
        ", ".join(str(v) for v in row.get("levels") or []),
        ", ".join(str(v) for v in row.get("stages") or []),
        ", ".join(str(v) for v in row.get("warnings") or []),
    ] for row in flattened_rows]
    setup_sheet(wb.create_sheet(), "Плоская", flat_headers, flat_rows)

    where_headers = ["Родитель", "Код", "Артикул", "Спецификация", "Уровень вверх", "Кол-во", "Итого к цели", "Этап"]
    where_rows = []
    for row in where_used_rows:
        parent = row.get("parent") or {}
        spec = row.get("spec") or {}
        stage = row.get("stage") or {}
        where_rows.append([
            str(parent.get("item_name") or ""),
            str(parent.get("item_code") or ""),
            str(parent.get("item_article") or ""),
            str(spec.get("spec_name") or spec.get("spec_code") or ""),
            int(row.get("level_up") or 0),
            _round_qty(_to_float(row.get("qty_per_parent")), 3),
            _round_qty(_to_float(row.get("total_qty_to_target")), 3),
            str(stage.get("name") or "") if isinstance(stage, dict) else "",
        ])
    setup_sheet(wb.create_sheet(), "Где используется", where_headers, where_rows)

    quality_headers = ["Проблема", "Серьезность", "Сообщение", "Артикул", "Номенклатура", "Спецификация"]
    quality_export_rows = []
    for issue in quality_rows:
        issue_item = issue.get("item") or {}
        quality_export_rows.append([
            str(issue.get("code") or ""),
            str(issue.get("severity") or ""),
            str(issue.get("message") or ""),
            str(issue_item.get("item_article") or issue_item.get("item_code") or ""),
            str(issue_item.get("item_name") or ""),
            "" if issue.get("spec_id") is None else int(issue.get("spec_id")),
        ])
    setup_sheet(wb.create_sheet(), "Качество", quality_headers, quality_export_rows)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"specification_{_safe_filename_part(str(root_payload.get('item_article') or root_payload.get('item_code') or item.item_id))}.xlsx"
    return {
        "status": "ok",
        "format": "xlsx",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": filename,
        "data_base64": base64.b64encode(bio.read()).decode("utf-8"),
        "total_rows": len(tree_rows),
        "flat_rows": len(flat_rows),
    }
