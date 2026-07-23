from __future__ import annotations

import base64
import io
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional

from sqlalchemy.orm import Session

from .planning_service import (
    get_run_purchases_grouped_by_category,
    get_run_rework_grouped_by_category,
)


def _require_openpyxl():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover - defensive import guard
        raise RuntimeError(f"openpyxl not available: {exc}")
    return Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter


def _autosize_columns(ws: Any, get_column_letter: Callable[[int], str], max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            try:
                text = str(value)
            except Exception:
                text = ""
            max_length = max(max_length, max((len(chunk) for chunk in text.splitlines()), default=0))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, 12), 48)


def _encode_workbook(wb: Any) -> str:
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return base64.b64encode(bio.read()).decode("utf-8")


def _group_title(group_name: Optional[str]) -> str:
    return f"Товарная группа: {(group_name or '').strip() or 'Без товарной группы'}"


def _sort_groups_for_export(groups: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return sorted(
        list(groups or []),
        key=lambda group: (
            1 if group.get("group_id") is None else 0,
            (group.get("group_name") or "").strip().lower(),
        ),
    )


def _rework_component_status(order: Dict[str, Any]) -> str:
    if bool(order.get("component_blocked")):
        return "Заблокирован"
    if bool(order.get("component_partial")):
        return "Частично ограничен"
    return "Без ограничений"


def _build_grouped_workbook(
    *,
    sheet_title: str,
    headers: List[str],
    groups: List[Dict[str, Any]],
    row_builder: Callable[[Dict[str, Any]], List[Any]],
) -> tuple[str, int]:
    Workbook, Alignment, Border, Font, PatternFill, Side, get_column_letter = _require_openpyxl()

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    group_fill = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")
    header_fill = PatternFill(fill_type="solid", start_color="D9E2F3", end_color="D9E2F3")
    zebra_fill = PatternFill(fill_type="solid", start_color="F7F9FC", end_color="F7F9FC")

    total_rows = 0
    for group in groups:
        ws.append([_group_title(group.get("group_name"))])
        group_row_idx = ws.max_row
        ws.merge_cells(start_row=group_row_idx, start_column=1, end_row=group_row_idx, end_column=len(headers))
        group_cell = ws.cell(row=group_row_idx, column=1)
        group_cell.font = Font(bold=True, color="FFFFFF")
        group_cell.alignment = Alignment(horizontal="left", vertical="center")
        group_cell.fill = group_fill
        group_cell.border = thin_border

        ws.append(headers)
        header_row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = thin_border

        orders = list(group.get("orders") or [])
        for idx, order in enumerate(orders):
            row_values = row_builder(order)
            ws.append(row_values)
            row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                if idx % 2 == 0:
                    cell.fill = zebra_fill
            total_rows += 1

        ws.append([])

    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        ws.delete_rows(1, 1)

    if ws.max_row > 0:
        ws.freeze_panes = "A2"
    _autosize_columns(ws, get_column_letter, len(headers))
    return _encode_workbook(wb), total_rows


def export_purchases_results_xlsx(
    db: Session,
    run_id: int,
    *,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> Dict[str, Any]:
    grouped = get_run_purchases_grouped_by_category(
        db=db,
        run_id=int(run_id),
        item_id=item_id,
        date_from=date_from,
        date_to=date_to,
        limit=1000,
        offset=0,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    groups = _sort_groups_for_export((grouped or {}).get("groups", []) or [])
    return export_purchases_snapshot_groups_xlsx(
        run_id=int(run_id),
        groups=groups,
        total_groups=int((grouped or {}).get("total_groups", 0) or 0),
    )


def export_purchases_snapshot_groups_xlsx(
    *,
    run_id: int,
    groups: List[Dict[str, Any]],
    total_groups: Optional[int] = None,
) -> Dict[str, Any]:
    """Render already persisted snapshot payloads; performs no DB reads."""
    groups = _sort_groups_for_export(groups)

    headers = [
        "Наименование",
        "Артикул",
        "Поставщик",
        "Категория",
        "Количество",
        "ЕИ",
        "Дата потребности",
        "Дата заказа",
        "Срок пополнения, дн.",
        "Поставщик 1С",
        "Пометка",
    ]

    def build_row(order: Dict[str, Any]) -> List[Any]:
        return [
            order.get("item_name") or "",
            order.get("item_article") or "",
            order.get("supplier_name") or order.get("supplier_ref1c") or "",
            order.get("category_name") or "",
            float(order.get("qty") or 0.0),
            order.get("unit") or "",
            order.get("need_date") or "",
            order.get("order_date") or "",
            int(order.get("lead_time_days") or 0),
            order.get("supplier_ref1c") or "",
            order.get("badge") or "",
        ]

    data_base64, total_rows = _build_grouped_workbook(
        sheet_title="Purchases",
        headers=headers,
        groups=groups,
        row_builder=build_row,
    )

    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": data_base64,
        "filename": f"mrp_purchases_run_{int(run_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "total_rows": int(total_rows),
        "total_groups": int(len(groups) if total_groups is None else total_groups),
    }


def export_rework_results_xlsx(
    db: Session,
    run_id: int,
    *,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> Dict[str, Any]:
    grouped = get_run_rework_grouped_by_category(
        db=db,
        run_id=int(run_id),
        item_id=item_id,
        date_from=date_from,
        date_to=date_to,
        limit=1000,
        offset=0,
        sort_by=sort_by,
        sort_dir=sort_dir,
    )
    groups = _sort_groups_for_export((grouped or {}).get("groups", []) or [])
    return export_rework_snapshot_groups_xlsx(
        run_id=int(run_id),
        groups=groups,
        total_groups=int((grouped or {}).get("total_groups", 0) or 0),
    )


def export_rework_snapshot_groups_xlsx(
    *,
    run_id: int,
    groups: List[Dict[str, Any]],
    total_groups: Optional[int] = None,
) -> Dict[str, Any]:
    """Render already persisted snapshot payloads; performs no DB reads."""
    groups = _sort_groups_for_export(groups)

    headers = [
        "Наименование",
        "Артикул",
        "Количество",
        "Запрошено",
        "К плану",
        "ЕИ",
        "Дата потребности",
        "Дата запуска",
        "Срок пополнения, дн.",
        "Спецификация",
        "Лимит по комплектующим",
        "Статус комплектующих",
    ]

    def build_row(order: Dict[str, Any]) -> List[Any]:
        return [
            order.get("item_name") or "",
            order.get("item_article") or "",
            float(order.get("qty") or 0.0),
            float(order.get("requested_qty") or 0.0),
            float(order.get("planned_qty") or 0.0),
            order.get("unit") or "",
            order.get("need_date") or "",
            order.get("order_date") or "",
            int(order.get("lead_time_days") or 0),
            order.get("spec_name") or order.get("spec_code") or "",
            float(order.get("component_limit") or 0.0) if order.get("component_limit") is not None else "",
            _rework_component_status(order),
        ]

    data_base64, total_rows = _build_grouped_workbook(
        sheet_title="Rework",
        headers=headers,
        groups=groups,
        row_builder=build_row,
    )

    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": data_base64,
        "filename": f"mrp_rework_run_{int(run_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        "total_rows": int(total_rows),
        "total_groups": int(len(groups) if total_groups is None else total_groups),
    }
