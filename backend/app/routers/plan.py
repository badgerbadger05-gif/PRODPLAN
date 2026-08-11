from typing import List, Dict, Any, Literal, Optional
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from pydantic import BaseModel, ConfigDict, Field

from ..database import get_db
from ..services.stage_directory import fetch_stages

from ..services.work_calendar_service import get_planning_anchor_date

from ..services.planning_service import (
    list_planning_runs,
    # Retained as test guards: tests/services/test_mrp_result_snapshot.py
    # monkeypatches these names on this module to prove the snapshot routes
    # never fall back to the legacy live getters.
    get_run_purchases,
    get_run_capacity,
    # Config management
    list_planning_configs,
    create_planning_config_version,
    activate_planning_config_version,
    get_active_planning_config_full
)
from ..schemas import ProductionGroupedResponse, PurchaseCategoryGroupedResponse, ReworkGroupedResponse
from ..services.mrp_result_snapshot import (
    read_mrp_result_manifest,
    read_mrp_result_rows,
)
from ..services.mrp_result_export import (
    export_purchases_snapshot_groups_xlsx,
    export_rework_snapshot_groups_xlsx,
)
from ..services.one_c_purchase_order_export import export_planned_purchases_to_1c
from ..services.period_plan_service import (
    add_item_to_period_plan,
    bulk_upsert_period_plan_lines,
    create_period_plan,
    create_mrp_snapshot_for_plan,
    delete_period_plan,
    delete_period_plan_item,
    fix_period_plan,
    get_period_plan,
    get_period_plan_execution_journal,
    get_period_plan_matrix,
    list_mrp_runs_for_plan,
    list_period_plans,
    repair_duplicate_plan_snapshots,
    update_period_plan_header,
)

router = APIRouter(prefix="/v1/plan", tags=["plan"])


def _read_all_mrp_snapshot_rows(
    db: Session,
    run_id: int,
    row_kind: str,
    *,
    snapshot_id: Optional[int],
    item_id: Optional[int] = None,
    root_item_id: Optional[int] = None,
    supplier_ref1c: Optional[str] = None,
    category_id: Optional[int] = None,
    category_ref1c: Optional[str] = None,
    area_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_dir: Optional[str] = None,
) -> tuple[list[dict[str, Any]], int]:
    """Read all export rows while pinning every page to one snapshot id."""
    rows: list[dict[str, Any]] = []
    offset = 0
    pinned_snapshot_id = snapshot_id
    while True:
        result = read_mrp_result_rows(
            db=db,
            run_id=int(run_id),
            row_kind=row_kind,
            snapshot_id=pinned_snapshot_id,
            item_id=item_id,
            root_item_id=root_item_id,
            supplier_ref1c=supplier_ref1c,
            category_id=category_id,
            category_ref1c=category_ref1c,
            area_id=area_id,
            date_from=date_from,
            date_to=date_to,
            limit=5000,
            offset=offset,
            sort_dir=sort_dir or "asc",
        )
        resolved_id = result.get("snapshot_id")
        if resolved_id is None:
            raise HTTPException(
                status_code=503,
                detail={
                    "code": "mrp_result_snapshot_required",
                    "run_id": int(run_id),
                    "truth_status": result.get("truth_status") or "unavailable",
                    "truth_reason": result.get("truth_reason")
                    or "MRP result snapshot is unavailable",
                    "rows": [],
                },
            )
        if pinned_snapshot_id is None:
            pinned_snapshot_id = int(resolved_id)
        elif int(resolved_id) != int(pinned_snapshot_id):
            raise HTTPException(
                status_code=409,
                detail={"code": "mrp_result_snapshot_changed"},
            )
        page = list(result.get("rows") or [])
        rows.extend(page)
        offset += len(page)
        if not page or offset >= int(result.get("total") or 0):
            return rows, int(pinned_snapshot_id)


def _page_groups(
    groups: list[dict[str, Any]],
    *,
    limit: int,
    offset: int,
    identity: dict[str, Any],
) -> dict[str, Any]:
    effective_limit = max(1, min(int(limit or 100), 5000))
    effective_offset = max(0, int(offset or 0))
    return {
        "groups": groups[effective_offset : effective_offset + effective_limit],
        "total_groups": len(groups),
        "total_orders": sum(len(group.get("orders") or []) for group in groups),
        "limit": effective_limit,
        "offset": effective_offset,
        **identity,
    }


def _mrp_snapshot_identity(
    db: Session, run_id: int, snapshot_id: int
) -> dict[str, Any]:
    manifest = read_mrp_result_manifest(
        db=db, run_id=int(run_id), snapshot_id=int(snapshot_id)
    )
    if manifest.get("snapshot_id") is None:
        raise HTTPException(
            status_code=503,
            detail={
                "code": "mrp_result_snapshot_required",
                "run_id": int(run_id),
                "truth_status": manifest.get("truth_status") or "unavailable",
                "truth_reason": manifest.get("truth_reason"),
            },
        )
    if int(manifest["snapshot_id"]) != int(snapshot_id):
        raise HTTPException(
            status_code=409, detail={"code": "mrp_result_snapshot_changed"}
        )
    return {
        "snapshot_id": int(manifest["snapshot_id"]),
        "ledger_generation": int(manifest["ledger_generation"]),
        "cutoff": str(manifest["cutoff"]),
        "truth_status": str(manifest["truth_status"]),
        "truth_reason": manifest.get("truth_reason"),
    }


def _category_groups_from_snapshot(
    rows: list[dict[str, Any]], *, rework: bool
) -> list[dict[str, Any]]:
    groups: dict[Optional[int], dict[str, Any]] = {}
    for source in rows:
        row = dict(source)
        group_id = row.get("category_id")
        if group_id is not None:
            group_id = int(group_id)
        group_name = (row.get("category_name") or "").strip()
        if not group_name:
            group_name = "Без товарной группы"
        group = groups.setdefault(
            group_id,
            {
                "group_id": group_id,
                "group_name": group_name,
                "orders": [],
                "sum_qty": 0.0,
                **(
                    {
                        "sum_requested_qty": 0.0,
                        "sum_planned_qty": 0.0,
                        "blocked_orders": 0,
                        "partial_orders": 0,
                    }
                    if rework
                    else {}
                ),
            },
        )
        group["orders"].append(row)
        group["sum_qty"] += float(row.get("qty") or 0)
        if rework:
            group["sum_requested_qty"] += float(row.get("requested_qty") or 0)
            group["sum_planned_qty"] += float(row.get("planned_qty") or 0)
            group["blocked_orders"] += int(bool(row.get("component_blocked")))
            group["partial_orders"] += int(bool(row.get("component_partial")))
    return sorted(
        groups.values(),
        key=lambda group: (
            1 if group.get("group_id") is None else 0,
            (group.get("group_name") or "").lower(),
        ),
    )


def _production_groups_from_snapshot(
    rows: list[dict[str, Any]],
    capacity_rows: list[dict[str, Any]],
    *,
    area_id: Optional[int] = None,
) -> list[dict[str, Any]]:
    groups: dict[Optional[int], dict[str, Any]] = {}
    for source in rows:
        row = dict(source)
        stages = list(row.get("stages") or [])
        main = max(stages, key=lambda stage: float(stage.get("hours") or 0), default={})
        main_area_id = main.get("area_id")
        if main_area_id is not None:
            main_area_id = int(main_area_id)
        if area_id is not None and main_area_id != int(area_id):
            continue
        area_name = (main.get("area_name") or "").strip() or "Без участка"
        group = groups.setdefault(
            main_area_id,
            {
                "area_id": main_area_id,
                "area_name": area_name,
                "orders": [],
                "norm_sum_hours": 0.0,
                "min_days_to_need": None,
                "cap_overload_hours": 0.0,
                "cap_overloaded_buckets": 0,
            },
        )
        row.setdefault(
            "agg_key", f"{int(row.get('item_id') or 0)}|{row.get('unit') or ''}"
        )
        row.setdefault("order_id", None)
        group["orders"].append(row)
        group["norm_sum_hours"] += float(row.get("norm_hours_total") or 0)

    for capacity in capacity_rows:
        capacity_area = capacity.get("area_id")
        if capacity_area is None:
            continue
        group = groups.get(int(capacity_area))
        if group is None:
            continue
        overload = float(capacity.get("overload_hours") or 0)
        group["cap_overload_hours"] += overload
        group["cap_overloaded_buckets"] += int(overload > 1e-9)
    return sorted(
        groups.values(), key=lambda group: (group.get("area_name") or "").lower()
    )


# Pydantic модели
class PlanningConfigCreate(BaseModel):
    config: Dict[str, Any]
    comment: Optional[str] = None
    created_by: Optional[str] = None
    activate: Optional[bool] = False


# ===== Period plans =====


class PeriodPlanCreateRequest(BaseModel):
    name: str
    period_from: str
    period_to: str
    created_by: Optional[str] = None
    comment: Optional[str] = None


class PeriodPlanLineEntry(BaseModel):
    item_id: int
    bucket_date: str
    qty: float


class PeriodPlanBulkUpsertRequest(BaseModel):
    entries: List[PeriodPlanLineEntry] = []


class PeriodPlanItemRequest(BaseModel):
    item_id: int


class PeriodPlanFixRequest(BaseModel):
    fixed_by: Optional[str] = None


class PeriodPlanMrpSnapshotRequest(BaseModel):
    # Transport hint only. The server chooses the deterministic refresh key.
    started_by: Optional[str] = None


class PeriodPlanRepairSnapshotsRequest(BaseModel):
    repaired_by: Optional[str] = None


class PeriodPlanUpdateRequest(BaseModel):
    name: Optional[str] = None
    period_from: Optional[str] = None
    period_to: Optional[str] = None
    comment: Optional[str] = None


class ExecutionJournalTruthMeta(BaseModel):
    model_config = ConfigDict(extra="forbid")

    accepted_at: str | None = None
    accepted_by: str | None = None
    truth_source: str | None = None
    unavailable_sections: list[str] | None = None
    unavailable_reason: str | None = None


class ExecutionJournalInformationLinkEvent(BaseModel):
    model_config = ConfigDict(extra="forbid")

    reservation_id: int
    url: str


class ExecutionJournalInformationLinks(BaseModel):
    model_config = ConfigDict(extra="forbid")

    reservation_events: list[ExecutionJournalInformationLinkEvent] = []


class ExecutionJournalLedgerEvent(BaseModel):
    model_config = ConfigDict(extra="forbid")

    event_id: int
    reservation_id: int
    sle_id: int | None = None
    fact_ref: str | None = None
    fact_line_ref: str | None = None
    match_rule: str | None = None


class ExecutionJournalLedgerLinks(BaseModel):
    model_config = ConfigDict(extra="forbid")

    item_id: int
    reservation_ids: list[int] = []
    events: list[ExecutionJournalLedgerEvent] = []


class ExecutionJournalWorkItem(BaseModel):
    model_config = ConfigDict(extra="forbid")

    type: Literal["production_order", "planned_order", "planned_purchase", "planned_rework"]
    qty: float
    product_id: int | None = None
    order_id: int | None = None
    order_number: str | None = None
    order_ref1c: str | None = None
    order_source: str | None = None
    one_c_opened: bool | None = None
    opened_at: str | None = None
    order_state: str | None = None
    purchase_id: int | None = None
    rework_id: int | None = None
    completed_qty: float | None = None
    remaining_qty: float | None = None
    need_date: str | None = None
    order_date: str | None = None
    lead_time_days: int | None = None
    forecast_date: str | None = None
    forecast_shift_days: int | None = None
    forecast_reason: str | None = None
    forecast_status: Literal["early", "on_time", "delayed", "critical", "unavailable"] | None = None


class ExecutionJournalRow(BaseModel):
    model_config = ConfigDict(extra="forbid")

    req_id: int
    item_id: int
    item_code: str
    item_name: str
    flow: Literal["production", "purchase", "rework"]
    bom_level: int
    gross_qty: float
    net_qty: float
    ordered_qty: float
    completed_qty: float | None
    coverage_pct: float | None
    remaining_qty: float | None
    work_items: list[ExecutionJournalWorkItem]
    status: Literal["net_zero", "covered", "partial", "ordered", "none", "execution_unavailable"]
    root_item_ids: list[int] = []
    information_links: ExecutionJournalInformationLinks
    reservation_ids: list[int] = []
    execution_events: list[dict[str, object]] = []
    execution_allocations: list[dict[str, object]] = []
    ledger_links: ExecutionJournalLedgerLinks | None = None
    item_article: str | None = None
    stock_qty: float | None = None
    covered_qty: float | None = None
    progress_base_qty: float | None = None
    execution_available: bool | None = None
    execution_unavailable_reason: str | None = None
    execution_source: str | None = None
    need_date: str | None = None
    status_label: str | None = None
    forecast_date: str | None = None
    forecast_shift_days: int | None = None
    forecast_reason: str | None = None
    forecast_status: Literal["early", "on_time", "delayed", "critical", "unavailable"] | None = None
    purchase_covered_qty: float | None = None
    purchase_to_order_qty: float | None = None
    unassigned_qty: float | None = None


class ExecutionJournalSummaryByFlow(BaseModel):
    model_config = ConfigDict(extra="forbid")

    completed_qty: float
    base_qty: float
    execution_pct: float | None
    available: bool = True
    total_base_qty: float | None = None
    confirmed_pct: float | None = None
    covered_pct: float | None = None
    to_order_pct: float | None = None
    purchase_covered_qty: float | None = None
    purchase_to_order_qty: float | None = None


class ExecutionJournalSummary(BaseModel):
    model_config = ConfigDict(extra="forbid")

    truth_status: str
    total_items: int
    execution_completed_qty: float | None = None
    execution_base_qty: float | None = None
    execution_available_base_qty: float | None = None
    execution_pct: float | None = None
    execution_confirmed_pct: float | None = None
    execution_partial: bool | None = None
    fully_covered: int | None = None
    partially_covered: int | None = None
    not_covered: int | None = None
    net_zero: int | None = None
    execution_by_flow: dict[str, ExecutionJournalSummaryByFlow] | None = None


class ExecutionJournalResponse(BaseModel):
    model_config = ConfigDict(extra="forbid")

    plan: dict[str, object]
    run_id: int
    rows: list[ExecutionJournalRow]
    summary: ExecutionJournalSummary
    total: int
    limit: int
    offset: int
    truth_status: str | None = None
    ledger_generation: str | int | None = None
    truth_generation_id: int | None = None
    cutoff: str | None = None
    truth_cutoff: str | None = None
    truth_meta: ExecutionJournalTruthMeta | dict[str, object] | None = None
    truth_reason: str | None = None
    reason: str | None = None
    facets: dict[str, list[int]] | None = None


class PurchaseOrder1CExportRequest(BaseModel):
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    purchase_ids: Optional[List[int]] = None
    dry_run: Optional[bool] = False
    # DEPRECATED: демо-гард записи в 1С удалён после go-live. Поле принимается
    # и игнорируется, чтобы существующие клиенты не получали 422.
    allow_production: Optional[bool] = False


# ===== Period plan routes =====

@router.get("/period-plans")
async def period_plans_list(
    status: Optional[str] = None,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    created_by: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    return list_period_plans(
        db,
        status=status,
        period_from=period_from,
        period_to=period_to,
        created_by=created_by,
        sort_by=sort_by,
        sort_dir=sort_dir,
        limit=limit,
        offset=offset,
    )


@router.post("/period-plans")
async def period_plans_create(
    req: PeriodPlanCreateRequest,
    db: Session = Depends(get_db),
):
    try:
        return create_period_plan(
            db,
            name=req.name,
            period_from=req.period_from,
            period_to=req.period_to,
            created_by=req.created_by,
            comment=req.comment,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/period-plans/{plan_id}")
async def period_plans_delete(plan_id: int, db: Session = Depends(get_db)):
    try:
        return delete_period_plan(db, plan_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/period-plans/{plan_id}")
async def period_plans_get(plan_id: int, db: Session = Depends(get_db)):
    try:
        return get_period_plan(db, plan_id)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.get("/period-plans/{plan_id}/matrix")
async def period_plans_matrix(plan_id: int, db: Session = Depends(get_db)):
    try:
        return get_period_plan_matrix(db, plan_id)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/period-plans/{plan_id}/items")
async def period_plans_add_item(
    plan_id: int,
    req: PeriodPlanItemRequest,
    db: Session = Depends(get_db),
):
    try:
        return add_item_to_period_plan(db, plan_id, req.item_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/period-plans/{plan_id}/items/{item_id}")
async def period_plans_delete_item(
    plan_id: int,
    item_id: int,
    db: Session = Depends(get_db),
):
    try:
        return delete_period_plan_item(db, plan_id, item_id)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/period-plans/{plan_id}/lines/bulk_upsert")
async def period_plans_bulk_upsert(
    plan_id: int,
    req: PeriodPlanBulkUpsertRequest,
    db: Session = Depends(get_db),
):
    try:
        payload = [e.model_dump() for e in (req.entries or [])]
        return bulk_upsert_period_plan_lines(db, plan_id, payload)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/period-plans/{plan_id}/fix")
async def period_plans_fix(
    plan_id: int,
    req: PeriodPlanFixRequest = PeriodPlanFixRequest(),
    db: Session = Depends(get_db),
):
    """Единственное действие «Зафиксировать»: атомарно фиксирует план и
    публикует его MRP-снимок одним поколением Ledger.

    Отдельной кнопки «MRP-снимок» в каноническом сценарии нет. Если снимок не
    опубликован, план остаётся `draft` (fail closed).
    """
    try:
        return fix_period_plan(db, plan_id, fixed_by=req.fixed_by)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/period-plans/{plan_id}")
async def period_plans_patch(
    plan_id: int,
    req: PeriodPlanUpdateRequest,
    db: Session = Depends(get_db),
):
    try:
        return update_period_plan_header(
            db,
            plan_id,
            name=req.name,
            period_from=req.period_from,
            period_to=req.period_to,
            comment=req.comment,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/period-plans/{plan_id}/runs")
async def period_plans_runs(
    plan_id: int,
    limit: int = 50,
    db: Session = Depends(get_db),
):
    try:
        return list_mrp_runs_for_plan(db, plan_id, limit=limit)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/period-plans/{plan_id}/mrp-snapshot")
async def period_plans_mrp_snapshot(
    plan_id: int,
    req: PeriodPlanMrpSnapshotRequest = PeriodPlanMrpSnapshotRequest(),
    db: Session = Depends(get_db),
):
    """Совместимый путь восстановления снимка уже зафиксированного плана.

    Основной сценарий — атомарный `POST /period-plans/{id}/fix`. Этот маршрут
    остаётся идемпотентным повтором: если у плана уже есть снимок в текущем
    принятом поколении, он возвращается как есть и ничего не форкается.
    Ключ формируется сервером автоматически из плана и текущего принятого
    поколения.
    """
    try:
        result = create_mrp_snapshot_for_plan(
            db,
            plan_id,
            started_by=req.started_by or "api",
        )
        db.commit()
        return result
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/period-plans/{plan_id}/repair-duplicate-snapshots")
async def period_plans_repair_duplicate_snapshots(
    plan_id: int,
    req: PeriodPlanRepairSnapshotsRequest = PeriodPlanRepairSnapshotsRequest(),
    db: Session = Depends(get_db),
):
    """Административный ремонт плана с несколькими текущими FIXED_SNAPSHOT.

    Старая гонка TOCTOU могла опубликовать два снимка одного плана. Такой план
    навсегда отвергается обычным путём («План имеет несколько текущих
    зафиксированных MRP-снимков»). Ремонт под тем же локом публикации
    детерминированно оставляет один прогон, остальные переводит в `SUPERSEDED`,
    после чего обычный путь снимка снова работает. Идемпотентен: для здорового
    плана возвращает `repaired=false` и ничего не меняет.
    """
    try:
        result = repair_duplicate_plan_snapshots(
            db, plan_id, repaired_by=req.repaired_by,
        )
        db.commit()
        return result
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get(
    "/period-plans/{plan_id}/execution-journal",
    response_model=ExecutionJournalResponse,
)
async def period_plans_execution_journal(
    plan_id: int,
    run_id: Optional[int] = None,
    root_item_id: Optional[int] = None,
    bom_level: Optional[int] = None,
    flow: Optional[str] = None,
    status: Optional[str] = None,
    include_net_zero: bool = True,
    sort_by: Literal["bom_level", "item_article", "item_code", "item_name", "flow", "gross_qty", "net_qty", "ordered_qty", "completed_qty", "remaining_qty", "need_date", "coverage_pct", "status"] = "bom_level",
    sort_dir: Literal["asc", "desc"] = "asc",
    limit: int = 100,
    offset: int = 0,
    db: Session = Depends(get_db),
) -> ExecutionJournalResponse:
    try:
        payload = get_period_plan_execution_journal(
            db,
            plan_id,
            run_id=run_id,
            root_item_id=root_item_id,
            bom_level=bom_level,
            flow=flow,
            status=status,
            include_net_zero=include_net_zero,
            sort_by=sort_by,
            sort_dir=sort_dir,
            limit=limit,
            offset=offset,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    return ExecutionJournalResponse.model_validate(payload)


@router.get("/anchor")
async def get_planning_anchor(
    db: Session = Depends(get_db),
):
    """Якорная дата для отображения планового окна.

    Семантика: первый НЕ закрытый рабочий день после последнего закрытого.
    """
    try:
        return get_planning_anchor_date(db=db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ===== MRP Planning API (runs and results) =====

# Retention & pinning DTOs
class RetentionCleanupRequest(BaseModel):
    older_than_days: Optional[int] = 30
    dry_run: Optional[bool] = False


class PinRequest(BaseModel):
    pinned: bool


# ===== Planning configuration management API =====

@router.get("/configs")
async def list_configs(
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Список версий конфигурации планирования (пагинация)"""
    try:
        return list_planning_configs(db=db, limit=limit, offset=offset)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/configs/active")
async def get_active_config(db: Session = Depends(get_db)):
    """Получить активную конфигурацию планирования (полный JSON снапшот)"""
    try:
        data = get_active_planning_config_full(db=db)
        return {"status": "ok", "config": data}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/configs")
async def create_config(
    req: PlanningConfigCreate,
    db: Session = Depends(get_db)
):
    """Создать новую версию конфигурации планирования (опционально сразу активировать)"""
    try:
        result = create_planning_config_version(
            db=db,
            config=req.config,
            comment=req.comment,
            created_by=req.created_by,
            activate=bool(req.activate or False),
        )
        return {"status": "ok", "created": result}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/configs/{config_id}/activate")
async def activate_config(
    config_id: int,
    db: Session = Depends(get_db)
):
    """Активировать указанную версию конфигурации (сняв active с предыдущей)"""
    try:
        result = activate_planning_config_version(db=db, config_id=int(config_id))
        return {"status": "ok", "activated": result}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))




@router.get("/runs")
async def get_planning_runs(
    limit: int = 50,
    offset: int = 0,
    db: Session = Depends(get_db)
):
    """Список прогонов планирования с краткой статистикой"""
    try:
        return list_planning_runs(db=db, limit=limit, offset=offset)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}")
async def get_planning_result_summary(
    run_id: int,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Сводка из сохранённого Ledger-bound снимка; расчёт не запускается."""
    try:
        return read_mrp_result_manifest(
            db=db, run_id=int(run_id), snapshot_id=snapshot_id
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/production")
async def get_planning_result_production(
    run_id: int,
    item_id: Optional[int] = None,
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Производственные обязательства из сохранённого снимка."""
    try:
        return read_mrp_result_rows(
            db=db,
            run_id=int(run_id),
            row_kind="production",
            snapshot_id=snapshot_id,
            item_id=item_id,
            root_item_id=root_item_id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset,
            sort_dir=sort_dir,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/results/{run_id}/production/grouped", response_model=ProductionGroupedResponse)
async def get_planning_result_production_grouped(
    run_id: int,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    area_id: Optional[int] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Группировка сохранённых производственных обязательств по участкам."""
    try:
        rows, snapshot_id = _read_all_mrp_snapshot_rows(
            db, int(run_id), "production",
            snapshot_id=None,
            item_id=item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )
        capacity_rows, _ = _read_all_mrp_snapshot_rows(
            db, int(run_id), "capacity",
            snapshot_id=snapshot_id,
            date_from=date_from,
            date_to=date_to,
        )
        groups = _production_groups_from_snapshot(
            rows, capacity_rows, area_id=area_id
        )
        return _page_groups(
            groups,
            limit=limit,
            offset=offset,
            identity=_mrp_snapshot_identity(db, int(run_id), snapshot_id),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/purchases")
async def get_planning_result_purchases(
    run_id: int,
    item_id: Optional[int] = None,
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    supplier_ref1c: Optional[str] = None,
    category_id: Optional[int] = None,
    category_ref1c: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Закупочные обязательства из сохранённого снимка."""
    try:
        return read_mrp_result_rows(
            db=db,
            run_id=int(run_id),
            row_kind="purchase",
            snapshot_id=snapshot_id,
            item_id=item_id,
            root_item_id=root_item_id,
            supplier_ref1c=supplier_ref1c,
            category_id=category_id,
            category_ref1c=category_ref1c,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset,
            sort_dir=sort_dir,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/purchases/grouped")
async def get_planning_result_purchases_grouped(
    run_id: int,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 1000,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Агрегированная выдача закупок из принятого снимка."""
    try:
        base_rows, snapshot_id = _read_all_mrp_snapshot_rows(
            db, int(run_id), "purchase",
            snapshot_id=None,
            date_from=date_from,
            date_to=date_to,
        )

        mapped = []
        for r in base_rows:
            try:
                iid = int(r.get("item_id"))
            except Exception:
                continue
            unit = r.get("unit")
            mapped.append(
                {
                    "agg_key": f"{iid}|{unit or ''}",
                    "item_id": iid,
                    "item_name": r.get("item_name"),
                    "item_article": r.get("item_article"),
                    "unit": unit,
                    "qty": float(r.get("qty") or 0.0),
                }
            )

        total = int(len(mapped))
        eff_limit = max(1, min(int(limit or 1000), 5000))
        eff_offset = max(0, int(offset or 0))
        page = mapped[eff_offset : eff_offset + eff_limit]
        return {
            "rows": page,
            "total": total,
            "limit": eff_limit,
            "offset": eff_offset,
            **_mrp_snapshot_identity(db, int(run_id), snapshot_id),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/rework")
async def get_planning_result_rework(
    run_id: int,
    item_id: Optional[int] = None,
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """Обязательства переработки из сохранённого снимка."""
    try:
        return read_mrp_result_rows(
            db=db,
            run_id=int(run_id),
            row_kind="rework",
            snapshot_id=snapshot_id,
            item_id=item_id,
            root_item_id=root_item_id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset,
            sort_dir=sort_dir,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/rework/grouped", response_model=ReworkGroupedResponse)
async def get_planning_result_rework_grouped(
    run_id: int,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Группированная выдача переработки из принятого снимка."""
    try:
        rows, snapshot_id = _read_all_mrp_snapshot_rows(
            db, int(run_id), "rework",
            snapshot_id=None,
            item_id=item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )
        groups = []
        if rows:
            groups = [{
                "group_id": None,
                "group_name": "Без товарной группы",
                "orders": rows,
                "sum_qty": sum(float(row.get("qty") or 0) for row in rows),
                "sum_requested_qty": sum(float(row.get("requested_qty") or 0) for row in rows),
                "sum_planned_qty": sum(float(row.get("planned_qty") or 0) for row in rows),
                "blocked_orders": sum(int(bool(row.get("component_blocked"))) for row in rows),
                "partial_orders": sum(int(bool(row.get("component_partial"))) for row in rows),
            }]
        return _page_groups(
            groups,
            limit=limit,
            offset=offset,
            identity=_mrp_snapshot_identity(db, int(run_id), snapshot_id),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/purchases/grouped-by-category", response_model=PurchaseCategoryGroupedResponse)
async def get_planning_result_purchases_grouped_by_category(
    run_id: int,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Закупки по сохранённой в снимке товарной группе."""
    try:
        rows, snapshot_id = _read_all_mrp_snapshot_rows(
            db, int(run_id), "purchase",
            snapshot_id=None,
            item_id=item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )
        return _page_groups(
            _category_groups_from_snapshot(rows, rework=False),
            limit=limit,
            offset=offset,
            identity=_mrp_snapshot_identity(db, int(run_id), snapshot_id),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/rework/grouped-by-category", response_model=ReworkGroupedResponse)
async def get_planning_result_rework_grouped_by_category(
    run_id: int,
    item_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    offset: int = 0,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Переработка по сохранённой в снимке товарной группе."""
    try:
        rows, snapshot_id = _read_all_mrp_snapshot_rows(
            db, int(run_id), "rework",
            snapshot_id=None,
            item_id=item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )
        return _page_groups(
            _category_groups_from_snapshot(rows, rework=True),
            limit=limit,
            offset=offset,
            identity=_mrp_snapshot_identity(db, int(run_id), snapshot_id),
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/capacity")
async def get_planning_result_capacity(
    run_id: int,
    area_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 200,
    offset: int = 0,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db)
):
    """Загрузка мощностей из сохранённого MRP-снимка."""
    try:
        return read_mrp_result_rows(
            db=db,
            run_id=int(run_id),
            row_kind="capacity",
            snapshot_id=snapshot_id,
            area_id=area_id,
            date_from=date_from,
            date_to=date_to,
            limit=limit,
            offset=offset,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/stages")
async def get_stages(db: Session = Depends(get_db)):
    """Получить список этапов производства"""
    try:
        return fetch_stages(db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# === Export endpoints for results ===


@router.get("/results/{run_id}/production/export")
async def export_planning_result_production(
    run_id: int,
    format: str = "csv",
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Экспорт результатов «Производство» в CSV или XLSX (base64).
    Колонки: Наименование, Артикул, Количество, Нормо-часы всего, Нормо-часы на ед., Дата потребности, Дата начала, Дата окончания, ЕИ
    """
    try:
        rows, resolved_snapshot_id = _read_all_mrp_snapshot_rows(
            db,
            int(run_id),
            "production",
            snapshot_id=snapshot_id,
            root_item_id=root_item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )

        headers = [
            "Наименование",
            "Артикул",
            "Количество",
            "Нормо-часы всего",
            "Нормо-часы на ед.",
            "Дата потребности",
            "Дата начала",
            "Дата окончания",
            "ЕИ",
            "Пометка",
        ]
        data_rows = []
        for r in rows:
            qty = float(r.get("qty") or 0.0)
            norm_total = float(r.get("norm_hours_total") or 0.0)
            norm_per_unit = r.get("norm_hours_per_unit")
            if norm_per_unit is None:
                norm_per_unit = (norm_total / qty) if qty > 0 else None
            data_rows.append(
                [
                    r.get("item_name") or "",
                    r.get("item_article") or "",
                    qty,
                    norm_total,
                    float(norm_per_unit) if norm_per_unit is not None else "",
                    r.get("need_date") or "",
                    r.get("start_date") or "",
                    r.get("finish_date") or "",
                    r.get("unit") or "",
                    r.get("badge") or "",
                ]
            )

        if (format or "csv").lower() == "xlsx":
            capacity_rows, _ = _read_all_mrp_snapshot_rows(
                db,
                int(run_id),
                "capacity",
                snapshot_id=resolved_snapshot_id,
                date_from=date_from,
                date_to=date_to,
            )
            groups = _production_groups_from_snapshot(rows, capacity_rows)

            import io, base64
            try:
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill, Font, Alignment
                from openpyxl.utils import get_column_letter
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"openpyxl not available: {e}")

            wb = Workbook()
            ws = wb.active
            ws.title = "Production"

            # Трекинг максимальной ширины контента по колонкам для псевдо-автоширины
            max_widths = {i: len(str(h)) for i, h in enumerate(headers, start=1)}

            def update_widths(values: list):
                for idx, val in enumerate(values, start=1):
                    text = "" if val is None else str(val)
                    # Учитываем переносы строк по наибольшей длине строки
                    length = max((len(line) for line in str(text).splitlines()), default=0)
                    if length > max_widths.get(idx, 0):
                        max_widths[idx] = length

            def style_header(row_idx: int):
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = Font(bold=True)

            def append_group_title(title: str):
                ws.append([title])
                r = ws.max_row
                # Подзаголовок на всю ширину таблицы
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
                cell = ws.cell(row=r, column=1)
                cell.font = Font(bold=True, color="FFFFFFFF")
                # Яркий синий фон для подзаголовка группы
                cell.fill = PatternFill(fill_type="solid", fgColor="FF4F81BD")
                cell.alignment = Alignment(horizontal="left")
                update_widths([title])

            if groups:
                # Для каждой группы добавляем подзаголовок с названием участка, затем шапку и строки
                for g in groups:
                    area_name = str(g.get("area_name") or f"ID {g.get('area_id') or ''}")
                    # Подзаголовок группы
                    append_group_title(f"Участок: {area_name}")
                    # Шапка колонок
                    ws.append(headers)
                    hdr_row = ws.max_row
                    style_header(hdr_row)
                    update_widths(headers)
                    orders = (g.get("orders", []) or [])
                    for o in orders:
                        qty = float(o.get("qty") or 0.0)
                        norm_total = float(o.get("norm_hours_total") or 0.0)
                        npu = o.get("norm_hours_per_unit")
                        if npu is None:
                            npu = (norm_total / qty) if qty > 0 else None
                        row_values = [
                            o.get("item_name") or "",
                            o.get("item_article") or "",
                            qty,
                            norm_total,
                            float(npu) if npu is not None else "",
                            "",  # Дата потребности (в агрегате может отсутствовать)
                            "",  # Дата начала
                            "",  # Дата окончания
                            o.get("unit") or "",
                            o.get("badge") or "",
                        ]
                        ws.append(row_values)
                        update_widths(row_values)
                    # Пустая строка между группами
                    ws.append([])
            else:
                # Фолбэк: плоский список без группировки
                ws.append(headers)
                hdr_row = ws.max_row
                style_header(hdr_row)
                update_widths(headers)
                for row in data_rows:
                    ws.append(row)
                    update_widths(row)

            # Установка ширины колонок в зависимости от контента (псевдо-автоширина)
            for col_idx in range(1, len(headers) + 1):
                letter = get_column_letter(col_idx)
                width = max_widths.get(col_idx, 10)
                # Коэффициент подбора ширины + небольшой запас, ограничения разумных пределов
                adjusted = min(max(width * 1.2 + 2, 12), 60)
                ws.column_dimensions[letter].width = adjusted

            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            b64 = base64.b64encode(bio.read()).decode("utf-8")
            return {
                "status": "ok",
                "format": "xlsx",
                "data_base64": b64,
                "filename": f"mrp_production_run_{run_id}.xlsx",
                "total_rows": len(data_rows) if not groups else sum(len((g.get("orders") or [])) for g in groups),
                "snapshot_id": resolved_snapshot_id,
            }
        else:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in data_rows:
                writer.writerow(row)
            return {
                "status": "ok",
                "format": "csv",
                "data": output.getvalue(),
                "filename": f"mrp_production_run_{run_id}.csv",
                "total_rows": len(data_rows),
                "snapshot_id": resolved_snapshot_id,
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
@router.get("/results/{run_id}/purchases/export")
async def export_planning_result_purchases(
    run_id: int,
    format: str = "csv",
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    supplier_ref1c: Optional[str] = None,
    category_id: Optional[int] = None,
    category_ref1c: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Экспорт результатов «Закупки» в CSV или XLSX (base64).
    Колонки: Наименование, Артикул, Количество, ЕИ
    """
    try:
        rows, resolved_snapshot_id = _read_all_mrp_snapshot_rows(
            db,
            int(run_id),
            "purchase",
            snapshot_id=snapshot_id,
            root_item_id=root_item_id,
            supplier_ref1c=supplier_ref1c,
            category_id=category_id,
            category_ref1c=category_ref1c,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )

        headers = ["Наименование", "Артикул", "Поставщик", "Категория", "Количество", "ЕИ", "Пометка"]
        data_rows = []
        for r in rows:
            data_rows.append([
                r.get("item_name") or "",
                r.get("item_article") or "",
                r.get("supplier_name") or r.get("supplier_ref1c") or "",
                r.get("category_name") or "",
                float(r.get("qty") or 0.0),
                r.get("unit") or "",
                r.get("badge") or "",
            ])

        if (format or "csv").lower() == "xlsx":
            groups = _category_groups_from_snapshot(rows, rework=False)
            result = export_purchases_snapshot_groups_xlsx(
                run_id=int(run_id),
                groups=groups,
            )
            result["snapshot_id"] = resolved_snapshot_id
            return result
        else:
            import io, csv
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in data_rows:
                writer.writerow(row)
            return {
                "status": "ok",
                "format": "csv",
                "data": output.getvalue(),
                "filename": f"mrp_purchases_run_{run_id}.csv",
                "total_rows": len(data_rows),
                "snapshot_id": resolved_snapshot_id,
            }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/results/{run_id}/purchases/export-to-1c")
async def export_planning_result_purchases_to_1c(
    run_id: int,
    req: PurchaseOrder1CExportRequest,
    db: Session = Depends(get_db),
):
    """
    Создать в 1С заказы поставщикам по результатам закупок MRP.
    Строки группируются по поставщику: один `Document_ЗаказПоставщику` на каждого поставщика.
    """
    try:
        return export_planned_purchases_to_1c(
            db=db,
            run_id=int(run_id),
            date_from=req.date_from,
            date_to=req.date_to,
            purchase_ids=req.purchase_ids,
            dry_run=bool(req.dry_run),
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class ReconcileRequest(BaseModel):
    dry_run: bool = False


class SpecificationRebaseRequest(BaseModel):
    dry_run: bool = False
    changed_spec_refs: List[str] = Field(default_factory=list)


class PendingSpecificationRebaseRequest(BaseModel):
    dry_run: bool = False


@router.post("/mrp/run/{run_id}/close")
async def close_mrp_run(
    run_id: int,
    req: ReconcileRequest = ReconcileRequest(),
    db: Session = Depends(get_db),
):
    """Явно закрыть плановый прогон (FIXED_SNAPSHOT -> CLOSED).

    Закрытие убирает активные резервы из рабочих очередей (release),
    требования и закупочные строки не перезаписывает. Доступно только для
    зафиксированных прогонах. Повторный вызов идемпотентен.
    """
    from ..services.period_plan_service import close_fixed_plan

    try:
        return close_fixed_plan(db, int(run_id), dry_run=bool(req.dry_run))
    except ValueError as e:
        detail = str(e)
        raise HTTPException(status_code=404 if "не найден" in detail else 400, detail=detail)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/mrp/run/{run_id}/rebase-specification")
async def rebase_mrp_run_for_specification(
    run_id: int,
    req: SpecificationRebaseRequest = SpecificationRebaseRequest(),
    db: Session = Depends(get_db),
):
    """Create a successor MRP only for the predecessor's unproduced roots."""
    from ..services.specification_mrp_rebase import (
        rebase_fixed_plan_remaining_roots,
    )

    try:
        return rebase_fixed_plan_remaining_roots(
            db,
            int(run_id),
            changed_spec_refs=req.changed_spec_refs,
            started_by=f"api:specification_rebase:{int(run_id)}",
            dry_run=bool(req.dry_run),
        )
    except ValueError as e:
        detail = str(e)
        raise HTTPException(
            status_code=404 if "не найден" in detail else 400,
            detail=detail,
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/mrp/specification-rebase/run-pending")
async def run_pending_specification_rebase(
    req: PendingSpecificationRebaseRequest = PendingSpecificationRebaseRequest(),
    db: Session = Depends(get_db),
):
    """Run/preview one durable automatic rebase item (benchmark entrypoint)."""
    from ..services.specification_rebase_worker import (
        run_one_pending_specification_rebase,
    )

    try:
        return run_one_pending_specification_rebase(
            db,
            dry_run=bool(req.dry_run),
            started_by="api:pending_specification_rebase",
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/results/{run_id}/rework/export")
async def export_planning_result_rework(
    run_id: int,
    format: str = "csv",
    root_item_id: Optional[int] = None,
    bucket_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = None,
    sort_dir: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    """
    Экспорт результатов «Переработка» в CSV или XLSX (base64).
    XLSX группируется по товарным группам.
    """
    try:
        rows, resolved_snapshot_id = _read_all_mrp_snapshot_rows(
            db,
            int(run_id),
            "rework",
            snapshot_id=snapshot_id,
            root_item_id=root_item_id,
            date_from=date_from,
            date_to=date_to,
            sort_dir=sort_dir,
        )

        headers = [
            "Наименование",
            "Артикул",
            "Количество",
            "Запрошено",
            "К плану",
            "ЕИ",
            "Дата потребности",
            "Дата запуска",
            "Срок пополнения, дн.",
            "Спецификация",
            "Лимит по комплектующим",
            "Статус комплектующих",
        ]
        data_rows = []
        for r in rows:
            if bool(r.get("component_blocked")):
                status = "Заблокирован"
            elif bool(r.get("component_partial")):
                status = "Частично ограничен"
            else:
                status = "Без ограничений"

            data_rows.append([
                r.get("item_name") or "",
                r.get("item_article") or "",
                float(r.get("qty") or 0.0),
                float(r.get("requested_qty") or 0.0),
                float(r.get("planned_qty") or 0.0),
                r.get("unit") or "",
                r.get("need_date") or "",
                r.get("order_date") or "",
                int(r.get("lead_time_days") or 0),
                r.get("spec_name") or r.get("spec_code") or "",
                float(r.get("component_limit") or 0.0) if r.get("component_limit") is not None else "",
                status,
            ])

        if (format or "csv").lower() == "xlsx":
            groups = _category_groups_from_snapshot(rows, rework=True)
            result = export_rework_snapshot_groups_xlsx(
                run_id=int(run_id),
                groups=groups,
            )
            result["snapshot_id"] = resolved_snapshot_id
            return result

        import io, csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        for row in data_rows:
            writer.writerow(row)
        return {
            "status": "ok",
            "format": "csv",
            "data": output.getvalue(),
            "filename": f"mrp_rework_run_{run_id}.csv",
            "total_rows": len(data_rows),
            "snapshot_id": resolved_snapshot_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
