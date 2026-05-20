"""
Экспорт учитываемых заказов поставщику в Excel (XLSX).
Данные берутся из БД (supplier_orders + supplier_order_items + items + suppliers).
"""

from __future__ import annotations

import base64
import io
from collections import OrderedDict
from datetime import datetime
from typing import Any, Dict

from sqlalchemy.orm import Session

from ..models import Supplier, SupplierOrder, SupplierOrderItem, Item
from .planning_service import SUPPLIER_ORDER_EXCLUDED_STATE_NAMES, _normalize_supplier_order_state_name


def _is_included_supplier_order(order: SupplierOrder) -> bool:
    if bool(getattr(order, "deletion_mark", False)):
        return False
    state_name = _normalize_supplier_order_state_name(getattr(order, "order_state_name", None))
    if not state_name and not str(getattr(order, "order_state_key", "") or "").strip():
        return False
    return state_name not in SUPPLIER_ORDER_EXCLUDED_STATE_NAMES


def export_supplier_orders_xlsx(db: Session) -> Dict[str, Any]:
    """
    Экспорт только тех заказов поставщику, которые учитываются MRP как ожидаемое поступление.
    Формат близок к экспорту заказов на производство: синие строки-группы по заказам
    и обычные строки деталей под ними. Колонки заказа при этом повторяются в деталях,
    чтобы Excel-автофильтр оставался полезным.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    rows = (
        db.query(SupplierOrder, SupplierOrderItem, Item, Supplier)
        .join(SupplierOrderItem, SupplierOrderItem.order_id == SupplierOrder.order_id)
        .join(Item, Item.item_id == SupplierOrderItem.item_id_ref)
        .outerjoin(Supplier, Supplier.supplier_id == SupplierOrder.supplier_id)
        .filter(SupplierOrder.deletion_mark.is_(False))
        .filter(SupplierOrderItem.remaining_qty > 0)
        .order_by(SupplierOrder.order_date, SupplierOrder.order_number, SupplierOrderItem.line_number)
        .all()
    )

    included_rows = [(order, row, item, supplier) for order, row, item, supplier in rows if _is_included_supplier_order(order)]

    grouped_rows: "OrderedDict[int, list[tuple[SupplierOrder, SupplierOrderItem, Item, Supplier]]]" = OrderedDict()
    for order, row, item, supplier in included_rows:
        grouped_rows.setdefault(int(order.order_id), []).append((order, row, item, supplier))

    headers = [
        "Заказ",
        "Дата заказа",
        "Статус",
        "Поставщик",
        "Номенклатура",
        "Артикул",
        "Код",
        "Строка",
        "Заказано",
        "Поступило",
        "Осталось",
        "Дата поступления",
        "Цена",
        "Сумма",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "ЗаказыПоставщику"

    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    order_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    order_font = Font(bold=True, size=11, color="FFFFFF")
    order_alignment = Alignment(horizontal="left", vertical="center")
    zebra_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = thin_border

    total_rows = 0
    row_num = 2
    for order_rows in grouped_rows.values():
        order, _, _, supplier = order_rows[0]
        order_date = order.order_date.strftime("%Y-%m-%d") if isinstance(order.order_date, datetime) else str(order.order_date or "")
        state_display = order.order_state_name or (f"ID: {str(order.order_state_key)[:8]}..." if order.order_state_key else "")
        supplier_name = getattr(supplier, "supplier_name", "") or ""
        order_header = f"Заказ №{order.order_number or ''} от {order_date}"
        if state_display:
            order_header += f" • {state_display}"
        if supplier_name:
            order_header += f" • {supplier_name}"

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        order_cell = ws.cell(row=row_num, column=1, value=order_header)
        order_cell.font = order_font
        order_cell.alignment = order_alignment
        order_cell.fill = order_fill
        order_cell.border = thin_border
        row_num += 1

        for order, order_item, item, supplier in order_rows:
            order_date = order.order_date.strftime("%Y-%m-%d") if isinstance(order.order_date, datetime) else str(order.order_date or "")
            state_display = order.order_state_name or (f"ID: {str(order.order_state_key)[:8]}..." if order.order_state_key else "")
            supplier_name = getattr(supplier, "supplier_name", "") or ""
            delivery_date_raw = order_item.delivery_date
            delivery_date = delivery_date_raw.strftime("%Y-%m-%d") if isinstance(delivery_date_raw, datetime) else str(delivery_date_raw or "")
            row_data = [
                order.order_number or "",
                order_date,
                state_display,
                supplier_name,
                getattr(item, "item_name", "") or "",
                getattr(item, "item_article", "") or "",
                getattr(item, "item_code", "") or "",
                order_item.line_number,
                float(order_item.quantity or 0.0),
                float(order_item.received_qty or 0.0),
                float(order_item.remaining_qty or 0.0),
                delivery_date,
                float(order_item.price or 0.0),
                float(order_item.amount or 0.0),
            ]

            fill = zebra_fill if total_rows % 2 == 0 else None
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
            row_num += 1
            total_rows += 1

        row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_num).column_letter
        for row in range(1, min(total_rows + 2, 300)):
            cell_value = ws.cell(row=row, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 55)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    b64 = base64.b64encode(bio.read()).decode("utf-8")
    filename = f"supplier_orders_included_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": b64,
        "filename": filename,
        "total_rows": total_rows,
        "orders_count": len(grouped_rows),
    }
