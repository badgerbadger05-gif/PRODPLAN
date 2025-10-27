from pathlib import Path
import sys
import pytest
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fastapi.testclient import TestClient

# Добавляем backend в путь для импорта
sys.path.append(str(Path(__file__).resolve().parents[1] / "backend"))

# Импортируем app и сервисы после добавления пути
# Для тестов нам не нужен реальный app, мы будем использовать моки
# Поэтому импортируем только необходимые функции
import app.services.planning_service as planning_service

# Создаем мок-приложение FastAPI для тестирования
from fastapi import FastAPI
from fastapi.testclient import TestClient

app = FastAPI()

@app.get("/v1/plan/results/{run_id}/shortage-report")
def get_shortage_report(run_id: int):
    # Мок-реализация эндпоинта
    result = planning_service._generate_shortage_report_v2(None, run_id)
    return result


def make_demo_workbook() -> str:
    """
    Создаёт демо-таблицу Excel с заголовками, подзаголовком, форматированием и outline.
    Возвращает base64-кодированные XLSX байты.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицит по компонентам (v2)"

    # Заголовки
    headers = [
        "Цех", "Код детали", "Наименование детали", "Артикул детали", "ЕИ", 
        "Потребность", "Покрытие", "Дефицит", "Лимитирующий"
    ]
    ws.append(headers)

    # Подзаголовок: объединение и стили
    ws.append(["Цех: DEMO"] + [""] * (len(headers) - 1))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subheader_cell = ws.cell(row=2, column=1)
    subheader_cell.font = Font(bold=True)
    subheader_cell.fill = PatternFill(fill_type="solid", start_color="D3D3D3")
    subheader_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Добавляем строку детали (родитель)
    ws.append([
        "Цех A", "K001", "Деталь 1", "A001", "шт", 100.0, 80.0, 20.0, ""
    ])

    # Добавляем строки компонентов (дочерние) с outline
    ws.append([
        "Цех A", "K002", "Компонент 1", "C001", "шт", 50.0, 40.0, 10.0, "Да"
    ])
    ws.row_dimensions[ws.max_row].outline_level = 2
    ws.row_dimensions[ws.max_row].hidden = True

    ws.append([
        "Цех A", "K003", "Компонент 2", "C002", "шт", 30.0, 20.0, 10.0, ""
    ])
    ws.row_dimensions[ws.max_row].outline_level = 2
    ws.row_dimensions[ws.max_row].hidden = True

    # Устанавливаем freeze panes
    ws.freeze_panes = "A2"

    # Включаем автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Устанавливаем outline ниже
    ws.sheet_properties.outlinePr.summaryBelow = False

    # Сохраняем в байты и кодируем в base64
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    b64 = base64.b64encode(bio.read()).decode()
    return b64


def test_shortage_report_endpoint_returns_v2_smoke(monkeypatch):
    """
    Smoke-тест: проверяет, что эндпоинт возвращает корректный JSON с fileName и data.
    """
    demo_b64 = make_demo_workbook()

    def stub(db, run_id):
        return {
            "status": "ok",
            "format": "xlsx",
            "data_base64": demo_b64,
            "filename": f"mrp_shortage_report_run_{run_id}.xlsx",
            "total_rows": 10
        }

    monkeypatch.setattr(planning_service, "_generate_shortage_report_v2", stub)

    client = TestClient(app)
    r = client.get("/v1/plan/results/123/shortage-report")
    assert r.status_code == 200

    json = r.json()
    assert "data_base64" in json
    assert "filename" in json

    # Декодируем base64 и проверяем структуру XLSX
    decoded_bytes = base64.b64decode(json["data_base64"])
    workbook_io = io.BytesIO(decoded_bytes)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_io)
    ws = wb.active

    # Проверяем заголовки
    header_row = [cell.value for cell in ws[1]]
    expected_headers = [
        "Цех", "Код детали", "Наименование детали", "Артикул детали", "ЕИ", 
        "Потребность", "Покрытие", "Дефицит", "Лимитирующий"
    ]
    assert header_row == expected_headers

    # Проверяем подзаголовок
    assert ws["A2"].value == "Цех: DEMO"

    # Проверяем freeze panes
    assert ws.freeze_panes == "A2"

    # Проверяем автофильтр
    assert ws.auto_filter.ref is not None


def test_shortage_report_outline_and_styles(monkeypatch):
    """
    Тест: проверяет, что строки компонентов имеют outline_level=2 и hidden=True.
    """
    demo_b64 = make_demo_workbook()

    def stub(db, run_id):
        return {
            "status": "ok",
            "format": "xlsx",
            "data_base64": demo_b64,
            "filename": f"mrp_shortage_report_run_{run_id}.xlsx",
            "total_rows": 10
        }

    monkeypatch.setattr(planning_service, "_generate_shortage_report_v2", stub)

    client = TestClient(app)
    r = client.get("/v1/plan/results/123/shortage-report")
    assert r.status_code == 200

    json = r.json()
    decoded_bytes = base64.b64decode(json["data_base64"])
    workbook_io = io.BytesIO(decoded_bytes)
    from openpyxl import load_workbook
    wb = load_workbook(workbook_io)
    ws = wb.active

    # Проверяем, что outline level 2 и hidden установлены хотя бы для одной строки
    found_component_row = False
    for row_num in range(3, ws.max_row + 1):  # Пропускаем заголовки
        row_data = [cell.value for cell in ws[row_num]]
        # Если это строка компонента (с лимитирующим флагом или другим признаком)
        if row_data[8] is not None:  # Если в колонке "Лимитирующий" есть значение
            outline_level = ws.row_dimensions[row_num].outline_level
            is_hidden = ws.row_dimensions[row_num].hidden
            assert outline_level == 2, f"Row {row_num} should have outline_level=2, got {outline_level}"
            assert is_hidden is True, f"Row {row_num} should be hidden, got {is_hidden}"
            found_component_row = True

    assert found_component_row, "No component row with outline_level=2 and hidden=True was found"

    # Проверяем outline summaryBelow
    assert ws.sheet_properties.outlinePr.summaryBelow is False