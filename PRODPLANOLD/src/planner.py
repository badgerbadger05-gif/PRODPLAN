from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Optional, Sequence, Any, cast

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import pandas as pd

from .database import get_connection
from .bom_calculator import get_root_products, explode_bom_for_root


HEADER_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="C8C8C8"),
    right=Side(style="thin", color="C8C8C8"),
    top=Side(style="thin", color="C8C8C8"),
    bottom=Side(style="thin", color="C8C8C8"),
)

COLOR_HEADER = "FFFFD700"   # Золотой (ARGB)
COLOR_WEEKEND = "FFE6E6E6"  # Серый (ARGB)
COLOR_SUBHEADER = "FFF0F0F0"  # Светло‑серый для подзаголовков изделий (ARGB)
COLOR_BG_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
COLOR_BG_WEEKEND = PatternFill("solid", fgColor=COLOR_WEEKEND)
COLOR_BG_SUBHEADER = PatternFill("solid", fgColor=COLOR_SUBHEADER)


def _read_nomen_index_map() -> dict[str, str]:
    """
    Возвращает карту {item_code -> item_article} из локального индекса output/nomenclature_index.json.
    Если файл отсутствует или поврежден — возвращает пустой словарь.
    """
    try:
        p = Path("output/nomenclature_index.json")
        if not p.exists():
            return {}
        import json
        with p.open("r", encoding="utf-8") as f:
            data = json.load(f) or {}
        items = data.get("items") if isinstance(data, dict) else None
        if not isinstance(items, list):
            return {}
        out: dict[str, str] = {}
        for it in items:
            try:
                code = str(it.get("code") or "").strip()
                article = str(it.get("article") or "").strip()
                if code:
                    out[code] = article
            except Exception:
                continue
        return out
    except Exception:
        return {}


def generate_plan_dataframe(
    db_path: Optional[Path] = None,
    horizon_days: int = 30,
    start_date: Optional[dt.date] = None,
) -> pd.DataFrame:
    """
    Построить DataFrame с раскладкой, максимально близкой к листу 'План производства' Excel,
    но без записи/чтения Excel. Источник данных — текущая логика получения корневых изделий.

    Колонки:
      - Номенклатурное наименование изделия
      - Артикул изделия
      - Выполнено (пока пусто)
      - Недовыполнено (пока пусто)
      - План на месяц (сумма по дневным колонкам)
      - Далее горизонт суточных колонок с заголовками дат в формате dd.mm.yy

    Примечание:
      - Суточные значения и, соответственно, 'План на месяц' сейчас инициализируются нулями (0.0),
        чтобы колонка сумм корректно считалась в UI. В Excel-версии здесь были пустые строки и формула.
      - На следующем этапе эти значения будут заполняться расчётом и пользовательскими правками из БД.
    """
    if start_date is None:
        start_date = dt.date.today()

    # 1) Данные корневых изделий из БД
    with get_connection(db_path) as conn:
        roots_df = get_root_products(conn)

    # 2) Заголовки дат
    dates = [(start_date + dt.timedelta(days=i)) for i in range(horizon_days)]
    date_headers = [d.strftime("%d.%m.%y") for d in dates]

    # 3) Формирование строк
    code_to_article = _read_nomen_index_map()
    rows: list[dict[str, Any]] = []
    for _, r in roots_df.iterrows():
        name = str(r.get("item_name", "") or "")
        code = str(r.get("item_code", "") or "")
        article = str(code_to_article.get(code, "") or "")

        # Инициализация дневных значений нулями для наглядного суммирования
        day_vals = [0.0] * horizon_days
        total = float(sum(day_vals))

        record: dict[str, Any] = {
            "Номенклатурное наименование изделия": name,
            "Артикул изделия": article,
            "Код изделия": code,
            "Выполнено": "",
            "Недовыполнено": "",
            "План на месяц": total,
        }
        for i, col in enumerate(date_headers):
            record[col] = day_vals[i]
        rows.append(record)

    columns = (
        ["Номенклатурное наименование изделия", "Артикул изделия", "Код изделия", "Выполнено", "Недовыполнено", "План на месяц"]
        + date_headers
    )
    df = pd.DataFrame(rows, columns=columns)
    return df
def generate_production_plan(
    db_path: Optional[Path] = None,
    output_path: Path | str = Path("output/production_plan.xlsx"),
    horizon_days: int = 30,
    start_date: Optional[dt.date] = None,
    auto_width: bool = True,
) -> Path:
    """
    Сформировать файл production_plan.xlsx (лист 'План производства'):

    Колонки:
    - A: Номенклатурное наименование изделия
    - B: Артикул изделия
    - C: Выполнено
    - D: Недовыполнено
    - E: План на месяц (формула = сумма 30 суточных колонок)
    - F..: 30 колонок с датами планирования (dd.mm.yy)
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if start_date is None:
        start_date = dt.date.today()

    # 1) Данные корневых изделий из БД
    with get_connection(db_path) as conn:
        roots_df = get_root_products(conn)

    # 2) Создать книгу/лист
    wb = Workbook(write_only=False)
    ws: Worksheet = cast(Worksheet, wb.active)
    ws.title = "План производства"

    # 3) Заголовки
    headers = ["Номенклатурное наименование изделия", "Артикул изделия", "Выполнено", "Недовыполнено", "План на месяц"]
    dates = [(start_date + dt.timedelta(days=i)) for i in range(horizon_days)]
    date_headers = [d.strftime("%d.%m.%y") for d in dates]
    full_headers = headers + date_headers

    ws.append(full_headers)

    # Стили заголовка
    for col_idx in range(1, len(full_headers) + 1):
        cell: Cell = cast(Cell, ws.cell(row=1, column=col_idx))
        cell.font = HEADER_FONT
        cell.fill = COLOR_BG_HEADER
        cell.border = THIN_BORDER
        if col_idx in (1, 2):
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = CENTER

    # 4) Данные строк
    # Если корневых изделий нет, всё равно сохранить пустой шаблон
    row_count = 0
    code_to_article = _read_nomen_index_map()
    for _, row in roots_df.iterrows():
        name = str(row.get("item_name", "") or "")
        code = str(row.get("item_code", "") or "")
        article = str(code_to_article.get(code, "") or "")

        # Базовые значения: выполнено/недовыполнено = пусто
        done = ""
        underdone = ""

        # Плейсхолдеры планов (пусто по умолчанию)
        day_values = [""] * horizon_days

        # Сформировать строку без формулы E, формула будет установлена ниже,
        # так как openpyxl не выставляет формулу при append удобно.
        record = [name, article, done, underdone, ""] + day_values
        ws.append(record)
        row_count += 1

        # Установить формулу для "План на месяц" = сумма от F до последней даты
        excel_row = row_count + 1  # т.к. 1 строка - заголовок
        first_day_col = 6  # F
        last_day_col = 5 + horizon_days
        sum_start = f"{get_column_letter(first_day_col)}{excel_row}"
        sum_end = f"{get_column_letter(last_day_col)}{excel_row}"
        cell_e: Cell = cast(Cell, ws.cell(row=excel_row, column=5))
        cell_e.value = f"=SUM({sum_start}:{sum_end})"

        # Стили строк данных: границы и выравнивание для всех ячеек
        for col_idx in range(1, last_day_col + 1):
            c: Cell = cast(Cell, ws.cell(row=excel_row, column=col_idx))
            c.border = THIN_BORDER
            if col_idx in (1, 2):
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = CENTER

    # 5) Выделить выходные дни в календарных колонках (F..)
    for i, d in enumerate(dates, start=0):
        col_idx = 6 + i  # F = 6
        is_weekend = d.weekday() >= 5  # 5=Суббота, 6=Воскресенье
        if is_weekend:
            col_letter = get_column_letter(col_idx)
            # Покрасить весь столбец (строки с 1 по 1+row_count)
            ws_any: Any = cast(Any, ws)
            for r in range(1, row_count + 2):  # включая заголовок
                ws_any[f"{col_letter}{r}"].fill = COLOR_BG_WEEKEND

    # 6) Заморозка панелей: образец не содержит заморозку — не применяем
    # ws.freeze_panes = "F2"

    # 7) Ширины колонок
    if auto_width:
        # Автоматическая ширина по содержимому (ограничение ширины и небольшой отступ)
        apply_auto_width_openpyxl(ws, max_width=80, padding=2)
    else:
        # Фиксированные ширины колонок по образцу
        ws.column_dimensions["A"].width = 68.0
        ws.column_dimensions["B"].width = 20.0
        ws.column_dimensions["C"].width = 14.0
        ws.column_dimensions["D"].width = 18.0
        ws.column_dimensions["E"].width = 18.0
        for i in range(horizon_days):
            ws.column_dimensions[get_column_letter(6 + i)].width = 13.0

    # 8) Добавить листы по этапам (по образцу)
    #    Группировка: один лист на каждый этап, подзаголовки по изделию.
    with get_connection(db_path) as conn2:
        _generate_stage_sheets(conn2, wb, roots_df, start_date, auto_width=auto_width)
        _generate_settings_sheet(conn2, wb, auto_width=auto_width)

    # 9) Сохранить
    wb.save(output_path)
    return output_path


def apply_auto_width_openpyxl(worksheet, max_width: int = 50, padding: int = 5) -> None:
    """
    Автоширина для всех колонок листа на основе максимальной длины значения.
    """
    for col_idx, column in enumerate(worksheet.columns, start=1):
        max_length = 0
        for cell in column:
            val = cell.value
            if val is None:
                continue
            try:
                length = len(str(val))
                if length > max_length:
                    max_length = length
            except Exception:
                pass
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = min(max_length + padding, max_width)


def _generate_stage_sheets(conn, wb: Workbook, roots_df, stocks_date: dt.date, auto_width: bool = True) -> None:
    """
    Сформировать листы по этапам производства.
    Структура листа этапа:
      1: Номенклатурное наименование детали/компонента
      2: Артикул детали
      3: Количество на одно изделие
      4: Остатки на дату загрузки
      5: Минимальная партия заказа
      6: Максимальная партия заказа
      7: Время пополнения (дни)
      8: В производство
    """
    # Получить список этапов (в заданном порядке)
    stages = conn.execute(
        "SELECT stage_id, stage_name FROM production_stages ORDER BY COALESCE(stage_order, stage_id)"
    ).fetchall()

    # Подготовить множество уже занятых имён листов (для уникальности)
    used_titles: set[str] = set(str(t) for t in wb.sheetnames)

    # Считать остатки один раз: только для номенклатуры, присутствующей в БД (берём напрямую из БД)
    rows = conn.execute("SELECT item_code, COALESCE(stock_qty, 0.0) AS qty FROM items").fetchall()
    stock_by_code = {str(r[0]): float(r[1]) for r in rows}

    for stage_id, stage_name in stages:
        # Переименование некоторых листов для совпадения с образцом
        stage_title_overrides: dict[str, str] = {
            "Закупка": "Закупные позиции",
        }
        raw_title = str(stage_name) if stage_name is not None else "Этап"
        title_key = raw_title.strip()
        display_title = stage_title_overrides.get(title_key, title_key)
        sheet_title = _sanitize_sheet_title(display_title, used_titles)
        ws: Worksheet = cast(Worksheet, wb.create_sheet(title=sheet_title))

        # Заголовки с датой в колонке "Остаток на …" в формате dd.mm.YYYY
        date_str = stocks_date.strftime("%d.%m.%Y")
        is_purchase = (display_title == "Закупные позиции" or title_key == "Закупка")
        if is_purchase:
            # Специальная шапка для листа "Закупные позиции" (8 колонок, без "В производство")
            headers = [
                "Номенклатурное наименование",
                "Артикул",
                "Количество на одно изделие",
                f"Остаток на {date_str}",
                "Минимальная партия заказа",
                "Максимальная партия заказа",
                "Заказано",
                "Срок пополнения (дни)",
            ]
        else:
            headers = [
                "Номенклатурное наименование детали/компонента",
                "Артикул детали",
                "Количество на одно изделие",
                f"Остаток на {date_str}",
                "Минимальная партия заказа",
                "Максимальная партия заказа",
                "Время пополнения (дни)",
                "В производство",
            ]
        ws.append(headers)
        # Стили заголовка (левое выравнивание для A,B; центр для остальных)
        for col_idx in range(1, len(headers) + 1):
            cell: Cell = cast(Cell, ws.cell(row=1, column=col_idx))
            cell.font = HEADER_FONT
            cell.fill = COLOR_BG_HEADER
            cell.border = THIN_BORDER
            if col_idx in (1, 2):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = CENTER

        current_row = 1

        # По каждому корневому изделию — подзаголовок + список компонентов этого этапа
        for _, root in roots_df.iterrows():
            root_code = str(root.get("item_code") or "")
            root_name = str(root.get("item_name") or "")

            # Подзаголовок изделия (merge A..H)
            current_row += 1
            end_col = len(headers)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=end_col)
            title_cell: Cell = cast(Cell, ws.cell(row=current_row, column=1))
            title_cell.value = f"{root_name} [{root_code}]"
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="left", vertical="center")
            title_cell.fill = COLOR_BG_SUBHEADER
            title_cell.border = THIN_BORDER
            # Проставить границы в объединённом диапазоне
            for c in range(1, end_col + 1):
                merged_cell: Cell = cast(Cell, ws.cell(row=current_row, column=c))
                merged_cell.border = THIN_BORDER

            # Вычислить потребность на 1 изделие и отфильтровать по этапу
            df = explode_bom_for_root(conn, root_code=root_code, order_qty=1.0, max_depth=15)
            if not df.empty:
                stage_df = df[df["stage_name"] == stage_name].copy()
                stage_df = stage_df.sort_values(["item_code"])

                for _, comp in stage_df.iterrows():
                    current_row += 1
                    name_v = str(comp.get("item_name") or "")
                    code_v = str(comp.get("item_code") or "")
                    qty_v = float(comp.get("required_qty") or 0.0)  # Количество на 1 изделие
                    # Динамическая длина строки под шапку листа
                    row_values = [name_v, code_v, qty_v] + [""] * (len(headers) - 3)
                    ws.append(row_values)
                    # Проставить остаток по артикулу в 4‑й колонке (Остаток на …)
                    stock_val = float(stock_by_code.get(code_v, 0.0)) if isinstance(stock_by_code, dict) else 0.0
                    ws.cell(row=current_row, column=4, value=stock_val)
                    # Стили строки и выравнивание
                    for col_idx in range(1, len(headers) + 1):
                        cell: Cell = cast(Cell, ws.cell(row=current_row, column=col_idx))
                        cell.border = THIN_BORDER
                        if col_idx in (1, 2):
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.alignment = CENTER

        # Фиксированные ширины колонок для листа этапа (по образцу)
        # Базовые значения + точечные переопределения для некоторых листов, чтобы совпасть с образцом
        widths = {
            "A": 79.0,
            "B": 19.0,
            "C": 31.0,
            "D": 26.0,
            "E": 30.0,
            "F": 31.0,
            "G": 27.0,
            "H": 19.0,
        }
        # Для листа "Закупные позиции" убрана последняя колонка
        overrides_map: dict[str, dict[str, float]] = {
            "Покраска": {"A": 81.0},
            "Фрезеровка": {"B": 22.0},
            "Гибка": {"B": 20.0},
            "Сверловка": {"B": 20.0},
            "Зенковка": {"A": 65.0, "B": 21.0},
            "Зачистка": {"A": 72.0},
            "Механическая обработка": {"A": 63.0, "B": 24.0},
            "Опресовка": {"A": 76.0},
            "Оклеивание наклеек": {"A": 77.0},
            # Ширины листа "Закупные позиции" по образцу (без последней колонки)
            "Закупка": {"A": 104.0, "B": 18.0, "G": 13.0, "H": 26.0},
        }

        overrides_map: dict[str, dict[str, float]] = {
            "Покраска": {"A": 81.0},
            "Фрезеровка": {"B": 22.0},
            "Гибка": {"B": 20.0},
            "Сверловка": {"B": 20.0},
            "Зенковка": {"A": 65.0, "B": 21.0},
            "Зачистка": {"A": 72.0},
            "Механическая обработка": {"A": 63.0, "B": 24.0},
            "Опресовка": {"A": 76.0},
            "Оклеивание наклеек": {"A": 77.0},
            # Ширины листа "Закупные позиции" по образцу
            "Закупка": {"A": 104.0, "B": 18.0, "G": 13.0, "H": 26.0, "I": 19.0},
        }
        if auto_width:
            # Автоматическая ширина по содержимому листа этапа
            apply_auto_width_openpyxl(ws, max_width=100, padding=2)
        else:
            key_for_override = (str(stage_name).strip() if stage_name is not None else "")
            if key_for_override in overrides_map:
                for col_letter, w in overrides_map[key_for_override].items():
                    widths[col_letter] = w

            for col_letter, w in widths.items():
                ws.column_dimensions[col_letter].width = w


def _generate_settings_sheet(conn, wb: Workbook, auto_width: bool = True) -> None:
    """
    Лист 'Таблица настроек' для визуального соответствия образцу.
    Колонки:
      - Этап (включая закупные позиции)
      - Сдвиг планирования
      - Диапазон планирования
      - Флаг активности этапа в расчёте
      - Время пополнения (дни)
    Ширины: A:36, B:23, C:26, D:36, E:25
    """
    ws: Worksheet = cast(Worksheet, wb.create_sheet(title="Таблица настроек", index=1))

    # Заголовки
    headers = [
        "Этап (включая закупные позиции)",
        "Сдвиг планирования",
        "Диапазон планирования",
        "Время пополнения (дни)",
        "Флаг активности этапа в расчёте",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell: Cell = cast(Cell, ws.cell(row=1, column=col_idx))
        cell.font = HEADER_FONT
        cell.fill = COLOR_BG_HEADER
        cell.border = THIN_BORDER
        if col_idx in (1, 2):
            cell.alignment = Alignment(horizontal="left", vertical="center")
        else:
            cell.alignment = CENTER

    # Данные: этапы из БД + синтетическая строка "Закупные позиции" (если отсутствует)
    stage_rows = conn.execute(
        "SELECT stage_name FROM production_stages ORDER BY COALESCE(stage_order, stage_id)"
    ).fetchall()
    names: list[str] = [str(row[0]) for row in stage_rows if row and row[0]]

    synthetic_label = "Закупные позиции"
    if synthetic_label not in names:
        try:
            idx = names.index("Закупка") + 1
            names.insert(idx, synthetic_label)
        except ValueError:
            names.append(synthetic_label)

    # Словарь значений по умолчанию для времени пополнения по этапам
    default_lead_times = {
        'Механообработка': 3,
        'Сборка': 2,
        'Закупка': 7,
        'Покраска': 2,
        'Фрезеровка': 3,
        'Гибка': 2,
        'Сверловка': 2,
        'Зенковка': 1,
        'Зачистка': 1,
        'Механическая обработка': 3,
        'Опресовка': 1,
        'Оклеивание наклеек': 1,
        'Закупные позиции': 7
    }

    current_row = 1
    for name in names:
        current_row += 1
        # Получаем значение времени пополнения по умолчанию для этапа
        lead_time = default_lead_times.get(name, 7)
        row_values = [name, 0, 30, "Да", lead_time]
        ws.append(row_values)
        for col_idx in range(1, 6):
            cell: Cell = cast(Cell, ws.cell(row=current_row, column=col_idx))
            cell.border = THIN_BORDER
            if col_idx in (1, 2):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = CENTER

    # Ширины колонок
    if auto_width:
        # Автоматическая ширина по содержимому
        apply_auto_width_openpyxl(ws, max_width=60, padding=2)
    else:
        # Фиксированные ширины колонок по образцу
        ws.column_dimensions["A"].width = 36.0
        ws.column_dimensions["B"].width = 23.0
        ws.column_dimensions["C"].width = 26.0
        ws.column_dimensions["D"].width = 36.0
        ws.column_dimensions["E"].width = 25.0
    
    # Добавить выпадающие списки да/нет в колонке D
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(type="list", formula1='"Да,Нет"', allow_blank=True)
    ws.add_data_validation(dv)
    # Применить ко всем ячейкам в колонке D, кроме заголовка
    for row in range(2, current_row + 1):
        dv.add(ws[f"D{row}"])


def _sanitize_sheet_title(title: str, used: set[str]) -> str:
    """
    Привести заголовок листа к требованиям Excel:
    - Запрет символов: []:*?/\
    - Длина не более 31
    - Уникальность в пределах книги (добавление _1, _2 и т.п.)
    """
    invalid = set('[]:*?/\\')
    sanitized = "".join(("_" if ch in invalid else ch) for ch in title).strip()
    if not sanitized:
        sanitized = "Лист"
    # Ограничить длину
    if len(sanitized) > 31:
        sanitized = sanitized[:31]
    base = sanitized
    i = 1
    while sanitized in used:
        suffix = f"_{i}"
        maxlen = 31 - len(suffix)
        sanitized = (base[:maxlen] if len(base) > maxlen else base) + suffix
        i += 1
    used.add(sanitized)
    return sanitized


if __name__ == "__main__":
    # Генерация полного файла плана (лист плана + листы этапов)
    generate_production_plan()