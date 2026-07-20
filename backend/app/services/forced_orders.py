from __future__ import annotations

from datetime import date as dt_date
from typing import Any, Dict, List, Optional

from sqlalchemy.orm import Session
from sqlalchemy import func

from ..models import (
    ForcedOrderRequest,
    ForcedOrderResult,
    PlanningRun,
    Item,
    Unit,
    DefaultSpecification,
    Specification,
    SpecComponent,
    ProductionResource,
    ResourceProductionKind,
)

from .order_quantity_calculator import OrderQuantityCalculator
from .planning_service import DEFAULT_PLANNING_CONFIG, get_active_planning_config


# NOTE: kept local (not app.utils.dates.to_date): this variant RAISES on bad
# input instead of returning None, does not handle datetime, and parses the
# full string (no [:10] truncation). Different contract; preserved as-is.
def _to_date(val: Any) -> dt_date:
    if isinstance(val, dt_date):
        return val
    if isinstance(val, str):
        return dt_date.fromisoformat(val)
    raise TypeError(f"Cannot convert {type(val)} to date")


def create_forced_order_request(
    db: Session,
    *,
    run_id: Optional[int],
    item_id: int,
    need_date: dt_date,
    requested_qty: float,
    created_by: Optional[str] = None,
    reason: Optional[str] = None,
) -> ForcedOrderRequest:
    rec = ForcedOrderRequest(
        run_id=int(run_id) if run_id is not None else None,
        item_id=int(item_id),
        need_date=_to_date(need_date),
        requested_qty=float(requested_qty or 0.0),
        created_by=created_by,
        reason=reason,
        status="PENDING",
        meta={},
    )
    db.add(rec)
    db.flush()
    return rec


def _build_order_qty_calculator_for_single_item(
    db: Session,
    *,
    snapshot: Dict[str, Any],
    item_id: int,
    requested_qty: float,
) -> OrderQuantityCalculator:
    # Default spec map + spec cache
    default_specs = db.query(DefaultSpecification).all()
    default_spec_map = {int(ds.item_id): int(ds.spec_id) for ds in default_specs}

    all_specs = db.query(Specification).all()
    spec_by_id = {int(s.spec_id): s for s in all_specs}

    def components_loader(spec_id: int) -> List[Any]:
        return db.query(SpecComponent).filter(SpecComponent.spec_id == int(spec_id)).all()

    # Resources
    all_resources = db.query(ProductionResource).all()
    res_by_id = {int(r.resource_id): r for r in all_resources}

    all_res_kinds = db.query(ResourceProductionKind).all()
    production_kinds_by_resource: Dict[int, set] = {}
    for rk in all_res_kinds:
        rid = int(rk.resource_id)
        pk = int(rk.production_kind_id)
        production_kinds_by_resource.setdefault(rid, set()).add(pk)

    # Item cache
    item = db.query(Item).filter(Item.item_id == int(item_id)).first()
    if not item:
        raise RuntimeError(f"Item {item_id} not found")
    item_by_id = {int(item.item_id): item}

    # Units cache
    units_all = db.query(Unit).all()
    units_by_ref = {getattr(u, "unit_ref1c"): u for u in units_all}

    # Stock + WIP
    stock_by_item = {int(item.item_id): float(item.stock_qty or 0.0)}
    wip_by_item: Dict[int, float] = {}
    include_wip = bool((snapshot or {}).get("toggles", {}).get("include_wip", False))
    if include_wip:
        try:
            from ..models import ProductionProduct

            wip_rows = (
                db.query(ProductionProduct.item_id, func.sum(ProductionProduct.quantity))
                .group_by(ProductionProduct.item_id)
                .all()
            )
            wip_by_item = {int(iid): float(qty or 0.0) for iid, qty in wip_rows}
        except Exception:
            wip_by_item = {}

    horizon_days = int((snapshot or {}).get("planning_horizon_days", 90) or 90)
    # Important: for forced orders we treat requested qty as "demand in horizon" so horizon_limit is not 0.
    total_demand_by_item = {int(item_id): float(requested_qty or 0.0)}

    return OrderQuantityCalculator(
        snapshot=snapshot or {},
        default_spec_map=default_spec_map,
        spec_by_id=spec_by_id,
        components_loader=components_loader,
        item_by_id=item_by_id,
        units_by_ref=units_by_ref,
        res_by_id=res_by_id,
        production_kinds_by_resource=production_kinds_by_resource,
        stock_by_item=stock_by_item,
        wip_by_item=wip_by_item,
        horizon_days=horizon_days,
        total_demand_by_item=total_demand_by_item,
    )


def process_forced_order_request(db: Session, request_id: int) -> Dict[str, Any]:
    req: Optional[ForcedOrderRequest] = (
        db.query(ForcedOrderRequest).filter(ForcedOrderRequest.id == int(request_id)).first()
    )
    if not req:
        raise RuntimeError(f"ForcedOrderRequest {request_id} not found")

    # Determine snapshot: prefer linked run snapshot if exists, else active config, else default
    snapshot: Dict[str, Any]
    if req.run_id is not None:
        run = db.query(PlanningRun).filter(PlanningRun.run_id == int(req.run_id)).first()
        snapshot = dict(getattr(run, "config_snapshot", None) or {}) if run else {}
    else:
        try:
            _, cfg = get_active_planning_config(db)
            snapshot = dict(cfg or {})
        except Exception:
            snapshot = dict(DEFAULT_PLANNING_CONFIG)

    requested_qty = float(req.requested_qty or 0.0)
    if requested_qty <= 0:
        raise RuntimeError("requested_qty must be > 0")

    calc = _build_order_qty_calculator_for_single_item(
        db,
        snapshot=snapshot,
        item_id=int(req.item_id),
        requested_qty=requested_qty,
    )

    final_qty_before, normalized_qty, details, warnings = calc.compute(int(req.item_id), requested_qty)

    # For forced orders we do NOT block by component_limit.
    planned_qty = float(normalized_qty or 0.0)
    if planned_qty <= 1e-9:
        # Safety: fall back to requested qty
        planned_qty = float(final_qty_before or requested_qty)

    # Upsert result
    existing: Optional[ForcedOrderResult] = (
        db.query(ForcedOrderResult).filter(ForcedOrderResult.request_id == int(req.id)).first()
    )
    payload_shortage = {
        "details": details,
        "warnings": warnings,
    }

    if existing:
        existing.planned_qty = planned_qty
        existing.normalized_qty = float(details.get("normalized_qty") or normalized_qty or 0.0)
        existing.horizon_limit = float(details.get("horizon_limit") or 0.0)
        existing.component_limit = float(details.get("component_limit") or 0.0)
        existing.shortage = payload_shortage
        res = existing
    else:
        res = ForcedOrderResult(
            request_id=int(req.id),
            planned_qty=planned_qty,
            normalized_qty=float(details.get("normalized_qty") or normalized_qty or 0.0),
            horizon_limit=float(details.get("horizon_limit") or 0.0),
            component_limit=float(details.get("component_limit") or 0.0),
            shortage=payload_shortage,
        )
        db.add(res)

    # Update request status
    req.status = "PROCESSED"
    req.error = None
    req.meta = {
        "planned_qty": planned_qty,
        "component_limit": float(details.get("component_limit") or 0.0),
    }
    db.flush()

    return {
        "request": {
            "id": int(req.id),
            "run_id": int(req.run_id) if req.run_id is not None else None,
            "item_id": int(req.item_id),
            "need_date": req.need_date.isoformat() if req.need_date else None,
            "requested_qty": float(req.requested_qty or 0.0),
            "status": str(req.status),
        },
        "result": {
            "planned_qty": float(planned_qty),
            "normalized_qty": float(details.get("normalized_qty") or normalized_qty or 0.0),
            "horizon_limit": float(details.get("horizon_limit") or 0.0),
            "component_limit": float(details.get("component_limit") or 0.0),
            "warnings": warnings,
        },
    }


def export_forced_order_xlsx(db: Session, request_id: int) -> Dict[str, Any]:
    req: Optional[ForcedOrderRequest] = (
        db.query(ForcedOrderRequest).filter(ForcedOrderRequest.id == int(request_id)).first()
    )
    if not req:
        raise RuntimeError(f"ForcedOrderRequest {request_id} not found")

    item = db.query(Item).filter(Item.item_id == int(req.item_id)).first()
    if not item:
        raise RuntimeError(f"Item {req.item_id} not found")

    res: Optional[ForcedOrderResult] = (
        db.query(ForcedOrderResult).filter(ForcedOrderResult.request_id == int(req.id)).first()
    )

    planned_qty = float(getattr(res, "planned_qty", None) or req.requested_qty or 0.0)
    component_limit = float(getattr(res, "component_limit", None) or 0.0)
    status = "OK"
    if component_limit <= 1e-9:
        status = "IGNORED_COMPONENT_SHORTAGE"

    # Unit display
    unit_display = ""
    try:
        uref = getattr(item, "unit", None)
        if uref:
            u = db.query(Unit).filter(Unit.unit_ref1c == uref).first()
            unit_display = (
                (getattr(u, "short_name", None) or "")
                or (getattr(u, "unit_name", None) or "")
                or (getattr(u, "unit_code", None) or "")
            )
    except Exception:
        unit_display = ""

    headers = [
        "Наименование",
        "Артикул",
        "Количество",
        "Дата потребности",
        "ЕИ",
        "Статус",
        "Лимит по комплектующим",
    ]
    row = [
        getattr(item, "item_name", "") or "",
        getattr(item, "item_article", "") or "",
        planned_qty,
        req.need_date.isoformat() if req.need_date else "",
        unit_display or "",
        status,
        component_limit,
    ]

    import io, base64

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "ForcedOrder"
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    ws.append(row)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    b64 = base64.b64encode(bio.read()).decode("utf-8")
    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": b64,
        "filename": f"forced_order_{request_id}.xlsx",
        "total_rows": 1,
    }


def export_shortage_report_for_run(db: Session, run_id: int) -> Dict[str, Any]:
    run: Optional[PlanningRun] = db.query(PlanningRun).filter(PlanningRun.run_id == int(run_id)).first()
    if not run:
        raise RuntimeError(f"Run {run_id} not found")

    warnings = list(getattr(run, "warnings", None) or [])
    rows = []
    for w in warnings:
        if not isinstance(w, dict):
            continue
        code = str(w.get("code") or "")
        if code not in {"COMPONENT_SHORTAGE_BLOCKED", "COMPONENT_SHORTAGE_PARTIAL"}:
            continue
        rows.append(
            {
                "code": code,
                "item_id": w.get("item_id"),
                "item_code": w.get("item_code"),
                "item_name": w.get("item_name"),
                "item_article": w.get("item_article"),
                "unit": w.get("unit"),
                "need_date": w.get("need_date"),
                "requested_qty": w.get("requested_qty"),
                "planned_qty": w.get("planned_qty"),
                "msg": w.get("msg") or w.get("message") or "",
            }
        )

    headers = [
        "Код",
        "Наименование",
        "Артикул",
        "ЕИ",
        "Дата потребности",
        "Запрошено",
        "Запланировано",
        "Статус",
        "Комментарий",
    ]

    # Resolve "участок" for grouping, similarly to production export: for shortage rows we may not have
    # planned_order stages (blocked cases). We group by the first mapped resource for item's production kind.
    default_specs = db.query(DefaultSpecification).all()
    default_spec_map = {int(ds.item_id): int(ds.spec_id) for ds in default_specs}
    all_specs = db.query(Specification).all()
    spec_by_id = {int(s.spec_id): s for s in all_specs}

    all_resources = db.query(ProductionResource).all()
    res_by_id = {int(r.resource_id): r for r in all_resources}
    all_res_kinds = db.query(ResourceProductionKind).all()
    kind_to_resources: Dict[int, List[int]] = {}
    for rk in all_res_kinds:
        kind_to_resources.setdefault(int(rk.production_kind_id), []).append(int(rk.resource_id))
    for kind_id in list(kind_to_resources.keys()):
        kind_to_resources[kind_id] = sorted(set(kind_to_resources[kind_id]))

    def resolve_area_name(item_id_val: Any) -> str:
        try:
            iid = int(item_id_val)
        except Exception:
            return "Без участка"
        spec_id = default_spec_map.get(iid)
        if not spec_id:
            return "Без участка"
        spec = spec_by_id.get(int(spec_id))
        kind_id = getattr(spec, "production_kind_id", None) if spec is not None else None
        if kind_id is None:
            return "Без участка"
        candidates = kind_to_resources.get(int(kind_id)) or []
        if not candidates:
            return "Без участка"
        res = res_by_id.get(int(candidates[0]))
        name = str(getattr(res, "resource_name", None) or "").strip()
        return name or "Без участка"

    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        area_name = resolve_area_name(r.get("item_id"))
        r["area_name"] = area_name
        grouped.setdefault(area_name, []).append(r)

    import io, base64

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицит"

    # Auto width tracking
    max_widths = {i: len(str(h)) for i, h in enumerate(headers, start=1)}

    def update_widths(values: list):
        for idx, val in enumerate(values, start=1):
            text = "" if val is None else str(val)
            length = max((len(line) for line in text.splitlines()), default=0)
            if length > max_widths.get(idx, 0):
                max_widths[idx] = length

    def style_header(row_idx: int):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)

    def append_group_title(title: str):
        ws.append([title])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF4F81BD")
        cell.alignment = Alignment(horizontal="left")
        update_widths([title])

    # Deterministic order
    for area_name in sorted(grouped.keys()):
        append_group_title(f"Участок: {area_name}")
        ws.append(headers)
        hdr_row = ws.max_row
        style_header(hdr_row)
        update_widths(headers)
        for r in grouped.get(area_name, []):
            status = "Блок" if r["code"] == "COMPONENT_SHORTAGE_BLOCKED" else "Частично"
            row_vals = [
                r.get("item_code") or "",
                r.get("item_name") or "",
                r.get("item_article") or "",
                r.get("unit") or "",
                r.get("need_date") or "",
                float(r.get("requested_qty") or 0.0),
                float(r.get("planned_qty") or 0.0),
                status,
                r.get("msg") or "",
            ]
            ws.append(row_vals)
            update_widths(row_vals)
        ws.append([])

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        width = max_widths.get(col_idx, 10)
        adjusted = min(max(width * 1.2 + 2, 12), 80)
        ws.column_dimensions[letter].width = adjusted

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    b64 = base64.b64encode(bio.read()).decode("utf-8")
    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": b64,
        "filename": f"mrp_shortage_report_run_{run_id}.xlsx",
        "total_rows": len(rows),
    }

