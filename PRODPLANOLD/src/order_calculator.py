"""
Модуль для расчета потребностей и формирования заказов.

Этот модуль предоставляет функции для:
- Расчета потребностей в компонентах на основе плана производства
- Анализа остатков и определения дефицитов
- Формирования заказов с цветовой индикацией статусов
- Экспорта заказов в Excel файлы
"""

from __future__ import annotations

import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd

from .database import get_connection


def calculate_component_needs(
    db_path: Optional[Path] = None,
    production_plan: Optional[Dict[str, int]] = None,
) -> pd.DataFrame:
    """
    Рассчитывает потребности в компонентах на основе плана производства.
    
    Args:
        db_path: Путь к базе данных SQLite
        production_plan: Словарь {код_изделия: количество} для расчета потребностей
        
    Returns:
        DataFrame с компонентами и их потребностями
    """
    if production_plan is None:
        production_plan = {}
    
    with get_connection(db_path) as conn:
        # Создаем временную таблицу заказов
        orders_data = list(production_plan.items())
        orders_df = pd.DataFrame(orders_data, columns=['item_code', 'qty'])
        orders_df.to_sql('temp_orders', conn, if_exists='replace', index=False)
        
        # Рекурсивный запрос для расчета потребностей
        query = """
        WITH RECURSIVE bom_explosion AS (
            -- Начальные изделия из заказов
            SELECT
                i.item_id,
                i.item_code,
                i.item_name,
                ps.stage_name,
                o.qty as total_qty,
                0 as level
            FROM temp_orders o
            JOIN items i ON i.item_code = o.item_code
            LEFT JOIN production_stages ps ON ps.stage_id = i.stage_id
            
            UNION ALL
            
            -- Рекурсивное развертывание компонентов
            SELECT
                child.item_id,
                child.item_code,
                child.item_name,
                ps.stage_name,
                parent.total_qty * b.quantity as total_qty,
                parent.level + 1
            FROM bom_explosion parent
            JOIN bom b ON b.parent_item_id = parent.item_id
            JOIN items child ON child.item_id = b.child_item_id
            LEFT JOIN production_stages ps ON ps.stage_id = child.stage_id
            WHERE parent.level < 15  -- защита от глубокой рекурсии
        )
        SELECT
            item_code,
            item_name,
            COALESCE(stage_name, 'Закупка') as stage_name,
            SUM(total_qty) as total_quantity_needed
        FROM bom_explosion
        WHERE level > 0  -- исключить корневые изделия
        GROUP BY item_code, item_name, stage_name
        ORDER BY stage_name, item_code;
        """
        
        result = pd.read_sql_query(query, conn)
        
        # Удаляем временную таблицу
        conn.execute("DROP TABLE IF EXISTS temp_orders")
        
        return result


def get_current_stock(db_path: Optional[Path] = None) -> Dict[str, float]:
    """
    Получает текущие остатки из базы данных.
    
    Args:
        db_path: Путь к базе данных SQLite
        
    Returns:
        Словарь {код_товара: остаток}
    """
    with get_connection(db_path) as conn:
        cursor = conn.execute("SELECT item_code, COALESCE(stock_qty, 0.0) FROM items")
        rows = cursor.fetchall()
        
        return {str(row[0]): float(row[1]) for row in rows}


def determine_order_statuses(
    component_needs: pd.DataFrame,
    stock_data: Dict[str, float],
    lead_times: Optional[Dict[str, int]] = None,
    db_path: Optional[Path] = None,
) -> pd.DataFrame:
    """
    Определяет статусы заказов на основе потребностей и остатков.
    
    Args:
        component_needs: DataFrame с потребностями в компонентах
        stock_data: Словарь текущих остатков {код_товара: остаток}
        lead_times: Словарь времен пополнения по этапам
        db_path: Путь к базе данных SQLite
        
    Returns:
        DataFrame с заказами и их статусами
    """
    if lead_times is None:
        lead_times = {
            'Механообработка': 3,
            'Сборка': 2,
            'Закупка': 7,
            'Покраска': 2,
            'Фрезеровка': 3,
            'Гибка': 2,
            'Сверловка': 2,
            'Зенковка': 1,
            'Зачистка': 1,
            'Механическая обработка': 3,
            'Опресовка': 1,
            'Оклеивание наклеек': 1
        }
    
    results = []
    
    for _, component in component_needs.iterrows():
        item_code = str(component['item_code'])
        item_name = str(component['item_name'])
        stage_name = str(component['stage_name'])
        required_qty = float(component['total_quantity_needed'])
        current_stock = stock_data.get(item_code, 0.0)
        
        shortage = max(0.0, required_qty - current_stock)
        
        if shortage <= 0:
            continue  # Дефицита нет
            
        # Получаем время пополнения
        lead_time = lead_times.get(stage_name, 7)
        
        # Инициализируем даты по умолчанию
        required_date = datetime.now().date() + timedelta(days=7)  # Упрощение
        order_date = required_date - timedelta(days=lead_time)
        
        # Для производственных этапов проверяем доступность комплектующих
        if stage_name != 'Закупка':
            # Здесь можно добавить более сложную логику проверки комплектующих
            # Пока используем упрощенную модель
            subcomponents_available = True
            
            if not subcomponents_available:
                status = 'RED'  # Невозможен из-за дефицита комплектующих
            else:
                # Проверяем сроки
                if order_date <= datetime.now().date():
                    status = 'BLUE'  # Просроченный
                else:
                    status = 'GREEN'  # Плановый
        else:
            # Для закупок - только проверка сроков
            if order_date <= datetime.now().date():
                status = 'BLUE'
            else:
                status = 'GREEN'
        
        results.append({
            'item_code': item_code,
            'item_name': item_name,
            'stage_name': stage_name,
            'required_qty': required_qty,
            'current_stock': current_stock,
            'shortage': shortage,
            'lead_time': lead_time,
            'status': status,
            'order_date': order_date.isoformat() if isinstance(order_date, (datetime, timedelta)) else str(order_date),
            'required_date': required_date.isoformat() if isinstance(required_date, (datetime, timedelta)) else str(required_date)
        })
    
    return pd.DataFrame(results)


def calculate_orders(
    db_path: Optional[Path] = None,
    production_plan: Optional[Dict[str, int]] = None,
    lead_times: Optional[Dict[str, int]] = None,
) -> pd.DataFrame:
    """
    Основная функция расчета заказов.
    
    Args:
        db_path: Путь к базе данных SQLite
        production_plan: Словарь {код_изделия: количество} для расчета потребностей
        lead_times: Словарь времен пополнения по этапам
        
    Returns:
        DataFrame с заказами и их статусами
    """
    # 1. Рассчитываем потребности в компонентах
    component_needs = calculate_component_needs(db_path, production_plan)
    
    # 2. Получаем текущие остатки
    stock_data = get_current_stock(db_path)
    
    # 3. Определяем статусы заказов
    orders = determine_order_statuses(component_needs, stock_data, lead_times, db_path)
    
    return orders


def export_orders_to_excel(
    orders: pd.DataFrame,
    output_dir: Path | str = Path("output"),
    db_path: Optional[Path] = None,
) -> Tuple[Path, Path]:
    """
    Экспортирует заказы в Excel файлы по категориям статусов.
    
    Args:
        orders: DataFrame с заказами
        output_dir: Директория для выходных файлов
        db_path: Путь к базе данных SQLite
        
    Returns:
        Кортеж с путями к созданным файлам (production_orders, purchase_orders)
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Проверяем, что DataFrame не пустой и содержит нужные колонки
    if orders.empty or 'status' not in orders.columns:
        # Если DataFrame пустой или нет колонки 'status', создаем пустые DataFrames
        red_orders = pd.DataFrame()
        blue_orders = pd.DataFrame()
        green_orders = pd.DataFrame()
        purchase_orders = pd.DataFrame()
        other_orders = pd.DataFrame()
    else:
        # Разделяем заказы на закупочные и остальные
        if 'stage_name' in orders.columns:
            purchase_orders = orders[orders['stage_name'] == 'Закупка']
            other_orders = orders[orders['stage_name'] != 'Закупка']
        else:
            purchase_orders = pd.DataFrame()
            other_orders = orders.copy()
        
        # Разделяем остальные заказы по категориям статусов
        red_orders = other_orders[other_orders['status'] == 'RED']
        blue_orders = other_orders[other_orders['status'] == 'BLUE']
        green_orders = other_orders[other_orders['status'] == 'GREEN']
    
    # Создаем файлы заказов на производство и закупку
    production_orders_path = output_dir / "orders_production.xlsx"
    purchase_orders_path = output_dir / "orders_purchase.xlsx"
    
    # Определяем стили
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill("solid", fgColor="FFD700")  # Золотой цвет
    border = Border(
        left=Side(style="thin", color="C8C8C8"),
        right=Side(style="thin", color="C8C8C8"),
        top=Side(style="thin", color="C8C8C8"),
        bottom=Side(style="thin", color="C8C8C8"),
    )
    
    # Словарь для перевода названий колонок на русский язык
    column_names_russian = {
        'item_code': 'Артикул',
        'item_name': 'Наименование',
        'stage_name': 'Этап',
        'required_qty': 'Требуется',
        'current_stock': 'Остаток',
        'shortage': 'Дефицит',
        'lead_time': 'Срок поставки (дни)',
        'status': 'Статус',
        'order_date': 'Дата заказа',
        'required_date': 'Дата потребности'
    }
    
    # Экспортируем заказы на производство
    with pd.ExcelWriter(production_orders_path, engine='openpyxl') as writer:
        # Создаем листы для каждого статуса
        status_groups = [
            ('RED статус', red_orders),
            ('BLUE статус', blue_orders),
            ('GREEN статус', green_orders)
        ]
        
        sheets_created = 0
        
        for sheet_name, df in status_groups:
            if not df.empty:
                # Переименовываем колонки на русский язык
                df_russian = df.rename(columns=column_names_russian)
                
                # Меняем местами колонки "Артикул" и "Наименование"
                cols = list(df_russian.columns)
                if 'Артикул' in cols and 'Наименование' in cols:
                    # Найдем индексы колонок
                    code_idx = cols.index('Артикул')
                    name_idx = cols.index('Наименование')
                    # Поменяем местами
                    cols[code_idx], cols[name_idx] = cols[name_idx], cols[code_idx]
                    df_russian = df_russian[cols]
                
                # Группируем по этапам
                if 'Этап' in df_russian.columns:
                    grouped = df_russian.groupby('Этап')
                    
                    # Создаем новый лист
                    workbook = writer.book
                    worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # Добавляем заголовок с датой формирования
                    current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
                    worksheet.merge_cells('A1:J1')
                    worksheet['A1'] = f"Заказы. Дата формирования: {current_date}"
                    worksheet['A1'].font = Font(bold=True, size=14)
                    worksheet['A1'].alignment = Alignment(horizontal="center")
                    worksheet['A1'].fill = header_fill
                    
                    row_num = 3
                    
                    # Для каждого этапа создаем блок
                    for stage, group in grouped:
                        # Добавляем подзаголовок с названием этапа
                        worksheet.merge_cells(f'A{row_num}:J{row_num}')
                        worksheet[f'A{row_num}'] = f"Этап: {stage}"
                        worksheet[f'A{row_num}'].font = Font(bold=True, size=12)
                        worksheet[f'A{row_num}'].alignment = Alignment(horizontal="center")
                        worksheet[f'A{row_num}'].fill = PatternFill("solid", fgColor="E6E6E6")  # Серый цвет
                        
                        row_num += 1
                        
                        # Добавляем заголовки таблицы
                        headers = list(group.columns)
                        # Убираем колонку "Этап" из заголовков
                        if 'Этап' in headers:
                            headers.remove('Этап')
                        for col_idx, header in enumerate(headers, 1):
                            cell = worksheet.cell(row=row_num, column=col_idx, value=header)
                            cell.font = header_font
                            cell.alignment = center_alignment
                            cell.fill = header_fill
                            cell.border = border
                        
                        row_num += 1
                        
                        # Добавляем данные
                        for _, row in group.iterrows():
                            # Убираем значение колонки "Этап" из данных
                            row_values = list(row)
                            if 'Этап' in list(group.columns):
                                stage_idx = list(group.columns).index('Этап')
                                row_values.pop(stage_idx)
                            for col_idx, value in enumerate(row_values, 1):
                                cell = worksheet.cell(row=row_num, column=col_idx, value=value)
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                                cell.border = border
                            row_num += 1
                        
                        # Добавляем пустую строку между этапами
                        row_num += 1
                        
                    sheets_created += 1
                else:
                    # Если нет колонки "Этап", просто экспортируем как есть
                    df_russian.to_excel(writer, sheet_name=sheet_name, index=False)
                    sheets_created += 1
                    
        # Если ни один лист не был создан, создаем пустой лист, чтобы избежать ошибки
        if sheets_created == 0:
            workbook = writer.book
            worksheet = workbook.create_sheet(title='Пусто')
            worksheet['A1'] = 'Нет данных для экспорта'
            worksheet['A1'].font = Font(bold=True, size=12)
    
    # Экспортируем заказы на закупку
    with pd.ExcelWriter(purchase_orders_path, engine='openpyxl') as writer:
        if not purchase_orders.empty:
            # Переименовываем колонки на русский язык
            purchase_orders_russian = purchase_orders.rename(columns=column_names_russian)
            
            # Меняем местами колонки "Артикул" и "Наименование"
            cols = list(purchase_orders_russian.columns)
            if 'Артикул' in cols and 'Наименование' in cols:
                # Найдем индексы колонок
                code_idx = cols.index('Артикул')
                name_idx = cols.index('Наименование')
                # Поменяем местами
                cols[code_idx], cols[name_idx] = cols[name_idx], cols[code_idx]
                purchase_orders_russian = purchase_orders_russian[cols]
            
            # Группируем по этапам (для закупок это будет один этап "Закупка")
            if 'Этап' in purchase_orders_russian.columns:
                grouped = purchase_orders_russian.groupby('Этап')
                
                # Создаем новый лист
                workbook = writer.book
                worksheet = workbook.create_sheet(title='Закупки')
                
                # Добавляем заголовок с датой формирования
                current_date = datetime.now().strftime("%d.%m.%Y %H:%M")
                worksheet.merge_cells('A1:J1')
                worksheet['A1'] = f"Заказы на закупку. Дата формирования: {current_date}"
                worksheet['A1'].font = Font(bold=True, size=14)
                worksheet['A1'].alignment = Alignment(horizontal="center")
                worksheet['A1'].fill = header_fill
                
                row_num = 3
                
                # Для каждого этапа создаем блок
                for stage, group in grouped:
                    # Добавляем подзаголовок с названием этапа
                    worksheet.merge_cells(f'A{row_num}:J{row_num}')
                    worksheet[f'A{row_num}'] = f"Этап: {stage}"
                    worksheet[f'A{row_num}'].font = Font(bold=True, size=12)
                    worksheet[f'A{row_num}'].alignment = Alignment(horizontal="center")
                    worksheet[f'A{row_num}'].fill = PatternFill("solid", fgColor="E6E6E6")  # Серый цвет
                    
                    row_num += 1
                    
                    # Добавляем заголовки таблицы
                    headers = list(group.columns)
                    # Убираем колонку "Этап" из заголовков
                    if 'Этап' in headers:
                        headers.remove('Этап')
                    for col_idx, header in enumerate(headers, 1):
                        cell = worksheet.cell(row=row_num, column=col_idx, value=header)
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.fill = header_fill
                        cell.border = border
                    
                    row_num += 1
                    
                    # Добавляем данные
                    for _, row in group.iterrows():
                        # Убираем значение колонки "Этап" из данных
                        row_values = list(row)
                        if 'Этап' in list(group.columns):
                            stage_idx = list(group.columns).index('Этап')
                            row_values.pop(stage_idx)
                        for col_idx, value in enumerate(row_values, 1):
                            cell = worksheet.cell(row=row_num, column=col_idx, value=value)
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                            cell.border = border
                        row_num += 1
                    
                    # Добавляем пустую строку между этапами
                    row_num += 1
            else:
                # Если нет колонки "Этап", просто экспортируем как есть
                purchase_orders_russian.to_excel(writer, sheet_name='Закупки', index=False)
        else:
            # Если заказов на закупку нет, создаем пустой лист
            workbook = writer.book
            worksheet = workbook.create_sheet(title='Пусто')
            worksheet['A1'] = 'Нет данных для экспорта'
            worksheet['A1'].font = Font(bold=True, size=12)
    
    # Применяем автоформатирование
    def apply_auto_formatting(file_path: Path):
        """Применяет автоформатирование к Excel файлу"""
        workbook = openpyxl.load_workbook(file_path)
        
        for worksheet in workbook.worksheets:
            # Автоширина колонок
            for column in worksheet.columns:
                max_length = 0
                column_idx = column[0].column if column[0].column is not None else 1
                column_letter = get_column_letter(column_idx)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Применяем стили к заголовкам
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    if cell.value:
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.fill = header_fill
                        cell.border = border
        
        workbook.save(file_path)
    
    # Применяем автоформатирование к обоим файлам
    if production_orders_path.exists():
        apply_auto_formatting(production_orders_path)
    
    if purchase_orders_path.exists():
        apply_auto_formatting(purchase_orders_path)
    
    return production_orders_path, purchase_orders_path


def load_production_plan_from_excel(
    excel_path: Path | str = Path("output/production_plan.xlsx"),
    db_path: Optional[Path] = None,
) -> Dict[str, int]:
    """
    Загружает план производства из Excel файла и применяет настройки.

    Args:
        excel_path: Путь к Excel файлу с планом производства
        db_path: Путь к базе данных SQLite (для проверки соответствия названий изделий)
        
    Returns:
        Словарь {код_изделия: количество} для расчета потребностей
    """
    import pandas as pd
    from datetime import datetime, timedelta
    
    # Преобразуем путь в Path
    excel_path = Path(excel_path)
    
    # Читаем первый лист (план производства)
    df_main = pd.read_excel(excel_path, sheet_name=0, engine='openpyxl')
    
    # Читаем лист "Таблица настроек"
    df_settings = pd.read_excel(excel_path, sheet_name='Таблица настроек', engine='openpyxl')
    
    # Создаем словарь настроек: {этап: (сдвиг, диапазон, активен, время_пополнения)}
    settings_dict = {}
    for _, row in df_settings.iterrows():
        # Используем индексы вместо имен колонок из-за возможных проблем с кодировкой
        stage_name = str(row.iloc[0]).strip()  # 'Этап (номенклатура этапа)'
        shift = int(row.iloc[1]) if not pd.isna(row.iloc[1]) else 0  # 'Сдвиг планирования'
        range_val = int(row.iloc[2]) if not pd.isna(row.iloc[2]) else 30  # 'Диапазон планирования'
        lead_time = int(row.iloc[3]) if not pd.isna(row.iloc[3]) else 7  # 'Время пополнения (дни)'
        is_active = str(row.iloc[4]).strip().lower() == 'да'  # 'Активен в расчете'
        settings_dict[stage_name] = (shift, range_val, is_active, lead_time)
    
    # Получаем сегодняшнюю дату
    today = datetime.now().date()
    
    production_plan = {}
    
    # Если путь к БД не указан, используем путь по умолчанию
    if db_path is None:
        db_path = Path("data/specifications.db")
    
    # Загружаем сопоставление названий и кодов из БД
    item_name_to_code = {}
    with get_connection(db_path) as conn:
        cursor = conn.execute("SELECT item_name, item_code FROM items")
        rows = cursor.fetchall()
        item_name_to_code = {str(row[0]).strip(): str(row[1]).strip() for row in rows}
    
    # Обрабатываем каждую строку первого листа
    for _, row in df_main.iterrows():
        item_name = str(row.iloc[0])  # Название изделия в первой колонке
        # Пропускаем пустые строки
        if not item_name or item_name.strip() == '':
            continue
            
        # Проверяем, есть ли настройки для этого изделия (по названию этапа)
        # Предполагаем, что название изделия совпадает с названием этапа в настройках
        shift, range_val, is_active, lead_time_default = settings_dict.get(item_name, (0, 30, True, 7))
        
        # Если этап не активен, пропускаем его
        if not is_active:
            continue
            
        # Определяем начальную дату с учетом сдвига
        start_date = today + timedelta(days=shift)
        
        # Собираем колонки с датами (все, кроме первых нескольких)
        # Предполагаем, что первая колонка - название, остальные - даты
        date_columns = [col for col in df_main.columns[1:] if isinstance(col, str) and col.count('.') == 2]
        
        # Сортируем даты
        try:
            date_columns_sorted = sorted(date_columns, key=lambda x: datetime.strptime(x, '%d.%m.%y'))
        except ValueError:
            # Если формат даты другой, пробуем другой формат
            try:
                date_columns_sorted = sorted(date_columns, key=lambda x: datetime.strptime(x, '%d.%m.%Y'))
            except ValueError:
                # Если не можем распарсить, используем как есть
                date_columns_sorted = date_columns
        
        # Определяем индексы колонок для диапазона
        # Найдем индекс начальной даты
        start_col_idx = None
        for i, col in enumerate(date_columns_sorted):
            try:
                col_date = datetime.strptime(col, '%d.%m.%y').date()
            except ValueError:
                try:
                    col_date = datetime.strptime(col, '%d.%m.%Y').date()
                except ValueError:
                    continue
                    
            if col_date >= start_date:
                start_col_idx = i
                break
                
        if start_col_idx is None:
            # Если не нашли подходящую дату, начинаем с начала
            start_col_idx = 0
            
        # Определяем конечную дату
        end_col_idx = min(start_col_idx + range_val, len(date_columns_sorted))
        
        # Суммируем значения в заданном диапазоне
        total_qty = 0.0
        for i in range(start_col_idx, end_col_idx):
            col_name = date_columns_sorted[i]
            value = row[col_name]
            if not pd.isna(value) and isinstance(value, (int, float)):
                total_qty += value
                
        # Округляем до целого
        total_qty = int(round(total_qty))
        
        # Добавляем в план, если количество больше 0
        if total_qty > 0:
            # Пытаемся получить код изделия из БД
            item_code = item_name_to_code.get(item_name.strip())
            if item_code:
                production_plan[item_code] = total_qty
            else:
                # Если код не найден, используем название (для обратной совместимости или отладки)
                production_plan[item_name] = total_qty
            
    return production_plan


def load_lead_times_from_excel(
    excel_path: Path | str = Path("output/production_plan.xlsx"),
) -> Dict[str, int]:
    """
    Загружает настройки времени пополнения из листа "Таблица настроек" Excel файла.

    Args:
        excel_path: Путь к Excel файлу с планом производства
        
    Returns:
        Словарь {этап: время_пополнения} для расчета заказов
    """
    import pandas as pd
    
    # Преобразуем путь в Path
    excel_path = Path(excel_path)
    
    # Читаем лист "Таблица настроек"
    df_settings = pd.read_excel(excel_path, sheet_name='Таблица настроек', engine='openpyxl')
    
    # Создаем словарь времен пополнения: {этап: время_пополнения}
    lead_times_dict = {}
    for _, row in df_settings.iterrows():
        # Используем индексы вместо имен колонок из-за возможных проблем с кодировкой
        stage_name = str(row.iloc[0]).strip()  # 'Этап (номенклатура этапа)'
        lead_time = int(row.iloc[3]) if not pd.isna(row.iloc[3]) else 7  # 'Время пополнения (дни)'
        lead_times_dict[stage_name] = lead_time
    
    return lead_times_dict


# Функции для командной строки
def cmd_calculate_orders(args: argparse.Namespace) -> None:
    """
    Команда CLI для расчета заказов.
    """
    db_path = Path(args.db) if args.db else None
    output_dir = Path(args.output) if args.output else Path("output")
    
    # Загружаем план производства из Excel файла с учетом настроек
    production_plan = load_production_plan_from_excel(
        excel_path=Path("output/production_plan.xlsx"),
        db_path=db_path
    )
    
    # Загружаем настройки времени пополнения из Excel файла
    lead_times = load_lead_times_from_excel(
        excel_path=Path("output/production_plan.xlsx")
    )
    
    # Рассчитываем заказы
    orders = calculate_orders(db_path=db_path, production_plan=production_plan, lead_times=lead_times)
    
    # Экспортируем в Excel
    production_file, purchase_file = export_orders_to_excel(orders, output_dir, db_path)
    
    print(f"Заказы на производство сохранены в: {production_file}")
    print(f"Заказы на закупку сохранены в: {purchase_file}")
    print(f"Всего заказов: {len(orders)}")


if __name__ == "__main__":
    import argparse
    
    # Создаем парсер аргументов командной строки для тестирования
    parser = argparse.ArgumentParser(description="Расчет заказов")
    parser.add_argument("--db", type=str, help="Путь к базе данных SQLite")
    parser.add_argument("--output", type=str, default="output", help="Директория для выходных файлов")
    
    args = parser.parse_args()
    cmd_calculate_orders(args)