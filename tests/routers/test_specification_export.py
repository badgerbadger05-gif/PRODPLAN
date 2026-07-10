from __future__ import annotations

import base64
import io

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.models import DefaultSpecification, Item, ProductionStage, SpecComponent, Specification, Unit
from app.routers.specification import router

API = "/api/v1/specification"


@pytest.fixture()
def db_session():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    session = sessionmaker(autocommit=False, autoflush=False, bind=engine)()
    try:
        yield session
    finally:
        session.close()
        Base.metadata.drop_all(bind=engine)


@pytest.fixture()
def client(db_session):
    app = FastAPI()
    app.include_router(router, prefix="/api")

    def override_get_db():
        yield db_session

    app.dependency_overrides[get_db] = override_get_db
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.clear()


def _decode_workbook(data_base64: str):
    return load_workbook(io.BytesIO(base64.b64decode(data_base64)))


def test_specification_export_returns_xlsx_with_tree_and_flat_sheets(client, db_session):
    unit = Unit(unit_ref1c="u-spec-xlsx", unit_name="шт", short_name="шт", precision=0)
    stage = ProductionStage(stage_name="Сборка", stage_order=1, stage_ref1c="stage-spec-xlsx")
    root = Item(
        item_code="ROOT-XLSX",
        item_name="Изделие export",
        item_article="ROOT-ART",
        replenishment_method="Производство",
        unit=unit.unit_ref1c,
    )
    component = Item(
        item_code="COMP-XLSX",
        item_name="Компонент export",
        item_article="COMP-ART",
        replenishment_method="Закупка",
        unit=unit.unit_ref1c,
    )
    db_session.add_all([unit, stage, root, component])
    db_session.flush()
    spec = Specification(spec_code="ROOT-XLSX", spec_name="Спецификация export", spec_ref1c="spec-xlsx")
    db_session.add(spec)
    db_session.flush()
    db_session.add(DefaultSpecification(item_id=root.item_id, spec_id=spec.spec_id))
    db_session.add(SpecComponent(spec_id=spec.spec_id, item_id=component.item_id, quantity=2, stage_id=stage.stage_id))
    db_session.commit()

    response = client.get(f"{API}/export", params={"item_id": root.item_id, "root_qty": 3})

    assert response.status_code == 200
    payload = response.json()
    assert payload["format"] == "xlsx"
    assert payload["filename"] == "specification_ROOT-ART.xlsx"
    workbook = _decode_workbook(payload["data_base64"])
    assert workbook.sheetnames == ["Дерево", "Плоская", "Где используется", "Качество"]

    tree = workbook["Дерево"]
    assert tree["A2"].value == "Уровень"
    assert tree["C4"].value == "Компонент export"
    assert tree["F4"].value == "Закупка"
    assert tree["I4"].value == 6

    flat = workbook["Плоская"]
    assert flat["A3"].value == "Компонент export"
    assert flat["D3"].value == "Закупка"
    assert flat["E3"].value == 6


def test_specification_export_respects_replenishment_method_filter(client, db_session):
    unit = Unit(unit_ref1c="u-spec-filter", unit_name="шт", short_name="шт", precision=0)
    root = Item(item_code="ROOT-FILTER", item_name="Root filter", replenishment_method="Производство", unit=unit.unit_ref1c)
    buy = Item(item_code="BUY-FILTER", item_name="Buy filter", replenishment_method="Закупка", unit=unit.unit_ref1c)
    make = Item(item_code="MAKE-FILTER", item_name="Make filter", replenishment_method="Производство", unit=unit.unit_ref1c)
    db_session.add_all([unit, root, buy, make])
    db_session.flush()
    spec = Specification(spec_code="ROOT-FILTER", spec_name="Spec filter", spec_ref1c="spec-filter")
    db_session.add(spec)
    db_session.flush()
    db_session.add(DefaultSpecification(item_id=root.item_id, spec_id=spec.spec_id))
    db_session.add_all([
        SpecComponent(spec_id=spec.spec_id, item_id=buy.item_id, quantity=1),
        SpecComponent(spec_id=spec.spec_id, item_id=make.item_id, quantity=1),
    ])
    db_session.commit()

    response = client.get(
        f"{API}/export",
        params={"item_id": root.item_id, "replenishment_method": "Закупка"},
    )

    assert response.status_code == 200
    workbook = _decode_workbook(response.json()["data_base64"])
    tree_names = [workbook["Дерево"][f"C{row}"].value for row in range(3, workbook["Дерево"].max_row + 1)]
    flat_names = [workbook["Плоская"][f"A{row}"].value for row in range(3, workbook["Плоская"].max_row + 1)]
    assert "Buy filter" in tree_names
    assert "Make filter" not in tree_names
    assert flat_names == ["Buy filter"]
