"""
Экспорт заказов на производство в Excel (XLSX).
Данные берутся из БД (production_orders + production_products + items).
"""

from __future__ import annotations

import io
import base64
from datetime import datetime
from typing import Dict, Any, List, Optional

from sqlalchemy.orm import Session, joinedload

from ..models import ProductionOrder, ProductionProduct, Item


def _norm_guid(val) -> str:
    """Нормализация GUID для сравнения (lowercase, без фигурных скобок и обёрток)."""
    s = str(val or "").strip().lower()
    if not s:
        return ""
    if s.startswith("{") and s.endswith("}"):
        s = s[1:-1].strip()
    if s.startswith("guid'") and s.endswith("'"):
        s = s[len("guid'") : -1].strip()
    return s


def export_production_orders_xlsx(db: Session) -> Dict[str, Any]:
    """
    Экспорт активных заказов на производство в XLSX (base64).

    Формат Excel:
    - Один лист
    - Группировка по заказам: заказ = подзаголовок (выделен цветом), детали = строки данных
    - Колонки: Номенклатура, Артикул, Характеристика, ЕИ, Заказано, Выполнено, Осталось
    
    Структура:
    | Заказ №XXX от YYYY-MM-DD | Состояние |
    | Номенклатура | Артикул | Характеристика | ЕИ | Заказано | Выполнено | Осталось |
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except Exception as e:
        raise RuntimeError(f"openpyxl not available: {e}")

    # Загружаем все активные (не удалённые) заказы с продукцией и номенклатурой
    orders = (
        db.query(ProductionOrder)
        .filter(ProductionOrder.deletion_mark == False)
        .options(
            joinedload(ProductionOrder.products)
            .joinedload(ProductionProduct.item)
        )
        .order_by(ProductionOrder.order_date, ProductionOrder.order_number)
        .all()
    )

    # Дополнительный фильтр "не завершён": в 1С/интеграции иногда появляются записи,
    # которые прошли по удалённости, но находятся в состоянии "Завершен".
    DONE_STATE_KEY = _norm_guid("ad28565a-991b-11eb-e39a-fa163e61326a")
    orders = [
        o for o in orders
        if _norm_guid(getattr(o, "order_state_key", None)) != DONE_STATE_KEY
    ]

    # Заголовки колонок для деталей
    headers = [
        "Номенклатура",
        "Артикул",
        "Характеристика",
        "ЕИ",
        "Заказано",
        "Выполнено",
        "Осталось",
    ]

    # Создаем Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ЗаказыНаПроизводство"

    # Стили
    header_font = Font(bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Стили для подзаголовка заказа
    order_header_font = Font(bold=True, size=11, color="FFFFFF")
    order_header_alignment = Alignment(horizontal="left", vertical="center")
    order_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Стили для чередования строк (zebra striping)
    light_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    
    # Границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Добавляем заголовки колонок
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Заполняем данными с группировкой по заказам
    row_num = 2
    total_detail_rows = 0

    for order_idx, order in enumerate(orders):
        # Состояние заказа (текстовое представление)
        state_display = "Активен"
        if order.order_state_key:
            state_display = order.order_state_name or f"ID: {order.order_state_key[:8]}..."

        # Дата заказа (форматируем)
        order_date_str = ""
        if order.order_date:
            if isinstance(order.order_date, datetime):
                order_date_str = order.order_date.strftime("%Y-%m-%d")
            else:
                order_date_str = str(order.order_date)

        # Строка-подзаголовок заказа
        order_header = f"Заказ №{order.order_number} от {order_date_str}"
        
        # Объединяем ячейки для подзаголовка (на всю ширину таблицы)
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        
        order_cell = ws.cell(row=row_num, column=1, value=order_header)
        order_cell.font = order_header_font
        order_cell.alignment = order_header_alignment
        order_cell.fill = order_header_fill
        order_cell.border = thin_border
        
        row_num += 1

        # Строки продукции заказа (сортируем по line_number)
        products = sorted(order.products, key=lambda p: p.line_number or 0)

        for product_idx, product in enumerate(products):
            item = product.item
            item_name = getattr(item, "item_name", "") or ""
            item_article = getattr(item, "item_article", "") or ""
            item_unit = getattr(item, "unit_name", "") or ""

            # Характеристика (если есть)
            characteristic = ""
            if product.characteristic_ref1c:
                characteristic = f"ID: {product.characteristic_ref1c[:8]}..."

            # Количество с учётом новых полей
            ordered_qty = float(product.quantity) if product.quantity else 0.0
            produced_qty = float(product.produced_qty) if hasattr(product, 'produced_qty') and product.produced_qty is not None else 0.0
            # Вычисляем remaining_qty явно, если поле не загружено или None
            if hasattr(product, 'remaining_qty') and product.remaining_qty is not None:
                remaining_qty = float(product.remaining_qty)
            else:
                remaining_qty = ordered_qty - produced_qty

            row_data = [
                item_name,
                item_article,
                characteristic,
                item_unit,
                ordered_qty,
                produced_qty,
                remaining_qty,
            ]

            # Чередование цветов для строк (zebra striping)
            cell_fill = light_fill if (total_detail_rows % 2 == 0) else None

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if cell_fill:
                    cell.fill = cell_fill

            row_num += 1
            total_detail_rows += 1

        # Добавляем пустую строку между заказами для визуального разделения
        row_num += 1

    # Авто-ширина колонок
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col_num).column_letter
        for row in range(1, min(row_num, 200)):  # Проверяем первые 200 строк для скорости
            try:
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем макс. ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    # Сохраняем в BytesIO
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    # Кодируем в base64
    b64 = base64.b64encode(bio.read()).decode("utf-8")

    # Формируем имя файла с датой
    filename = f"production_orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return {
        "status": "ok",
        "format": "xlsx",
        "data_base64": b64,
        "filename": filename,
        "total_rows": total_detail_rows,
        "orders_count": len(orders),
    }
